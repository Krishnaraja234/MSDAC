"""
COVER SHEET generation - CONFIRMED: pick & place the reference
COVER_SHEET.dxf content directly at (0,0), no translation. Always
exactly ONE sheet (no SPARE SHEETS count for this row in FIELD PG.NO,
unlike STATION LAYOUT/FLOOR PLAN). The reference file's own MTEXT
placeholders ({HUT NAME}, {SECTION1-SECTION2}, {AUTOSECTION - SIGNAL},
{DOUBLE LINE}, {COVER SHEET}) are left as-is for now - no confirmed
Excel-driven substitution values yet.
"""

import ezdxf
from ezdxf.addons import importer

from signal_core import insert_border_title, resolve_template_path

TITLE_TEXT = "COVER SHEET"


def get_cover_sheet_number(xlsx_path: str) -> str:
    """Reads COVER SHEET's sheet number (e.g. '000') from FIELD PG.NO."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(row=r, column=2).value or "").strip().upper()
        if label == "COVER SHEET":
            value = ws.cell(row=r, column=3).value
            if value is None or str(value).strip() == "":
                raise ValueError("FIELD PG.NO has no Sheet Number for the COVER SHEET row")
            return str(value).strip()
    raise ValueError("Could not find a 'COVER SHEET' row in FIELD PG.NO")


def get_next_sheet_number_after(xlsx_path: str, current_label: str) -> str:
    """Reads the SHEET NUMBER of whichever row comes immediately after
    `current_label` in FIELD PG.NO - e.g. after 'COVER SHEET' comes
    'INDEX' (sheet '00A'). CONFIRMED: this is NOT a simple '+1' on the
    current sheet number, since alphanumeric sequences (000 -> 00A ->
    00B...) don't increment that way - it has to be read from whatever
    FIELD PG.NO actually says comes next."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    found_current = False
    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(row=r, column=2).value or "").strip().upper()
        if found_current:
            value = ws.cell(row=r, column=3).value
            if value is not None and str(value).strip() != "":
                return str(value).strip()
            continue  # skip blank rows, keep looking for the next real one
        if label == current_label.strip().upper():
            found_current = True
    raise ValueError(f"Could not find a row after '{current_label}' in FIELD PG.NO")


def generate_cover_sheet(xlsx_path: str, border_template_path: str):
    """
    Returns (list of (out_name, doc, sht_number_str), next_sheet_number),
    same shape as every other generate_all_* function - always exactly
    ONE sheet.
    """
    sht = get_cover_sheet_number(xlsx_path)
    next_sht = get_next_sheet_number_after(xlsx_path, "COVER SHEET")

    source_doc = ezdxf.readfile(resolve_template_path("COVER_SHEET.dxf"))
    doc = ezdxf.new()
    imp = importer.Importer(source_doc, doc)
    imp.import_modelspace()
    imp.finalize()
    # CONFIRMED: pick & place at (0,0) - no translation needed, the
    # reference content is already at its correct final position.

    insert_border_title(
        doc=doc,
        border_template_path=border_template_path,
        template_name="COVER_SHEET",
        sig_name="", ahead_sig="", loc="",
        sht_number=str(sht),
        cont_number=str(next_sht),
        title_text_override=TITLE_TEXT,
    )

    out_name = f"COVER_SHEET_SHT{sht}.dxf"
    return [(out_name, doc, sht)], next_sht
