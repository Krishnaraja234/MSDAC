"""
MSDAC Communication Circuit Generator - core assembly logic.

Reads the COMMUNICATION sheet's row-groups (each ~18 rows: a sub-header
row with hut names, a column-header row ('I/O BIT'/'INPUT'/id/'OUTPUT'),
and 16 bit-data rows). Each group has a LEFT panel (columns B/C/D) and a
RIGHT panel (columns F/G/H), each identified by a UFSBI id like '1A'.

CONFIRMED mechanism:
  - Each panel (e.g. '1A') generates 4 SEPARATE SHEETS for MAIN, one per
    batch of 4 bits (1-4, 5-8, 9-12, 13-16). Each sheet is a FRESH,
    UNMODIFIED-POSITION copy of UFSBI_MAIN.dxf ("pick 0,0 place 0,0" -
    no repositioning, since each batch is its own separate file).
  - REDUN mirrors the SAME bits/relay identities (same INPUT column) but
    uses K4's contact groups instead of J4's, and has no output coil -
    it only wires the input side.
  - Sheet order: all 4 MAIN sheets for a panel, then all 4 REDUN sheets
    for that same panel, then move to the next panel.
  - S_NAME/R_NAME split on the FIRST SPACE (e.g. 'MCN8 HZR' -> 'MCN8'/'HZR').
  - A/F/B/C values come from FIELD PG.NO!J4 (MAIN) / K4 (REDUN): 4 groups
    per file (e.g. 'A1,A2/A3,A4/A5,A6/A7,A8'), FIXED and reused for every
    relay in every sheet (not incrementing).
  - LOC1/LOC2 placeholders <- the row-group's sub-header hut names
    (e.g. A2 and D2 for the LEFT panel).
  - 'UFSBI-XX' placeholder <- 'UFSBI-' + the panel's id cell (e.g. C3).
  - REDUN's title id swaps the MAIN panel's trailing letter A->B (e.g.
    panel '1A' -> REDUN title uses '1B').

NOTE: the L-X sequential label (seen in UFSBI_REDUN.dxf) is not yet
implemented - deferred per explicit instruction, revisit later.
"""

import ezdxf
from ezdxf.addons import importer
import os
import re

from relay_rack_core import split_bit_name, read_name2_whitelist

from signal_core import (
    TEMPLATES_DIR,
    resolve_template_path,
    _safe_str,
    insert_border_title as _insert_border_title_generic,
)

BITS_PER_SHEET = 4
TOTAL_BITS = 16


def _split_first_space(text: str):
    """
    'MCN8 HZR' -> ('MCN8', 'HZR'), splitting on the FIRST space only.
    Single-word text (e.g. 'SPARE', 'EB') goes into R_NAME, not S_NAME.
    """
    parts = str(text).strip().split(" ", 1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    return "", parts[0].strip()


def _parse_contact_groups(value: str):
    """'A1,A2/A3,A4/A5,A6/A7,A8' -> [['A1','A2'], ['A3','A4'], ['A5','A6'], ['A7','A8']]"""
    groups = []
    for chunk in str(value).split("/"):
        chunk = chunk.strip()
        if not chunk:
            continue
        values = [v.strip() for v in chunk.split(",")]
        groups.append(values)
    return groups


def _afc_for_subvalue(subvalue: str):
    """'A1' -> (letter='A', number='1')."""
    m = re.match(r"^([A-Za-z]+)(\d+)$", subvalue)
    if not m:
        raise ValueError(f"Could not parse contact sub-value {subvalue!r} (expected e.g. 'A1')")
    return m.group(1), m.group(2)


def get_contact_groups(xlsx_path: str):
    """Reads FIELD PG.NO!J4 (MAIN) and K4 (REDUN) contact group definitions."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    j4 = ws["J4"].value
    k4 = ws["K4"].value
    if not j4 or not k4:
        raise ValueError("FIELD PG.NO!J4 (MAIN) and K4 (REDUN) must both be filled in")
    return _parse_contact_groups(j4), _parse_contact_groups(k4)


def read_communication_panels(xlsx_path: str):
    """
    Dynamically finds every row-group in the COMMUNICATION sheet and
    returns a list of panel dicts, one per LEFT/RIGHT panel found:
        {id, loc1, loc2, own_hut_name, input_col, output_col, bits: [(input_text, output_text), ...]}

    CONFIRMED (replaces the earlier per-panel "group_location" design,
    which required a dedicated location row above EVERY panel-group's
    own subheader - unreliable, since not every panel-group actually has
    one): 'own_hut_name' is read ONCE from row 1 (the sheet's own merged
    title cell, e.g. "MSDAC HUT-1") and applies to EVERY panel on this
    sheet - it represents which hut THIS WHOLE GENERATION RUN belongs
    to. Contact Analysis compares this fixed value against each panel's
    OWN LOC1 (to decide whether CONTACTS count) and LOC2 (to decide
    whether the COIL counts) - NOT the relay's own rack identifier from
    Relay Rack, which is often just a generic label like "R1" with no
    hut name of its own.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["COMMUNICATION"]

    own_hut_name = str(ws.cell(row=1, column=1).value or "").strip()

    panels = []
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or "").strip() == "I/O\nBIT":
            colheader_row = r
            subheader_row = r - 1
            data_start = r + 1
            data_end = r + TOTAL_BITS  # 16 bit rows

            for side, (loc1_col, id_col, loc2_col, input_col, output_col) in {
                "LEFT": (1, 3, 4, 2, 4),
                "RIGHT": (5, 7, 8, 6, 8),
            }.items():
                panel_id = ws.cell(row=colheader_row, column=id_col).value
                if panel_id is None or str(panel_id).strip() == "":
                    continue
                loc1 = ws.cell(row=subheader_row, column=loc1_col).value or ""
                if not str(loc1).strip():
                    # CONFIRMED FIX: some workbooks have LOC1 shifted one
                    # column to the right of where it's normally expected
                    # (e.g. from a merged summary row above disturbing the
                    # sub-header row's alignment) - fall back to the
                    # adjacent column rather than silently using a blank
                    # LOC1, which would make every contact/coil location
                    # check fail for this entire panel.
                    loc1 = ws.cell(row=subheader_row, column=loc1_col + 1).value or ""
                loc2 = ws.cell(row=subheader_row, column=loc2_col).value or ""

                bits = []
                for br in range(data_start, data_end + 1):
                    input_text = ws.cell(row=br, column=input_col).value or ""
                    output_text = ws.cell(row=br, column=output_col).value or ""
                    bits.append((str(input_text).strip(), str(output_text).strip()))

                panels.append({
                    "id": str(panel_id).strip(),
                    "loc1": str(loc1).strip(),
                    "loc2": str(loc2).strip(),
                    "own_hut_name": own_hut_name,
                    "bits": bits,
                })

    return panels


def _substitute_placeholders(text: str, loc1: str, loc2: str, ufsbi_id: str) -> str:
    if text is None:
        return text
    new_text = text
    new_text = new_text.replace("LOC1", _safe_str(loc1))
    new_text = new_text.replace("LOC2", _safe_str(loc2))
    new_text = new_text.replace("UFSBI-XX", f"UFSBI-{_safe_str(ufsbi_id)}")
    return new_text


def _redun_id(main_id: str) -> str:
    """'1A' -> '1B' (swap trailing letter A->B, keep the numeric prefix)."""
    m = re.match(r"^(\d*)([A-Za-z])$", main_id)
    if m:
        prefix, letter = m.group(1), m.group(2)
        return f"{prefix}{chr(ord(letter) + 1)}"
    return main_id


def _import_whole_file_no_rename(doc, source_filename: str):
    """
    Like _import_whole_file, but avoids block-name-renaming when the same
    block definition gets re-imported multiple times across sheets/calls.
    Imports each INSERT's block DEFINITION only once per document (safe
    check via doc.blocks), then creates a fresh blockref preserving the
    original name. Raw entities (MTEXT, etc.) are copied fresh each time
    (no renaming risk for non-block entities). Always at (0,0) - no
    translation needed for this module's usage pattern.
    """
    source_doc = ezdxf.readfile(resolve_template_path(source_filename))
    source_msp = source_doc.modelspace()
    target_msp = doc.modelspace()

    new_entities = []
    for e in source_msp:
        if e.dxftype() == "INSERT":
            block_name = e.dxf.name
            if block_name not in doc.blocks:
                imp = importer.Importer(source_doc, doc)
                imp.import_block(block_name)
                imp.finalize()
            ins = target_msp.add_blockref(block_name, (e.dxf.insert.x, e.dxf.insert.y))
            ins.dxf.xscale = e.dxf.xscale
            ins.dxf.yscale = e.dxf.yscale
            ins.dxf.zscale = e.dxf.zscale
            ins.dxf.rotation = e.dxf.rotation
            for att in e.attribs:
                dxfattribs = {
                    "height": att.dxf.height,
                    "style": att.dxf.style,
                    "rotation": att.dxf.rotation,
                    "flags": att.dxf.flags,
                }
                ins.add_attrib(att.dxf.tag, att.dxf.text, att.dxf.insert, dxfattribs)
            new_entities.append(ins)
        else:
            existing_ids = set(id(x) for x in target_msp)
            imp = importer.Importer(source_doc, doc)
            imp.import_entity(e, target_msp)
            imp.finalize()
            new_entities.extend(x for x in target_msp if id(x) not in existing_ids)

    return new_entities


def _import_whole_file(doc, source_filename: str):
    """Copy the ENTIRE source file's modelspace content into doc, unchanged (pick 0,0 -> place 0,0)."""
    source_doc = ezdxf.readfile(resolve_template_path(source_filename))
    target_msp = doc.modelspace()
    existing_ids = set(id(e) for e in target_msp)

    imp = importer.Importer(source_doc, doc)
    imp.import_modelspace()
    imp.finalize()

    return [e for e in target_msp if id(e) not in existing_ids]


def _inserts_named(entities, name):
    return [e for e in entities if e.dxftype() == "INSERT" and e.dxf.name == name]


def generate_one_communication_sheet(
    panel: dict,
    bit_batch: list,
    contact_groups: list,
    is_main: bool,
    sht_number: str,
    cont_number: str,
    border_template_path: str,
    first_bit_number: int = 1,
    corresponding_redun_sht: str = None,
    corresponding_main_sht: str = None,
    sheet_index: int = 0,
    total_sheets: int = 1,
    real_bit_count: int = TOTAL_BITS,
    previous_sht: str = None,
    name2_whitelist: list = None,
    is_true_last: bool = True,
    next_side_first_sht: str = None,
):
    """
    Builds ONE sheet covering 4 bits, for either MAIN or REDUN.
    bit_batch: list of up to 4 (input_text, output_text) tuples.
    first_bit_number: the actual bit number (1, 5, 9, or 13) that this
        batch starts at - used to fill UBSBI_IN_BIT/UBSBIT_OUT_BIT.
    sheet_index: 0-based position of this sheet within the panel's batch
        sequence (0=first sheet -> gets FUSE; last -> gets CONT_UP+CONT
        or TERMINATION depending on is_true_last; everything else ->
        gets CONT_UP+CONT).
    is_true_last: CONFIRMED - TERMINATION only appears on REDUN's actual
        last sheet. MAIN's own last sheet instead gets CONT pointing
        forward to REDUN's first sheet (next_side_first_sht), continuing
        the chain rather than terminating it.
    real_bit_count: how many of this panel's 16 bits are NOT literally
        "SPARE" text - determines where the L-X/OUT_TERMINAL sequence
        switches from real numbers to "SPARE".
    Returns (output_filename, ezdxf.Drawing).
    """
    source_filename = "UFSBI_MAIN.dxf" if is_main else "UFSBI_REDUN.dxf"
    doc = ezdxf.readfile(resolve_template_path(source_filename))
    msp = doc.modelspace()

    # CONFIRMED: use NATIVE file order (as entities appear in the DXF),
    # NOT sorted by Y position - the template's own internal entity order
    # does not follow a simple top-to-bottom sequence.
    coil_inserts = _inserts_named(list(msp), "Relay_Coil_ACI") if is_main else []
    front_contacts = _inserts_named(list(msp), "Front_Contact")
    back_contacts = _inserts_named(list(msp), "Back_Contact")

    # Each relay slot uses 2 Front + 2 Back contacts - group them in
    # consecutive pairs matching the number of bits in this batch.
    num_slots = len(bit_batch)
    for slot_index in range(num_slots):
        input_text, output_text = bit_batch[slot_index]
        s_name, r_name = split_bit_name(input_text, name2_whitelist) if input_text else ("", "")

        slot_fronts = front_contacts[slot_index * 2: slot_index * 2 + 2]
        slot_backs = back_contacts[slot_index * 2: slot_index * 2 + 2]

        # Front1, Front2 <- group[0], group[1]; Back1, Back2 <- group[2], group[3]
        if len(contact_groups) < 4:
            raise ValueError(f"Contact groups need at least 4 entries, got {len(contact_groups)}: {contact_groups!r}")

        # CONFIRMED FIX: a SPARE bit isn't a real relay and is never
        # updated in Contact Analysis, so it must not get any contact
        # position (C/F/A/B) assigned either - only the "SPARE" label
        # itself (via R_NAME) should show. Leave the position fields
        # blank instead of filling them from the fixed contact_groups
        # layout.
        is_spare_input = (input_text or "").strip().upper() == "SPARE"

        for e, subval in zip(slot_fronts, [contact_groups[0][0], contact_groups[1][0]]):
            letter, num = _afc_for_subvalue(subval)
            for att in e.attribs:
                if att.dxf.tag == "S_NAME":
                    att.dxf.text = _safe_str(s_name)
                elif att.dxf.tag == "R_NAME":
                    att.dxf.text = _safe_str(r_name)
                elif att.dxf.tag == "C":
                    att.dxf.text = "" if is_spare_input else letter
                elif att.dxf.tag == "F":
                    att.dxf.text = "" if is_spare_input else num
                elif att.dxf.tag == "A":
                    att.dxf.text = "" if is_spare_input else str(int(num) + 1)
                elif att.dxf.tag == "R(POS)":
                    att.dxf.text = ""  # CONFIRMED: blank unless the Excel relay-position is ticked

        for e, subval in zip(slot_backs, [contact_groups[2][0], contact_groups[3][0]]):
            letter, num = _afc_for_subvalue(subval)
            for att in e.attribs:
                if att.dxf.tag == "S_NAME":
                    att.dxf.text = _safe_str(s_name)
                elif att.dxf.tag == "R_NAME":
                    att.dxf.text = _safe_str(r_name)
                elif att.dxf.tag == "C":
                    att.dxf.text = "" if is_spare_input else letter
                elif att.dxf.tag == "A":
                    att.dxf.text = "" if is_spare_input else num
                elif att.dxf.tag == "B":
                    att.dxf.text = "" if is_spare_input else str(int(num) + 1)
                elif att.dxf.tag == "R(POS)":
                    att.dxf.text = ""  # CONFIRMED: blank unless the Excel relay-position is ticked

        if is_main and slot_index < len(coil_inserts):
            out_s_name, out_r_name = split_bit_name(output_text, name2_whitelist) if output_text else ("", "")
            is_spare_output = (output_text or "").strip().upper() == "SPARE"
            for att in coil_inserts[slot_index].attribs:
                if att.dxf.tag == "S_NAME":
                    att.dxf.text = _safe_str(out_s_name)
                elif att.dxf.tag == "R_NAME":
                    att.dxf.text = _safe_str(out_r_name)
                elif att.dxf.tag == "R(POS)":
                    att.dxf.text = ""  # CONFIRMED: blank unless the Excel relay-position is ticked
                elif att.dxf.tag in ("L+", "L-"):
                    # CONFIRMED FIX: previously never touched at all, so
                    # it always showed the raw template default ("R2"/
                    # "R1") regardless of real vs SPARE. Blank it for
                    # SPARE (no relay there, never tracked in CA);
                    # restore the normal R2/R1 label for a real bit.
                    if is_spare_output:
                        att.dxf.text = ""
                    else:
                        att.dxf.text = "R2" if att.dxf.tag == "L+" else "R1"

    # UBSBI_IN_BIT / UBSBIT_OUT_BIT: fill BIT1-4 with the actual bit
    # numbers this sheet covers (e.g. sheet 2 -> 5,6,7,8).
    for bit_block_name in ("UBSBI_IN_BIT", "UBSBIT_OUT_BIT"):
        for e in _inserts_named(list(msp), bit_block_name):
            for att in e.attribs:
                if att.dxf.tag in ("BIT1", "BIT2", "BIT3", "BIT4"):
                    offset = int(att.dxf.tag[-1]) - 1
                    bit_num = first_bit_number + offset
                    att.dxf.text = str(bit_num) if offset < num_slots else ""

    # CONFIRMED: FUSE's own L-X ('L-1') is a SEPARATE/ADDITIONAL position -
    # the L-X block's own 8-per-sheet sequence starts at L-2, not L-1.
    # highest_real_lx: with 4 real bits (1 sheet), this should be L-9
    # (1 for FUSE + 4 bits x 2 = 9) - CONFIRMED via worked example.
    highest_real_lx = 1 + real_bit_count * 2

    # OUT_TERMINAL (MAIN only): plain numbers, ALWAYS numeric (no FUSE offset - separate sequence).
    for e in _inserts_named(list(msp), "OUT_TERMINAL"):
        for att in e.attribs:
            if att.dxf.tag.startswith("BIT") and att.dxf.tag[3:].isdigit():
                slot_in_sheet = int(att.dxf.tag[3:])  # 1-8
                position = sheet_index * 8 + slot_in_sheet
                att.dxf.text = str(position)

    # L-X block (8 attributes L-1..L-8 per instance): sequential numbering
    # WITH "L-" prefix, offset by +1 to account for FUSE's own L-1.
    for e in _inserts_named(list(msp), "L-X"):
        for att in e.attribs:
            if att.dxf.tag.startswith("L-") and att.dxf.tag[2:].isdigit():
                slot_in_sheet = int(att.dxf.tag[2:])  # 1-8
                position = sheet_index * 8 + slot_in_sheet + 1
                att.dxf.text = f"L-{position}"

    def _letter_for(n):
        """0->'A', 1->'B', ..., 25->'Z', 26->'A1', 27->'B1', ... (CONFIRMED wrap rule)."""
        cycle, remainder = divmod(n, 26)
        letter = chr(ord("A") + remainder)
        return f"{letter}{cycle}" if cycle > 0 else letter

    is_first = sheet_index == 0
    is_last = sheet_index == total_sheets - 1
    gets_termination = is_last  # CONFIRMED (reverted): both MAIN's own last sheet AND REDUN's own last sheet each get their own TERMINATION

    # FUSE - only on the FIRST sheet, pick(0,0)->place(0,0).
    if is_first:
        fuse_pasted = _import_whole_file_no_rename(doc, "UFSBI_FUSE.dxf")
        for e in fuse_pasted:
            if e.dxftype() == "INSERT":
                for att in e.attribs:
                    if att.dxf.tag == "L-X":
                        att.dxf.text = "L-1"
                    elif att.dxf.tag == "VOLT":
                        # CONFIRMED: VOLT's rotation is baked into the source
                        # template as 180 degrees (upside-down/"mirrored"
                        # appearance) - reset it to render right-side-up.
                        att.dxf.rotation = 0

    # CONT_UP - on every NON-FIRST sheet: CONT attribute = PREVIOUS sheet's
    # number, A attribute = PREVIOUS sheet's own letter.
    if not is_first and previous_sht:
        pasted = _import_whole_file_no_rename(doc, "UFSBI_CONT_UP.dxf")
        for e in pasted:
            if e.dxftype() == "INSERT":
                for att in e.attribs:
                    if att.dxf.tag == "CONT":
                        att.dxf.text = _safe_str(previous_sht)
                    elif att.dxf.tag == "A":
                        att.dxf.text = _letter_for(sheet_index - 1)

    # CONT - on every sheet EXCEPT the one getting TERMINATION. CONFIRMED:
    # MAIN's own last sheet gets CONT pointing forward to REDUN's FIRST
    # sheet (continuing the chain) rather than TERMINATION - only REDUN's
    # true last sheet actually terminates.
    if not gets_termination:
        pasted = _import_whole_file_no_rename(doc, "UFSBI_CONT.dxf")
        cont_value = cont_number
        for e in pasted:
            if e.dxftype() == "INSERT":
                for att in e.attribs:
                    if att.dxf.tag == "CONT":
                        att.dxf.text = _safe_str(cont_value)
                    elif att.dxf.tag == "A":
                        att.dxf.text = _letter_for(sheet_index)

    # TERMINATION - only on REDUN's true final sheet, REPLACING CONT there.
    if gets_termination:
        term_pasted = _import_whole_file_no_rename(doc, "UFSBI_TERMINATION.dxf")
        has_spare = highest_real_lx < 33  # CONFIRMED: if all 4 sheets are fully real, no spare summary needed
        for e in term_pasted:
            if e.dxftype() == "INSERT":
                for att in e.attribs:
                    if att.dxf.tag == "L-X":
                        att.dxf.text = "L-34"  # CONFIRMED: always fixed, regardless of data
            if e.dxftype() == "MTEXT" and e.text and "L-XX TO XX" in e.text:
                if has_spare:
                    first_spare = highest_real_lx + 1
                    e.text = e.text.replace("L-XX TO XX", f"L-{first_spare} TO 33")
                else:
                    doc.modelspace().delete_entity(e)  # CONFIRMED: not required when nothing is spare

    # SH.NO:XXX (MAIN only) -> the CORRESPONDING REDUN sheet's number
    # (same bit batch, generated right after all MAIN sheets for this panel).
    if is_main and corresponding_redun_sht:
        for e in msp:
            if e.dxftype() == "MTEXT" and e.text and "SH.NO:XXX" in e.text:
                e.text = e.text.replace("SH.NO:XXX", f"SH.NO:{corresponding_redun_sht}")

    # SH.NO:XXX (REDUN only) -> mirrors back to the CORRESPONDING MAIN
    # sheet's number (same bit batch).
    if not is_main and corresponding_main_sht:
        for e in msp:
            if e.dxftype() == "MTEXT" and e.text and "SH.NO:XXX" in e.text:
                e.text = e.text.replace("SH.NO:XXX", f"SH.NO:{corresponding_main_sht}")

    # LOC1/LOC2/UFSBI-XX placeholder substitution across TEXT/MTEXT.
    ufsbi_id = panel["id"] if is_main else _redun_id(panel["id"])
    for e in msp:
        if e.dxftype() == "TEXT":
            e.dxf.text = _substitute_placeholders(e.dxf.text, panel["loc1"], panel["loc2"], ufsbi_id)
        elif e.dxftype() == "MTEXT":
            e.text = _substitute_placeholders(e.text, panel["loc1"], panel["loc2"], ufsbi_id)

    title_text = f"UFSBI {ufsbi_id}- {panel['loc1']} - {panel['loc2']} CIRCUITS"
    _insert_border_title_generic(
        doc=doc,
        border_template_path=border_template_path,
        template_name="COMMUNICATION",
        sig_name=None,
        ahead_sig=None,
        loc=None,
        sht_number=sht_number,
        cont_number=cont_number,
        title_text_override=title_text,
    )

    # CONFIRMED: carries the row-group's own "home hut" label through to
    # the generated sheet as a hidden extra attribute (no matching ATTDEF
    # needed - just a data-carrying tag), so Contact Analysis can read it
    # back later without needing to re-open the Excel workbook. Used to
    # decide whether this panel's contacts (LOC1 match) and/or coil
    # (LOC2 match) actually count.
    for e in msp:
        if e.dxftype() == "INSERT" and e.dxf.name in ("TITLE", "TITLEBLOCK"):
            e.add_attrib("GROUP_LOC", panel["own_hut_name"], (0, 0), dxfattribs={"flags": 1})
            # CONFIRMED FIX: LOC1/LOC2 are now embedded directly as their
            # own hidden attributes, rather than making the reader
            # regex-parse them back out of the human-readable TITLE text.
            # That parsing was fragile - a hut name that itself contains
            # a hyphen (e.g. "MSDAC HUT-1") broke the LOC1/LOC2 split,
            # since the regex couldn't tell that internal hyphen apart
            # from the one separating LOC1 and LOC2.
            e.add_attrib("LOC1", panel["loc1"], (0, 0), dxfattribs={"flags": 1})
            e.add_attrib("LOC2", panel["loc2"], (0, 0), dxfattribs={"flags": 1})
            break

    out_name = f"UFSBI_{ufsbi_id}_SHT{sht_number}.dxf"
    return out_name, doc


def get_communication_start_sheet_number(xlsx_path: str) -> int:
    """Reads COMMUNICATION's starting sheet number from FIELD PG.NO."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(row=r, column=2).value or "").strip().upper()
        if label == "COMMUNICATION":
            value = ws.cell(row=r, column=3).value
            if value is None or str(value).strip() == "":
                raise ValueError("FIELD PG.NO has no Sheet Number for the COMMUNICATION row")
            return int(str(value).strip())
    raise ValueError("Could not find a 'COMMUNICATION' row in FIELD PG.NO")


def get_next_circuit_after_communication(xlsx_path: str) -> str:
    """Reads the starting sheet number of whatever circuit follows COMMUNICATION in FIELD PG.NO."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(row=r, column=2).value or "").strip().upper()
        if label == "COMMUNICATION":
            value = ws.cell(row=r + 1, column=3).value
            return str(value).strip() if value is not None else ""
    return ""


def generate_all_communication_sheets(xlsx_path: str, start_sheet_number: int, border_template_path: str):
    """
    Generates every sheet for every panel found in the COMMUNICATION
    sheet: MAIN sheets then REDUN sheets per panel, sequential sheet
    numbers throughout. Each MAIN sheet's 'SH.NO:XXX' placeholders get
    filled with its CORRESPONDING REDUN sheet's number.

    Returns (list of (output_filename, ezdxf.Drawing, sht_number_str), next_sheet_number)
    """
    panels = read_communication_panels(xlsx_path)
    if not panels:
        raise ValueError("No UFSBI panels found in COMMUNICATION sheet")

    main_groups, redun_groups = get_contact_groups(xlsx_path)
    name2_whitelist = read_name2_whitelist(xlsx_path)

    # PASS 1: compute the sheet-number layout for every panel first, so
    # each sheet's TITLEBLOCK CONT can correctly skip over reserved-but-
    # unused slots and correctly chain across panel boundaries.
    layout = []
    sheet_num = start_sheet_number
    for panel in panels:
        bits = panel["bits"]
        batches = [bits[i:i + BITS_PER_SHEET] for i in range(0, len(bits), BITS_PER_SHEET)]

        # CONFIRMED: MAIN (and REDUN, identically) skips trailing batches
        # that are ENTIRELY spare (no real bit in them) - the sheet-number
        # SLOT is still reserved for all 4 batches, so numbering stays
        # predictable. REDUN uses the SAME batch_count as MAIN, since it
        # covers the same underlying bits.
        main_batch_count = 0
        for b in batches:
            if any(input_text and input_text.strip().upper() != "SPARE" for input_text, _ in b):
                main_batch_count = batches.index(b) + 1
        if main_batch_count == 0:
            # CONFIRMED: a panel with NO real bits at all (every bit is
            # SPARE) is skipped entirely - no MAIN/REDUN sheets generated,
            # no sheet numbers consumed for it. Doesn't affect sheet
            # numbering for any OTHER panel, since sheet_num only advances
            # for panels that actually produce sheets.
            continue

        generated_bit_count = main_batch_count * BITS_PER_SHEET
        main_shts = [f"{sheet_num + i:03d}" for i in range(len(batches))]
        redun_shts = [f"{sheet_num + len(batches) + i:03d}" for i in range(len(batches))]

        layout.append({
            "panel": panel, "batches": batches, "main_batch_count": main_batch_count,
            "generated_bit_count": generated_bit_count, "main_shts": main_shts, "redun_shts": redun_shts,
        })
        sheet_num += len(batches) * 2

    # PASS 2: generate content, now knowing each sheet's correct "next
    # sheet" (skip-aware within a side, cross-panel-aware at panel boundaries).
    results = []
    for panel_index, entry in enumerate(layout):
        panel = entry["panel"]
        batches = entry["batches"]
        main_batch_count = entry["main_batch_count"]
        generated_bit_count = entry["generated_bit_count"]
        main_shts = entry["main_shts"]
        redun_shts = entry["redun_shts"]

        next_panel_first_sht = layout[panel_index + 1]["main_shts"][0] if panel_index + 1 < len(layout) else None

        for is_main, groups, shts in (
            (True, main_groups, main_shts),
            (False, redun_groups, redun_shts),
        ):
            for batch_index in range(main_batch_count):
                batch = batches[batch_index]
                sht = shts[batch_index]
                is_last_in_side = batch_index == main_batch_count - 1
                if not is_last_in_side:
                    cont = shts[batch_index + 1]  # next batch, same side - always sequential, no skip
                elif is_main:
                    cont = redun_shts[0]  # MAIN's own last sheet -> REDUN's actual first sheet
                else:
                    cont = next_panel_first_sht or f"{int(sht) + 1:03d}"  # REDUN's true last sheet -> next panel's first sheet (placeholder if this is the very last sheet overall - app.py overwrites it)
                first_bit_number = batch_index * BITS_PER_SHEET + 1
                corresponding_redun_sht = redun_shts[batch_index] if is_main else None
                corresponding_main_sht = main_shts[batch_index] if not is_main else None
                previous_sht = shts[batch_index - 1] if batch_index > 0 else None
                try:
                    out_name, doc = generate_one_communication_sheet(
                        panel, batch, groups, is_main, sht, cont, border_template_path,
                        first_bit_number, corresponding_redun_sht, corresponding_main_sht,
                        batch_index, main_batch_count, generated_bit_count, previous_sht,
                        name2_whitelist,
                    )
                except Exception as e:
                    raise ValueError(f"Panel {panel['id']!r} ({'MAIN' if is_main else 'REDUN'}): {e}") from e
                results.append((out_name, doc, sht))

    final_sheet_num = start_sheet_number + sum(len(entry["batches"]) * 2 for entry in layout)
    return results, final_sheet_num
