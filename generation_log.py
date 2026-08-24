"""
Generation logging and cross-circuit contact-uniqueness tracking.

CONFIRMED requirements:
  - Every error/warning during a generation run should be logged (not
    just shown transiently in job status) - a log file gets included in
    the output zip so the user can review everything after the fact.
  - Contact repetition: if the SAME relay's SAME contact (e.g. SKL 2's
    D1,D2) gets used more than once - whether on the same sheet or a
    different one, and regardless of which circuit type assigned it -
    that's a genuine physical conflict and must be flagged.
  - Missing/incomplete Excel input (a required cell empty, a row with
    no usable data, etc.) should also be captured in the log, not just
    silently skipped.
"""

import datetime


class GenerationLog:
    """Collects INFO/WARNING/ERROR entries during a generation run."""

    def __init__(self):
        self.entries = []  # list of (level, message)

    def info(self, message: str):
        self.entries.append(("INFO", message))

    def warning(self, message: str):
        self.entries.append(("WARNING", message))

    def error(self, message: str):
        self.entries.append(("ERROR", message))

    @property
    def has_warnings_or_errors(self) -> bool:
        return any(level in ("WARNING", "ERROR") for level, _ in self.entries)

    @property
    def has_errors(self) -> bool:
        return any(level == "ERROR" for level, _ in self.entries)

    def summary_counts(self):
        warnings = sum(1 for level, _ in self.entries if level == "WARNING")
        errors = sum(1 for level, _ in self.entries if level == "ERROR")
        return warnings, errors

    def render(self) -> str:
        """Renders the full log as plain text, for inclusion in the output zip."""
        lines = [
            "MSDAC IFC Generator - Generation Log",
            f"Generated: {datetime.datetime.now().isoformat(timespec='seconds')}",
            "=" * 60,
            "",
        ]
        if not self.entries:
            lines.append("No warnings or errors - everything generated cleanly.")
        else:
            warnings, errors = self.summary_counts()
            lines.append(f"{errors} error(s), {warnings} warning(s):")
            lines.append("")
            for level, message in self.entries:
                lines.append(f"[{level}] {message}")
        return "\n".join(lines)

    def save_excel(self, path: str):
        """Writes the log as an .xlsx file - one row per entry, with Level/Message columns."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Generation Log"
        ws.append(["Level", "Message"])
        for cell in ws[1]:
            cell.font = Font(bold=True)

        error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        if not self.entries:
            ws.append(["INFO", "No warnings or errors - everything generated cleanly."])
        else:
            for level, message in self.entries:
                ws.append([level, message])
                row = ws.max_row
                if level == "ERROR":
                    ws.cell(row=row, column=1).fill = error_fill
                elif level == "WARNING":
                    ws.cell(row=row, column=1).fill = warning_fill

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 120
        wb.save(path)


class ContactRegistry:
    """
    Tracks every (relay_name, contact_id) pair used across ALL circuit
    types in a single generation run, flagging repeats.

    contact_id should uniquely identify a physical contact position,
    e.g. 'D1' or 'D2' (not the whole pair 'D1,D2' - each half of a pair
    is its own physical terminal and could theoretically be reused
    independently, though in practice they're almost always used together).
    """

    def __init__(self, log: GenerationLog):
        self.log = log
        self._used = {}  # (relay_name, contact_id) -> list of (circuit_type, sheet_number)

    def register(self, relay_name: str, contact_id: str, circuit_type: str, sheet_number: str):
        if not relay_name or not contact_id:
            return
        # CONFIRMED FIX: Excel cells can hold a raw number (e.g. an
        # AR/AZR "own name" entered as literal 1 instead of text "1" or
        # "SKL 2") - openpyxl then hands back an int, and calling
        # .strip() on it directly crashed with "'int' object has no
        # attribute 'strip'". Convert to string first.
        key = (str(relay_name).strip().upper(), str(contact_id).strip().upper())
        if key in self._used:
            prior = self._used[key]
            prior_desc = ", ".join(f"{ct} sheet {sht}" for ct, sht in prior)
            # CONFIRMED: contact repetition is a CRITICAL error - it must
            # stop generation before any output is produced, not just warn.
            self.log.error(
                f"Contact repetition: relay {relay_name!r} contact {contact_id!r} "
                f"was already used in {prior_desc} - now also used in {circuit_type} sheet {sheet_number}"
            )
        self._used.setdefault(key, []).append((circuit_type, sheet_number))

    def register_pair(self, relay_name: str, letter: str, num1, num2, circuit_type: str, sheet_number: str):
        """Convenience: registers BOTH halves of a 'D1,D2'-style pair."""
        if num1 is not None:
            self.register(relay_name, f"{letter}{num1}", circuit_type, sheet_number)
        if num2 is not None:
            self.register(relay_name, f"{letter}{num2}", circuit_type, sheet_number)
