"""
STATION LAYOUT sheet generation - CONFIRMED: these are just blank
placeholder sheets (border/title-block only, no actual content or
table) - the number of sheets to create comes from FIELD PG.NO's
'SPARE SHEETS' column for the 'STATION LAYOUT' row (e.g. Sheet Number
'00D', SPARE SHEETS=2 -> generates sheets '00D' and '00E', titled
'STATION LAYOUT - 1' and 'STATION LAYOUT - 2').
"""

import ezdxf

from signal_core import insert_border_title

TITLE_TEXT_PREFIX = "STATION LAYOUT - "


def get_station_layout_config(xlsx_path: str):
    """Reads STATION LAYOUT's starting sheet number (e.g. '00D') and
    sheet count (from SPARE SHEETS) from FIELD PG.NO."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(row=r, column=2).value or "").strip().upper()
        if label == "STATION LAYOUT":
            sht_value = ws.cell(row=r, column=3).value
            count_value = ws.cell(row=r, column=4).value
            if sht_value is None or str(sht_value).strip() == "":
                raise ValueError("FIELD PG.NO has no Sheet Number for the STATION LAYOUT row")
            if count_value is None or str(count_value).strip() == "":
                raise ValueError("FIELD PG.NO has no SPARE SHEETS count for the STATION LAYOUT row")
            return str(sht_value).strip(), int(count_value)
    raise ValueError("Could not find a 'STATION LAYOUT' row in FIELD PG.NO")


def _increment_alpha_suffix(sht: str) -> str:
    """'00D' -> '00E' -> ... -> '00Z' -> '00AA' (same convention as
    index_core.py's own version - kept local to avoid a cross-module
    dependency for one small shared helper)."""
    m = 0
    while m < len(sht) and sht[len(sht) - 1 - m].isalpha():
        m += 1
    if m == 0:
        return str(int(sht) + 1)
    prefix, suffix = sht[:-m], sht[-m:]
    chars = list(suffix)
    i = len(chars) - 1
    while i >= 0:
        if chars[i] != "Z":
            chars[i] = chr(ord(chars[i]) + 1)
            break
        chars[i] = "A"
        i -= 1
    else:
        chars.insert(0, "A")
    return prefix + "".join(chars)


def generate_all_station_layout_sheets(xlsx_path: str, border_template_path: str):
    """
    Returns (list of (out_name, doc, sht_number_str), next_sheet_number),
    same shape as every other generate_all_* function - one blank sheet
    per SPARE SHEETS count, titled "STATION LAYOUT - N".
    """
    start_sht, count = get_station_layout_config(xlsx_path)

    results = []
    sht = start_sht
    for page_number in range(1, count + 1):
        next_sht = _increment_alpha_suffix(sht)
        doc = ezdxf.new()
        insert_border_title(
            doc=doc,
            border_template_path=border_template_path,
            template_name="STATION_LAYOUT",
            sig_name="", ahead_sig="", loc="",
            sht_number=str(sht),
            cont_number=str(next_sht),
            title_text_override=f"{TITLE_TEXT_PREFIX}{page_number}",
        )
        out_name = f"STATION_LAYOUT_SHT{sht}.dxf"
        results.append((out_name, doc, sht))
        sht = next_sht

    return results, sht
