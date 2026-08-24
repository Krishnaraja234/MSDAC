"""
MSDAC Relay Rack Generator - core assembly logic.

Reads the RELAY RACK sheet's grid (rack row letters A, B, C... down the
left edge; numbered rack positions 1-21 across the top) and stamps out
one RELAY_BLOCK.dxf copy per grid cell - including blank cells, which
get a "SPARE" placeholder block rather than being skipped.

Excel layout (CONFIRMED): each rack row-letter occupies exactly 2 Excel
rows - a "tag" row (e.g. row 3 for 'A') immediately followed by a "type"
row (row 4). This repeats consistently: B=rows5-6, C=rows7-8, D=rows9-10,
etc. The sheet ends at a row whose column A value is "END".

Per-cell attribute mapping (CONFIRMED):
  - Tag cell (e.g. "F54 TSPR") splits on the FIRST space:
        NAME1 = "F54", NAME2 = "TSPR"
  - Type cell directly below (e.g. "QN1,8F-8B") splits on comma:
        RELTYP = "QN1", CNTCNF = "8F-8B"
  - Blank tag cell -> NAME1/NAME2/RELTYP/CNTCNF all blank, NAME3="SPARE"

Positioning (CONFIRMED, measured directly from the reference file):
  - X_SPACING = 17.8198477209488 (ADDED per column, not subtracted)
  - Y_SPACING = 24.7188341 (SUBTRACTED per row, going down the rack)
  - First position (row A, column 1) lands at FIRST_POSITION_POINT
  - RELAY_BLOCK.dxf's own native pick point is (64.60997830802603, 142.2331889705842)

RELAY_RACK.dxf (the base/border) is inserted at (0,0) -> (0,0), i.e. no
translation at all - it's just the starting container document.
"""

import ezdxf
from ezdxf.addons import importer
import os

from signal_core import (
    TEMPLATES_DIR,
    resolve_template_path,
    _safe_str,
    insert_border_title as _insert_border_title_generic,
)

RELAY_BLOCK_NATIVE_PICK = (64.60997830802603, 142.2331889705842)


def read_relay_rack_position_config(xlsx_path: str) -> dict:
    """
    Reads position values from the 'RELAY_RACK_POSITIONS' sheet, so
    positions can be adjusted directly in Excel without touching code.

    Expected sheet layout (header row + one row per named value):
        Parameter                | X       | Y
        FIRST_POSITION_POINT     | 33.2163 | 268.8717
        X_SPACING                | 17.8198 |
        Y_SPACING                | 24.7188 |

    Returns a dict: {name: (x, y)} for point values, {name: x} for
    single-value rows (X_SPACING, Y_SPACING - no Y).
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "RELAY_RACK_POSITIONS" not in wb.sheetnames:
        raise ValueError(
            "Workbook has no 'RELAY_RACK_POSITIONS' sheet. Add one with columns "
            "Parameter/X/Y to control Relay Rack element positions - see "
            "relay_rack_core.py's read_relay_rack_position_config() docstring for the layout."
        )

    ws = wb["RELAY_RACK_POSITIONS"]
    config = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or str(row[0]).strip() == "":
            continue
        name = str(row[0]).strip()
        x_val = row[1]
        y_val = row[2] if len(row) > 2 else None
        if x_val is None:
            raise ValueError(f"RELAY_RACK_POSITIONS row {name!r} has no X value")
        if y_val is None or str(y_val).strip() == "":
            config[name] = float(x_val)
        else:
            config[name] = (float(x_val), float(y_val))

    required = ["FIRST_POSITION_POINT", "X_SPACING", "Y_SPACING"]
    missing = [name for name in required if name not in config]
    if missing:
        raise ValueError(f"RELAY_RACK_POSITIONS sheet is missing required rows: {missing}")

    return config


def read_name2_whitelist(xlsx_path: str) -> list:
    """
    CONFIRMED: the known/possible NAME2 (relay type) values live in
    RELAY_RACK_POSITIONS!E2 as a comma-separated list (e.g. 'DZR,HHZR,
    HZR,...,DN SDF,UP SDF,UP AT,DN AT,LV LOAD,...,TS(M),TS(R),TS(A),
    TS(B)'). CONFIRMED: everything else in a tag cell belongs to NAME1
    (which may itself contain spaces, e.g. 'LC 22'). This is what makes
    correct splitting possible for multi-word NAME1 values - naive
    "split on first/last space" breaks those cases (e.g. 'LC 22 LCPR1'
    would wrongly split into NAME1='LC'/NAME2='22 LCPR1' on first-space,
    or NAME1='LC 22 LCPR'/NAME2='1' on last-space).

    Returns entries sorted so multi-word values (e.g. 'DN SDF') are
    tried before single-word ones, and longer values before shorter -
    avoids a shorter suffix accidentally matching first.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["RELAY_RACK_POSITIONS"]
    raw = ws["E2"].value or ""
    values = [v.strip() for v in str(raw).split(",") if v.strip()]
    values.sort(key=lambda v: (-len(v.split()), -len(v)))
    return values


def split_bit_name(tag_text: str, name2_whitelist: list = None):
    """
    CONFIRMED shared splitter: given a relay "tag" cell (e.g.
    'LC 22 LCPR1', 'F54 TSPR', 'EB'), returns (NAME1, NAME2) using the
    whitelist - tries each known NAME2 value (longest/multi-word first)
    as a whole-word suffix; whatever's left (may itself contain spaces)
    is NAME1. Falls back to legacy "split on first space" (single word
    -> NAME1='', NAME2=that word) if nothing in the whitelist matches -
    e.g. genuinely unrecognized/mistyped data - so this never raises,
    just degrades gracefully.
    """
    text = str(tag_text).strip()
    if name2_whitelist:
        text_upper = text.upper()

        # PASS 1 (preferred/precise): candidate must be a whole-word
        # suffix, preceded by a space - avoids accidentally matching
        # inside an unrelated word.
        for candidate in name2_whitelist:
            cand_upper = candidate.strip().upper()
            if not cand_upper:
                continue
            if text_upper == cand_upper:
                return "", text
            suffix = " " + cand_upper
            if text_upper.endswith(suffix):
                name1 = text[: -len(suffix)].strip()
                name2 = text[-len(candidate):].strip()
                return name1, name2

        # PASS 2 (fallback for data-entry mistakes): CONFIRMED some tag
        # cells have the location and relay type glued together with NO
        # space at all (e.g. 'CGL 76TS(M)' instead of 'CGL 76 TS(M)',
        # or 'LOC DNSDF' instead of 'LOC DN SDF') - match the suffix
        # directly even without a space boundary. For multi-word
        # candidates (e.g. "DN SDF"), also try matching against the
        # candidate with its OWN internal space removed ("DNSDF").
        for candidate in name2_whitelist:
            cand_upper = candidate.strip().upper()
            if not cand_upper:
                continue
            cand_glued = cand_upper.replace(" ", "")
            if text_upper.endswith(cand_glued) and len(text) > len(cand_glued):
                name1 = text[: -len(cand_glued)].strip()
                name2 = candidate.strip()  # use the whitelist's own spacing for the match
                if name1:
                    return name1, name2

    # Fallback: legacy behavior - split on first space if 2+ words,
    # otherwise treat a lone word as NAME2 (relay type), not NAME1.
    parts = text.split(" ", 1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    return "", parts[0].strip()


def _split_tag(tag_text: str, name2_whitelist: list = None):
    """Kept for backward compatibility - delegates to split_bit_name()."""
    return split_bit_name(tag_text, name2_whitelist)


def _split_type(type_text: str):
    """'QN1,8F-8B' -> ('QN1', '8F-8B'), splitting on comma."""
    parts = str(type_text).split(",", 1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    return parts[0].strip(), ""


def read_all_relay_racks(xlsx_path: str):
    """
    CONFIRMED fixed structure: each rack's grid is bounded to band
    letters A-G (7 letter-rows, i.e. 14 Excel rows of tag/type pairs)
    and columns 1-21 (positions) - data beyond either boundary within a
    single rack is not part of the grid and is never read.

    CONFIRMED (reverted from an earlier, incorrect assumption): "END"
    only ends the CURRENT rack - reading then continues to look for
    another rack header right after it (e.g. "R2"), and keeps reading
    every rack found this way until a header row is genuinely blank.
    Multiple racks (R1, R2, ...) are real, needed data, not stale
    leftovers - all of them get processed.

    Reads, per rack:
      - starts with a "header" row (column A holds the rack's own
        identifier, e.g. 'R1' - CONFIRMED at row 2 for the first rack)
      - immediately followed by its grid data (tag/type row pairs),
        capped at 7 bands (A-G)
      - ends at a row whose column A value is "END" (or after 7 bands,
        whichever comes first)
      - the NEXT rack (if any) starts at the row right after that 'END'

    Returns a list of (rack_num, grid_rows) tuples, one per rack found.
    """
    import openpyxl

    MAX_BANDS = 7          # CONFIRMED: bands A-G only, per rack
    MAX_DATA_COLUMN = 22   # CONFIRMED: column B (2) through column V (22) = 21 positions

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["RELAY RACK"]
    max_col = min(ws.max_column, MAX_DATA_COLUMN)
    name2_whitelist = read_name2_whitelist(xlsx_path)

    racks = []
    header_row = 2  # CONFIRMED: first rack's header/identifier row
    while True:
        rack_num_val = ws.cell(row=header_row, column=1).value
        if rack_num_val is None or str(rack_num_val).strip() == "":
            break  # no more racks
        rack_num = str(rack_num_val).strip()

        grid_rows = []
        excel_row = header_row + 1
        end_row = excel_row  # fallback if MAX_BANDS is hit without ever seeing "END"
        for _band_index in range(MAX_BANDS):
            row_label = ws.cell(row=excel_row, column=1).value
            if row_label is None or str(row_label).strip().upper() == "END":
                end_row = excel_row
                break

            tag_row = excel_row
            type_row = excel_row + 1
            cells = []
            for col in range(2, max_col + 1):  # column B is the first data column
                tag_val = ws.cell(row=tag_row, column=col).value
                if tag_val is None or str(tag_val).strip() == "":
                    cells.append({"NAME1": "", "NAME2": "", "RELTYP": "", "CNTCNF": "", "NAME3": "SPARE"})
                else:
                    name1, name2 = _split_tag(tag_val, name2_whitelist)
                    type_val = ws.cell(row=type_row, column=col).value or ""
                    reltyp, cntcnf = _split_type(type_val)
                    cells.append({"NAME1": name1, "NAME2": name2, "RELTYP": reltyp, "CNTCNF": cntcnf, "NAME3": ""})

            grid_rows.append(cells)
            excel_row += 2
            end_row = excel_row  # keeps advancing in case MAX_BANDS is hit with no explicit "END"

        racks.append((rack_num, grid_rows))
        header_row = end_row + 1  # next rack's header, if any, starts right after this rack's end

    return racks


def build_relay_position_lookup(xlsx_path: str, log=None):
    """
    Builds two lookups from the RELAY RACK sheet:
      1. by_pair:  {(NAME1.upper(), NAME2.upper()): "R(POS)" code}
         Primary match - e.g. rack 'R1', band letter 'A' (1st tag/type
         row pair), column B (1st data column) -> "R1A1". CONFIRMED
         convention (same one every circuit-type module already uses):
         a Relay Rack tag cell like 'F656 HZRP' splits into
         NAME1='F656'/NAME2='HZRP' - ANY Front_Contact/Back_Contact/
         Relay_Coil instance whose S_NAME='F656' AND R_NAME='HZRP' is
         that SAME physical relay.
      2. by_type: {NAME2.upper(): "R(POS)" code or None if ambiguous}
         Fallback for blocks that carry NO S_NAME at all (CONFIRMED:
         AR & AZR's and Track's own relay coil, tagged R_POS not
         R(POS)) - matched by R_NAME/NAME2 alone. CONFIRMED this is
         accepted as less precise/possibly ambiguous: if more than one
         Relay Rack entry shares the same NAME2, by_type maps to None
         (caller treats that as "ambiguous", not a clean vital error).

    log: CONFIRMED FIX (from Krish) - optional GenerationLog. If the
        SAME relay name (NAME1+NAME2 pair) is found in MORE than one
        Relay Rack position, that's a genuine data-entry mistake in the
        source Excel (the same relay accidentally typed into two
        different physical cells) - previously this was a silent
        dict-overwrite with no warning at all, meaning one of the two
        physical positions became invisible with no indication anything
        was wrong. Now raised as a VITAL error (log.error) naming the
        relay and both colliding positions, stopping generation before
        any output - same as every other genuine data conflict this
        app catches. Kept optional so any other/older caller that
        doesn't pass log still works, just without this specific check.

    Returns (by_pair, by_type).
    """
    by_pair = {}
    by_type = {}
    seen_type_count = {}

    for rack_num, grid_rows in read_all_relay_racks(xlsx_path):
        for band_index, row in enumerate(grid_rows):
            band_letter = chr(ord("A") + band_index)
            for col_index, cell in enumerate(row):
                if cell.get("NAME3", "").strip().upper() == "SPARE":
                    continue
                name1 = cell.get("NAME1", "").strip()
                name2 = cell.get("NAME2", "").strip()
                if not name2:
                    continue  # CONFIRMED: NAME1 can legitimately be blank (e.g. "DN SDF" alone) - only NAME2 is required
                position_number = col_index + 1  # column B (index 0) -> position 1
                position_code = f"{rack_num}{band_letter}{position_number}"

                pair_key = (name1.upper(), name2.upper())
                if log is not None and pair_key in by_pair:
                    log.error(
                        f"[RELAY RACK] Duplicate relay: '{name1} {name2}' found at both "
                        f"{by_pair[pair_key]} and {position_code} - the same relay name was "
                        "entered in more than one Relay Rack position. Fix the Excel data "
                        "before generating."
                    )
                by_pair[pair_key] = position_code

                type_key = name2.upper()
                seen_type_count[type_key] = seen_type_count.get(type_key, 0) + 1
                by_type[type_key] = position_code if seen_type_count[type_key] == 1 else None

    return by_pair, by_type





RELAY_BLOCK_NAME = "SingleRelay_20"


def _import_relay_block_definition(doc):
    """
    Import RELAY_BLOCK.dxf's block DEFINITION once, so every rack cell can
    insert a reference to this SAME block name instead of each cell
    bringing its own renamed copy (SingleRelay_20, _201, _202, ...).
    """
    if RELAY_BLOCK_NAME in doc.blocks:
        return  # already imported
    source_doc = ezdxf.readfile(resolve_template_path("RELAY_BLOCK.dxf"))
    imp = importer.Importer(source_doc, doc)
    imp.import_block(RELAY_BLOCK_NAME)
    imp.finalize()


def get_relay_rack_sheet_number(xlsx_path: str) -> str:
    """
    Reads the Relay Rack sheet number from FIELD PG.NO (row where
    Circuits == 'RELAY RACK'). Returns '' if blank (not yet assigned).
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(row=r, column=2).value or "").strip().upper()
        if label == "RELAY RACK":
            value = ws.cell(row=r, column=3).value
            return str(value).strip() if value is not None else ""
    return ""


def get_next_circuit_after_relay_rack(xlsx_path: str) -> str:
    """Reads the starting sheet number of whatever circuit follows RELAY RACK in FIELD PG.NO."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(row=r, column=2).value or "").strip().upper()
        if label == "RELAY RACK":
            value = ws.cell(row=r + 1, column=3).value
            return str(value).strip() if value is not None else ""
    return ""


def generate_one_relay_rack_drawing(rack_num: str, grid_rows: list, positions: dict, border_template_path: str, sht_number: str, cont_number: str):
    """
    Builds ONE Relay Rack drawing from already-read grid data.
    Returns (output_filename, ezdxf.Drawing).
    """
    if not grid_rows:
        raise ValueError(f"Rack {rack_num!r} has no grid rows (nothing before 'END')")

    first_x, first_y = positions["FIRST_POSITION_POINT"]
    x_spacing = positions["X_SPACING"]
    y_spacing = positions["Y_SPACING"]

    # Base doc: RELAY_RACK.dxf, inserted at (0,0) -> (0,0), i.e. unchanged.
    doc = ezdxf.readfile(resolve_template_path("RELAY_RACK.dxf"))
    msp = doc.modelspace()

    # RELAYRACKNUM is the FULL string "RELAY RACK - R1" (matching the
    # block's own native default format, "RELAY RACK - 1").
    relay_rack_num_text = f"RELAY RACK - {rack_num}"
    for e in msp:
        if e.dxftype() == "INSERT" and e.dxf.name == "RELAYRACK":
            for att in e.attribs:
                if att.dxf.tag == "RELAYRACKNUM":
                    att.dxf.text = _safe_str(relay_rack_num_text)

    _import_relay_block_definition(doc)

    for row_index, cells in enumerate(grid_rows):
        row_y = first_y - row_index * y_spacing
        for col_index, cell in enumerate(cells):
            cell_x = first_x + col_index * x_spacing
            ins = msp.add_blockref(RELAY_BLOCK_NAME, (cell_x, row_y))
            ins.add_auto_attribs({tag: _safe_str(val) for tag, val in cell.items()})

    _insert_border_title_generic(
        doc=doc,
        border_template_path=border_template_path,
        template_name="RELAY_RACK",
        sig_name=None,
        ahead_sig=None,
        loc=None,
        sht_number=sht_number,
        cont_number=cont_number,
        title_text_override=relay_rack_num_text,
    )

    out_name = f"RELAY_RACK_{rack_num}.dxf"
    return out_name, doc


def _increment_sheet_designator(designator: str) -> str:
    """
    Increments a sheet number designator by one step:
      - Pure numeric ('000') -> next integer, same zero-padded width.
      - Numeric prefix + trailing letter ('00F') -> increments the letter
        ('00F' -> '00G'), prefix unchanged.
    """
    if designator.isdigit():
        width = len(designator)
        return str(int(designator) + 1).zfill(width)

    import re
    m = re.match(r"^(\d*)([A-Za-z])$", designator)
    if m:
        prefix, letter = m.group(1), m.group(2)
        next_letter = chr(ord(letter) + 1)
        return f"{prefix}{next_letter}"

    raise ValueError(f"Don't know how to increment sheet designator {designator!r}")


def generate_all_relay_rack_drawings(xlsx_path: str, positions: dict, border_template_path: str):
    """
    Builds ONE drawing PER relay rack found in the sheet (each rack gets
    its own separate output file, per confirmation). Sheet numbers run
    sequentially starting from FIELD PG.NO's RELAY RACK row, with the
    LAST rack's CONT pointing to whatever circuit follows RELAY RACK.

    Returns list of (output_filename, ezdxf.Drawing, sht_number_str).
    """
    racks = read_all_relay_racks(xlsx_path)
    if not racks:
        raise ValueError("No relay racks found in RELAY RACK sheet")

    start_sheet_number_str = get_relay_rack_sheet_number(xlsx_path) or "000"
    next_circuit_start = get_next_circuit_after_relay_rack(xlsx_path) or start_sheet_number_str

    current = start_sheet_number_str
    results = []
    for i, (rack_num, grid_rows) in enumerate(racks):
        is_last = i == len(racks) - 1
        sht = current
        cont = next_circuit_start if is_last else _increment_sheet_designator(current)
        current = cont if not is_last else current

        out_name, doc = generate_one_relay_rack_drawing(
            rack_num, grid_rows, positions, border_template_path, sht, cont
        )
        results.append((out_name, doc, sht))

    return results
