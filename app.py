import os
import uuid
import shutil
import zipfile
import ezdxf
from retemplate_core import apply_new_template
import threading
import secrets
from concurrent.futures import ThreadPoolExecutor

import openpyxl
from flask import Flask, request, jsonify, render_template, send_from_directory, session, redirect, url_for
from functools import wraps

import auth_core
import email_core

from signal_core import (
    generate_signal_drawing_set,
    get_signal_start_sheet_number,
    get_next_circuit_start_sheet_number,
    set_cont_value,
    read_typical_config,
)
from track_core import (
    generate_all_track_sheets,
    get_track_start_sheet_number,
    get_next_circuit_after_track,
)
from ar_azr_core import (
    generate_all_ar_azr_sheets,
    get_ar_azr_start_sheet_number,
    get_next_circuit_after_ar_azr,
    get_ar_azr_contact_groups_h3,
    get_ar_azr_contact_groups_g3,
    read_ar_azr_position_config,
)
from relay_rack_core import (
    generate_all_relay_rack_drawings,
    read_relay_rack_position_config,
    build_relay_position_lookup,
    read_name2_whitelist,
)
from communication_core import (
    generate_all_communication_sheets,
    get_communication_start_sheet_number,
    get_next_circuit_after_communication,
)
from sdf_core import (
    generate_all_sdf_sheets,
    get_sdf_start_sheet_number,
    get_next_circuit_after_sdf,
)
from datalogger_core import (
    generate_all_datalogger_sheets,
    get_datalogger_start_sheet_number,
    get_next_circuit_after_datalogger,
    build_used_contacts_registry,
)
from custom_circuits_core import (
    generate_all_custom_circuit_sheets,
    get_custom_circuits_start_sheet_number,
    get_next_circuit_after_custom_circuits,
)
from generation_log import GenerationLog, ContactRegistry
from contact_analysis_core import (
    generate_all_contact_analysis_sheets,
    get_contact_analysis_start_sheet_number,
    ContactAnalysisCancelled,
)
from index_core import (
    get_index_start_sheet_number,
    build_index_entries_with_spare_gaps,
    add_index_self_listing,
    generate_all_index_sheets,
    _sheet_number_sort_key,
)
from cover_sheet_core import generate_cover_sheet
from station_layout_core import generate_all_station_layout_sheets
from floor_plan_core import generate_all_floor_plan_sheets
import attsync_core
from lcpr_allocation_core import apply_lcpr_allocations, validate_lcpr_relay_availability
from relay_position_core import apply_relay_positions, clear_placeholder_values, apply_filename_attribute, apply_hut_name_keyword
from oda_convert import convert_folder_to_dwg, convert_single_file_to_dxf, OdaConversionError

# CONFIRMED: some situations (e.g. Track/SDF relays consistently missing
# from Relay Rack in pairs - see contact_analysis_core.py) need a REAL
# pause-and-ask prompt in the UI, not just a log entry - the job thread
# blocks on an Event until /api/confirm/<job_id> answers it.
pending_confirmations = {}   # job_id -> threading.Event
confirmation_answers = {}    # job_id -> bool


def ask_user_confirmation(job_id, message: str) -> bool:
    event = threading.Event()
    with jobs_lock:
        pending_confirmations[job_id] = event
        jobs[job_id]["state"] = "awaiting_confirmation"
        jobs[job_id]["confirmation_message"] = message
    event.wait()
    with jobs_lock:
        answer = confirmation_answers.pop(job_id, False)
        pending_confirmations.pop(job_id, None)
    return answer

APP_DIR = os.path.dirname(os.path.abspath(__file__))
JOBS_DIR = os.path.join(APP_DIR, "jobs")
UPLOADS_DIR = os.path.join(APP_DIR, "uploads")
os.makedirs(JOBS_DIR, exist_ok=True)
os.makedirs(UPLOADS_DIR, exist_ok=True)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 25 * 1024 * 1024  # 25 MB
def _get_or_create_secret_key():
    """
    CONFIRMED: with multiple active users, regenerating a random key on
    every restart (the old behavior) logs EVERYONE out every time the
    server restarts - annoying at 20-user scale. Prefers
    MSDAC_SECRET_KEY (env var) if set; otherwise persists a
    once-generated key to a local file, so sessions survive restarts
    even without manual configuration.
    """
    env_key = os.environ.get("MSDAC_SECRET_KEY")
    if env_key:
        return env_key
    key_file = os.path.join(os.path.dirname(__file__), ".secret_key")
    if os.path.exists(key_file):
        with open(key_file, "r") as f:
            return f.read().strip()
    new_key = secrets.token_hex(32)
    with open(key_file, "w") as f:
        f.write(new_key)
    return new_key


app.secret_key = _get_or_create_secret_key()

auth_core.init_db()


def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("userid"):
            if request.path.startswith("/api/"):
                return jsonify({"error": "Not logged in"}), 401
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return wrapper


@app.before_request
def _track_online_activity():
    # CONFIRMED: updates a "last active" timestamp on every request from
    # a logged-in user - this is what the admin panel's "online" status
    # is based on (active within the last few minutes), not a live
    # persistent connection.
    userid = session.get("userid")
    if userid:
        auth_core.touch_last_active(userid)


def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("is_admin"):
            if request.path.startswith("/api/"):
                return jsonify({"error": "Admin access required"}), 403
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return wrapper

# In-memory job status store. Keyed by job_id (UUID), each job is isolated
# to its own folder under JOBS_DIR, so concurrent engineers never collide.
jobs = {}
jobs_lock = threading.Lock()

executor = ThreadPoolExecutor(max_workers=8)

# CONFIRMED: pause/resume mechanism for the live contact-conflict picker.
# Contact Analysis runs inside a background job thread - when it hits a
# repeated-contact conflict, it needs to BLOCK that one thread (not the
# whole server) until the user picks an alternative via the browser,
# then continue with that choice. Keyed by job_id, one Event per job
# that currently has a conflict awaiting a choice.
contact_choice_events = {}
contact_choice_events_lock = threading.Lock()


def wait_for_contact_choice(job_id, relay_key, available_pairs, sheet_nums=None, circuit_types=None, occurrence_choices=None, contact_code=None):
    """
    Called from inside the Contact Analysis pipeline (background job
    thread) when a repeated-contact conflict is found. Blocks THIS
    thread only - other jobs' threads are unaffected - until the user
    submits a choice via /api/resolve-contact/<job_id>.

    occurrence_choices: list of {"sheet_number", "circuit_type"} dicts,
    one per conflicting occurrence - the user picks BOTH an alternative
    contact pair AND which specific occurrence (sheet) gets reassigned
    to it; every other occurrence is left untouched.

    Returns (chosen_pair, chosen_occurrence_index), or None if no
    options were available at all (caller should fall back to logging
    the conflict as a hard error, same as before this feature existed).
    """
    if not available_pairs:
        return None

    event = threading.Event()
    with contact_choice_events_lock:
        contact_choice_events[job_id] = event

    _set_status(
        job_id,
        state="awaiting_contact_choice",
        conflict_relay=f"{relay_key[0]} {relay_key[1]}".strip(),
        conflict_code=contact_code,
        conflict_options=[f"{p[0]}/{p[1]}" for p in available_pairs],
        conflict_sheets=list(sheet_nums or []),
        conflict_circuit_types=list(circuit_types or []),
        conflict_occurrences=occurrence_choices or [],
    )

    event.wait()  # blocks until /api/resolve-contact sets it

    with jobs_lock:
        cancelled = jobs[job_id].get("contact_choice_cancelled")
        chosen = jobs[job_id].get("chosen_contact_pair")
        chosen_index = jobs[job_id].get("chosen_occurrence_index")
    with contact_choice_events_lock:
        contact_choice_events.pop(job_id, None)

    if cancelled:
        return "CANCELLED"
    if not chosen or chosen_index is None:
        return None
    first, second = chosen.split("/")
    return ((first, second), chosen_index)


# CONFIRMED: raised from 4 to support ~20 simultaneous users - not all
# 20 submitting a generation job in the exact same instant is unlikely
# in practice (jobs are periodic, not constant), so 8 concurrent DXF
# generations comfortably covers realistic simultaneous load without
# over-committing CPU/memory. If your server has fewer than ~8 CPU
# cores, consider lowering this - DXF generation is genuinely
# CPU/memory-intensive work, and raising this number doesn't create
# more capacity than the actual hardware has; it only changes how many
# jobs get to compete for that hardware at once instead of queueing.


def _set_status(job_id, **kwargs):
    with jobs_lock:
        jobs[job_id].update(kwargs)
        if kwargs.get("state") == "done":
            job = jobs[job_id]
            userid = job.get("userid")
            download_name = job.get("download_name", "")
            job_type = job.get("job_type") or (download_name.replace(".zip", "").replace("_", " ").title() if download_name else "Full IFC")
            if userid:
                auth_core.record_job_history(job_id, userid, job_type, download_name, job.get("hut_name", ""), job.get("section", ""))


def _sheet_title_text(doc):
    # CONFIRMED FIX: don't assume the title block is named "TITLE" or
    # "TITLEBLOCK" - after Re-template swaps in a new template, the
    # block gets named after the new template file instead. Look for
    # ANY block instance carrying a "TITLE" attribute tag, regardless
    # of what the block itself is called.
    for e in doc.modelspace():
        if e.dxftype() == "INSERT":
            for att in e.attribs:
                if att.dxf.tag == "TITLE":
                    return (att.dxf.text or "").strip()
    return ""


def _generate_index_for_pending(pending_pairs, xlsx_path, template_path, filename_pattern, dxf_folder, log, generate_index_enabled):
    """
    CONFIRMED: same INDEX-generation logic as the Full IFC job, scoped
    to whatever pending_pairs was actually passed in - lets an
    individual generate-button run (or the Re-template page) also
    produce its own INDEX sheet covering just that batch of files, not
    the full project. No-op if the checkbox wasn't ticked. Appends the
    generated INDEX sheet(s) directly onto pending_pairs (same list,
    modified in place) so they get included in that endpoint's own
    output zip.

    xlsx_path is now OPTIONAL: if not given (e.g. the Re-template page,
    which has no Excel of its own), the INDEX's own starting sheet
    number is instead inferred by looking for an EXISTING sheet among
    pending_pairs whose own TITLE text contains "INDEX" - reusing that
    sheet's own SHT value as the starting point, exactly the way it
    already works when Excel IS available. If no such sheet is found
    and no xlsx_path was given either, INDEX generation is skipped with
    a clear warning rather than guessing.
    """
    if not generate_index_enabled:
        return
    try:
        index_source_entries = []
        existing_index_sht = None
        for final_name, doc in pending_pairs:
            title_text = _sheet_title_text(doc)
            for e in doc.modelspace():
                if e.dxftype() == "INSERT":
                    found_sht = False
                    for att in e.attribs:
                        if att.dxf.tag == "SHT":
                            sht = (att.dxf.text or "").strip()
                            name_no_ext, _ext = os.path.splitext(final_name)
                            index_source_entries.append((title_text, sht, name_no_ext))
                            if "INDEX" in title_text.strip().upper():
                                existing_index_sht = sht
                            found_sht = True
                    if found_sht:
                        break

        if xlsx_path:
            index_start_sht = get_index_start_sheet_number(xlsx_path)
        elif existing_index_sht:
            index_start_sht = existing_index_sht
        else:
            log.warning("Index: no Excel workbook given and no existing INDEX sheet found among the files to infer a starting sheet number from - skipped.")
            return

        ordered_sheets = sorted(index_source_entries, key=lambda t: _sheet_number_sort_key(t[1]))
        ordered_sheets = add_index_self_listing(ordered_sheets, index_start_sht, filename_pattern)
        ordered_sheets.sort(key=lambda t: _sheet_number_sort_key(t[1]))
        index_entries = build_index_entries_with_spare_gaps(ordered_sheets, filename_pattern)
        index_results, _next_index_sht = generate_all_index_sheets(index_entries, index_start_sht, template_path)
        for out_name, index_doc, index_sht in index_results:
            name_no_ext, ext = os.path.splitext(out_name)
            final_name = _build_output_filename(name_no_ext, index_sht, filename_pattern) + ext
            index_doc.saveas(os.path.join(dxf_folder, final_name))
            pending_pairs.append((final_name, index_doc))
    except Exception as e:
        log.warning(f"Index: {e}")


def _apply_relay_position_check(pending_pairs, xlsx_path, relay_position_enabled, log, filename_to_circuit_type=None):
    """
    pending_pairs: list of (final_name, doc) tuples for whatever circuit
    type just finished generating. No-op if the checkbox wasn't ticked.
    Shared across every individual generate-button job AND the Full IFC
    job, so Relay Position works the same regardless of which button
    was used - CONFIRMED requirement (checkbox applies wherever ticked,
    not just on a combined run). filename_to_circuit_type is only
    meaningful for the Full IFC job (individual endpoints only ever deal
    with one circuit type at a time anyway).
    """
    # CONFIRMED FIX: LCPR allocation must run BEFORE Relay Position -
    # it's what resolves which LCPR/LCPR1/LCPR2 instance each block
    # actually belongs to (writing the real R_NAME). Relay Position's
    # own lookup needs that ALREADY resolved to find the correct
    # R(POS); running it first meant every LCPR block's position got
    # looked up using the raw, unresolved "LCPR" literal regardless of
    # which signal it really belonged to - producing a wrong R(POS)
    # without raising an error, since the lookup still technically
    # found *some* position (just not accounting for the LCPR1/LCPR2
    # distinction). Runs regardless of whether Relay Position was
    # ticked, and regardless of which individual button was used
    # (Signal or otherwise) - LCPR allocation is its own separate
    # feature, purely driven by whether the SIGNAL sheet's own "LCPR"
    # column is filled. Harmless no-op for non-Signal endpoints (finds
    # no LCPR_FRONT/BACK blocks tagged with a matching owning signal,
    # so nothing changes).
    try:
        apply_lcpr_allocations(pending_pairs, xlsx_path, log)
    except Exception as e:
        log.warning(f"LCPR Allocation: {e}")

    if relay_position_enabled:
        try:
            position_lookup = build_relay_position_lookup(xlsx_path, log)
            apply_relay_positions(pending_pairs, position_lookup, log, filename_to_circuit_type)
        except Exception as e:
            log.warning(f"Relay Position: {e}")


def _stop_if_vital_errors(job_id, job_folder, log) -> bool:
    """
    Returns True (and already set the job's error status) if the log has
    ANY vital error logged against it - Relay Position's unmatched
    relay, Contact Analysis's duplicate contact usage or unmatched
    relay, or any other log.error() call - same stop-before-any-output
    behavior used everywhere in this app.
    """
    if log.has_errors:
        log_xlsx_path = os.path.join(job_folder, "generation_log.xlsx")
        log.save_excel(log_xlsx_path)
        error_messages = [msg for level, msg in log.entries if level == "ERROR"]
        _set_status(
            job_id,
            state="error",
            error="Critical error(s) found - generation stopped before producing output:\n"
            + "\n".join(error_messages),
            log_errors=len(error_messages),
        )
        return True
    return False


def _build_download_zip_name(base_name: str, hut_name: str, section: str) -> str:
    """
    CONFIRMED: download zip filenames now include HUT NAME, SECTION,
    and the current date/time (at download-generation time), to avoid
    confusion between multiple downloads and to give a clear record for
    database/job-history logging - e.g. "HUT1_SECTION-A_20260728_1430_signal_circuits.zip".
    Falls back to just the date/time + base name if HUT NAME/SECTION
    weren't provided (kept optional, not required).
    """
    import datetime
    import re as _re

    def _sanitize(value):
        value = (value or "").strip()
        return _re.sub(r"[^A-Za-z0-9_-]", "_", value)

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    hut_part = _sanitize(hut_name)
    section_part = _sanitize(section)

    parts = [p for p in (hut_part, section_part, timestamp) if p]
    return "_".join(parts) + f"_{base_name}"


def _build_output_filename(descriptive_name_no_ext, sht_number, pattern):
    """
    descriptive_name_no_ext: e.g. 'SKL 8_HZR_HHZR_DZR_SHT001' (used only when
        no pattern is given, so existing behavior is preserved)
    sht_number: e.g. '002'
    pattern: whatever the user typed, e.g. 'PR_[]' or 'PR[]'. '[]' is
        replaced with JUST the sheet number - signal name / typical are
        dropped entirely from the output filename when a pattern is given.
        If pattern has no '[]', the sheet number is appended after it.
    """
    pattern = (pattern or "").strip()
    if not pattern:
        return descriptive_name_no_ext
    if "[]" in pattern:
        return pattern.replace("[]", sht_number)
    return f"{pattern}{sht_number}"


def _add_source_files_to_zip(zf, job_folder):
    """
    CONFIRMED: bundle the original uploaded source Excel workbook,
    template file, and (for Custom Circuits / Full IFC) the uploaded
    custom circuits.zip into the downloadable zip alongside the
    generated outputs, so "download all" always includes both inputs
    and outputs together - not just outputs. All three live at
    predictable paths within job_folder (source.xlsx,
    border_template.<ext>, circuits.zip) for every generation job, so
    no extra bookkeeping is needed to find them.
    """
    source_xlsx = os.path.join(job_folder, "source.xlsx")
    if os.path.exists(source_xlsx):
        zf.write(source_xlsx, arcname="source_files/source.xlsx")
    uploaded_circuits_zip = os.path.join(job_folder, "circuits.zip")
    if os.path.exists(uploaded_circuits_zip):
        zf.write(uploaded_circuits_zip, arcname="source_files/uploaded_circuits.zip")
    for fname in os.listdir(job_folder):
        if fname.startswith("border_template"):
            zf.write(os.path.join(job_folder, fname), arcname=f"source_files/{fname}")


def _group_job_history(history_rows):
    """
    CONFIRMED: group flat job_history rows into hut_name+section groups
    for the foldable history/admin pages. Each group keeps its jobs in
    the same order as the input (already most-recent-first from the DB
    query), and groups themselves stay ordered by their own most recent
    job, since the first occurrence of a given hut/section determines
    where the group is inserted.
    """
    groups = []
    group_index = {}
    for h in history_rows:
        key = (h["hut_name"] or "", h["section"] or "")
        if key not in group_index:
            group_index[key] = {"hut_name": key[0], "section": key[1], "jobs": []}
            groups.append(group_index[key])
        group_index[key]["jobs"].append(h)
    return groups


def _read_signal_rows(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["SIGNAL"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    headers = rows[0]
    data_rows = []
    for r in rows[1:]:
        row_dict = dict(zip(headers, r))
        if row_dict.get("SIG NAME") in (None, ""):
            continue
        data_rows.append(row_dict)
    return data_rows


def _run_generation_job(job_id, xlsx_path, template_path, filename_pattern, relay_position_enabled=False, hut_name="", section="", script_type="both"):
    job_folder = os.path.join(JOBS_DIR, job_id)
    dxf_folder = os.path.join(job_folder, "dxf")
    dwg_folder = os.path.join(job_folder, "dwg")
    os.makedirs(dxf_folder, exist_ok=True)
    os.makedirs(dwg_folder, exist_ok=True)

    try:
        # Border/title-block template may be uploaded as .dwg - convert to
        # .dxf first since block-import works on DXF via ezdxf.
        if template_path.lower().endswith(".dwg"):
            _set_status(job_id, state="preparing_template", progress=2)
            template_path = convert_single_file_to_dxf(template_path)

        _set_status(job_id, state="reading_excel", progress=5)
        rows = _read_signal_rows(xlsx_path)
        if not rows:
            raise ValueError("No signal rows found in SIGNAL sheet")

        start_sheet_number = get_signal_start_sheet_number(xlsx_path)
        next_circuit_start = get_next_circuit_start_sheet_number(xlsx_path)
        sheet_sets_override, title_map_override, lcpr_config = read_typical_config(xlsx_path)

        _set_status(job_id, state="generating_dxf", progress=10, total_signals=len(rows))

        sheet_num = start_sheet_number
        pending = []  # (final_name, doc) - saved only after the CONT override below
        for i, row in enumerate(rows):
            try:
                sheets, sheet_num = generate_signal_drawing_set(
                    row, start_sheet_number=sheet_num, border_template_path=template_path,
                    sheet_sets_override=sheet_sets_override, title_map_override=title_map_override,
                    lcpr_config=lcpr_config, hut_name_input=hut_name,
                )
            except Exception as row_err:
                _set_status(
                    job_id,
                    state="error",
                    error=f"Row {i + 2} (SIG NAME={row.get('SIG NAME')}): {row_err}",
                )
                return

            for out_name, doc, sht in sheets:
                name_no_ext, ext = os.path.splitext(out_name)
                final_name = _build_output_filename(name_no_ext, sht, filename_pattern) + ext
                pending.append((final_name, doc))

            pct = 10 + int(50 * (i + 1) / len(rows))
            _set_status(job_id, progress=pct, signals_done=i + 1)

        # The very last Signal sheet's CONT should point to the next
        # circuit type's starting sheet (FIELD PG.NO!C7), not a simple
        # +1 continuation of Signal's own numbering.
        if pending:
            last_name, last_doc = pending[-1]
            set_cont_value(last_doc, next_circuit_start)

        log = GenerationLog()
        _apply_relay_position_check(pending, xlsx_path, relay_position_enabled, log)
        clear_placeholder_values(pending)
        apply_filename_attribute(pending)
        apply_hut_name_keyword(pending, hut_name)
        if _stop_if_vital_errors(job_id, job_folder, log):
            return

        generated_files = []
        for final_name, doc in pending:
            out_path = os.path.join(dxf_folder, final_name)
            doc.saveas(out_path)
            generated_files.append(final_name)

        _set_status(job_id, state="converting_dwg", progress=65)
        convert_folder_to_dwg(dxf_folder, dwg_folder)

        # CONFIRMED: same ATTSYNC script generation as Full IFC, now
        # also available for this individual button - respects the same
        # script_type selection (autocad/both/arescad).
        try:
            if script_type in ("arescad", "both"):
                scr_text = attsync_core.generate_attsync_scr(dxf_folder, target="ares")
                with open(os.path.join(dwg_folder, "attsync.scr"), "w", encoding="utf-8") as f:
                    f.write(scr_text)
            if script_type in ("autocad", "both"):
                scr_text_autocad = attsync_core.generate_attsync_scr(dxf_folder, target="autocad")
                with open(os.path.join(dwg_folder, "attsync_autocad.scr"), "w", encoding="utf-8") as f:
                    f.write(scr_text_autocad)
        except Exception:
            pass  # non-fatal - a missing ATTSYNC script shouldn't fail the whole job

        _set_status(job_id, state="zipping", progress=90)
        zip_path = os.path.join(job_folder, "signal_circuits.zip")
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            _add_source_files_to_zip(zf, job_folder)
            for fname in os.listdir(dwg_folder):
                zf.write(os.path.join(dwg_folder, fname), arcname=fname)
            # CONFIRMED: always include the log, not just when something
            # went wrong - so its absence is never ambiguous ("did this
            # run cleanly, or not run at all?" was a real debugging dead
            # end). "No warnings or errors" is itself useful information.
            log.save_excel(os.path.join(job_folder, "generation_log.xlsx"))
            zf.write(os.path.join(job_folder, "generation_log.xlsx"), arcname="generation_log.xlsx")

        _final_download_name = _build_download_zip_name("signal_circuits.zip", hut_name, section)
        os.rename(zip_path, os.path.join(job_folder, _final_download_name))
        _set_status(
            job_id,
            state="done",
            progress=100,
            download_name=_final_download_name,
            hut_name=hut_name,
            section=section,
            sheet_count=len(generated_files),
        )

    except OdaConversionError as e:
        _set_status(job_id, state="error", error=str(e))
    except Exception as e:
        _set_status(job_id, state="error", error=str(e))


def _read_track_rows(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["TRACK"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    headers = rows[0]
    data_rows = []
    for r in rows[1:]:
        row_dict = dict(zip(headers, r))
        if row_dict.get("TRACK NAME") in (None, ""):
            continue
        data_rows.append(row_dict)
    return data_rows


def _run_track_generation_job(job_id, xlsx_path, template_path, filename_pattern, relay_position_enabled=False, hut_name="", section="", script_type="both"):
    job_folder = os.path.join(JOBS_DIR, job_id)
    dxf_folder = os.path.join(job_folder, "dxf")
    dwg_folder = os.path.join(job_folder, "dwg")
    os.makedirs(dxf_folder, exist_ok=True)
    os.makedirs(dwg_folder, exist_ok=True)

    try:
        if template_path.lower().endswith(".dwg"):
            _set_status(job_id, state="preparing_template", progress=2)
            template_path = convert_single_file_to_dxf(template_path)

        _set_status(job_id, state="reading_excel", progress=5)
        rows = _read_track_rows(xlsx_path)
        if not rows:
            raise ValueError("No track rows found in TRACK sheet")

        start_sheet_number = get_track_start_sheet_number(xlsx_path)
        next_circuit_start = get_next_circuit_after_track(xlsx_path)

        _set_status(job_id, state="generating_dxf", progress=10, total_signals=len(rows))

        try:
            sheets, _ = generate_all_track_sheets(
                rows, start_sheet_number=start_sheet_number, border_template_path=template_path,
                hut_name_input=hut_name,
            )
        except Exception as row_err:
            _set_status(job_id, state="error", error=str(row_err))
            return

        _set_status(job_id, progress=60, signals_done=len(rows))

        if sheets:
            _, last_doc, _ = sheets[-1]
            set_cont_value(last_doc, next_circuit_start)

        named = []
        for out_name, doc, sht in sheets:
            name_no_ext, ext = os.path.splitext(out_name)
            final_name = _build_output_filename(name_no_ext, sht, filename_pattern) + ext
            named.append((final_name, doc))

        log = GenerationLog()
        _apply_relay_position_check(named, xlsx_path, relay_position_enabled, log)
        clear_placeholder_values(named)
        apply_filename_attribute(named)
        apply_hut_name_keyword(named, hut_name)
        if _stop_if_vital_errors(job_id, job_folder, log):
            return

        generated_files = []
        for final_name, doc in named:
            out_path = os.path.join(dxf_folder, final_name)
            doc.saveas(out_path)
            generated_files.append(final_name)

        _set_status(job_id, state="converting_dwg", progress=65)
        convert_folder_to_dwg(dxf_folder, dwg_folder)

        try:
            if script_type in ("arescad", "both"):
                scr_text = attsync_core.generate_attsync_scr(dxf_folder, target="ares")
                with open(os.path.join(dwg_folder, "attsync.scr"), "w", encoding="utf-8") as f:
                    f.write(scr_text)
            if script_type in ("autocad", "both"):
                scr_text_autocad = attsync_core.generate_attsync_scr(dxf_folder, target="autocad")
                with open(os.path.join(dwg_folder, "attsync_autocad.scr"), "w", encoding="utf-8") as f:
                    f.write(scr_text_autocad)
        except Exception:
            pass

        _set_status(job_id, state="zipping", progress=90)
        zip_path = os.path.join(job_folder, "track_circuits.zip")
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            _add_source_files_to_zip(zf, job_folder)
            for fname in os.listdir(dwg_folder):
                zf.write(os.path.join(dwg_folder, fname), arcname=fname)
            # CONFIRMED: always include the log, not just when something
            # went wrong - so its absence is never ambiguous ("did this
            # run cleanly, or not run at all?" was a real debugging dead
            # end). "No warnings or errors" is itself useful information.
            log.save_excel(os.path.join(job_folder, "generation_log.xlsx"))
            zf.write(os.path.join(job_folder, "generation_log.xlsx"), arcname="generation_log.xlsx")

        _final_download_name = _build_download_zip_name("track_circuits.zip", hut_name, section)
        os.rename(zip_path, os.path.join(job_folder, _final_download_name))
        _set_status(
            job_id,
            state="done",
            progress=100,
            download_name=_final_download_name,
            hut_name=hut_name,
            section=section,
            sheet_count=len(generated_files),
        )

    except OdaConversionError as e:
        _set_status(job_id, state="error", error=str(e))
    except Exception as e:
        _set_status(job_id, state="error", error=str(e))


def _read_ar_azr_rows(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["AR & AZR"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    headers = rows[0]
    data_rows = []
    for r in rows[1:]:
        row_dict = dict(zip(headers, r))
        if row_dict.get("AR/AZR") in (None, ""):
            continue
        data_rows.append(row_dict)
    return data_rows


def _run_ar_azr_generation_job(job_id, xlsx_path, template_path, filename_pattern, relay_position_enabled=False, hut_name="", section="", script_type="both"):
    job_folder = os.path.join(JOBS_DIR, job_id)
    dxf_folder = os.path.join(job_folder, "dxf")
    dwg_folder = os.path.join(job_folder, "dwg")
    os.makedirs(dxf_folder, exist_ok=True)
    os.makedirs(dwg_folder, exist_ok=True)

    try:
        if template_path.lower().endswith(".dwg"):
            _set_status(job_id, state="preparing_template", progress=2)
            template_path = convert_single_file_to_dxf(template_path)

        _set_status(job_id, state="reading_excel", progress=5)
        rows = _read_ar_azr_rows(xlsx_path)
        if not rows:
            raise ValueError("No rows found in AR & AZR sheet")

        start_sheet_number = get_ar_azr_start_sheet_number(xlsx_path)
        next_circuit_start = get_next_circuit_after_ar_azr(xlsx_path)
        contact_groups_h3 = get_ar_azr_contact_groups_h3(xlsx_path)
        contact_groups_g3 = get_ar_azr_contact_groups_g3(xlsx_path)
        positions = read_ar_azr_position_config(xlsx_path)

        _set_status(job_id, state="generating_dxf", progress=10, total_signals=len(rows))

        try:
            name2_whitelist = read_name2_whitelist(xlsx_path)
            sheets, _ = generate_all_ar_azr_sheets(
                rows, contact_groups_h3, contact_groups_g3, start_sheet_number=start_sheet_number,
                border_template_path=template_path, positions=positions, name2_whitelist=name2_whitelist,
            )
        except Exception as row_err:
            _set_status(job_id, state="error", error=str(row_err))
            return

        _set_status(job_id, progress=60, signals_done=len(rows))

        if sheets:
            _, last_doc, _ = sheets[-1]
            set_cont_value(last_doc, next_circuit_start)

        named = []
        for out_name, doc, sht in sheets:
            name_no_ext, ext = os.path.splitext(out_name)
            final_name = _build_output_filename(name_no_ext, sht, filename_pattern) + ext
            named.append((final_name, doc))

        log = GenerationLog()
        _apply_relay_position_check(named, xlsx_path, relay_position_enabled, log)
        clear_placeholder_values(named)
        apply_filename_attribute(named)
        apply_hut_name_keyword(named, hut_name)
        if _stop_if_vital_errors(job_id, job_folder, log):
            return

        generated_files = []
        for final_name, doc in named:
            out_path = os.path.join(dxf_folder, final_name)
            doc.saveas(out_path)
            generated_files.append(final_name)

        _set_status(job_id, state="converting_dwg", progress=65)
        convert_folder_to_dwg(dxf_folder, dwg_folder)

        try:
            if script_type in ("arescad", "both"):
                scr_text = attsync_core.generate_attsync_scr(dxf_folder, target="ares")
                with open(os.path.join(dwg_folder, "attsync.scr"), "w", encoding="utf-8") as f:
                    f.write(scr_text)
            if script_type in ("autocad", "both"):
                scr_text_autocad = attsync_core.generate_attsync_scr(dxf_folder, target="autocad")
                with open(os.path.join(dwg_folder, "attsync_autocad.scr"), "w", encoding="utf-8") as f:
                    f.write(scr_text_autocad)
        except Exception:
            pass

        _set_status(job_id, state="zipping", progress=90)
        zip_path = os.path.join(job_folder, "ar_azr_circuits.zip")
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            _add_source_files_to_zip(zf, job_folder)
            for fname in os.listdir(dwg_folder):
                zf.write(os.path.join(dwg_folder, fname), arcname=fname)
            # CONFIRMED: always include the log, not just when something
            # went wrong - so its absence is never ambiguous ("did this
            # run cleanly, or not run at all?" was a real debugging dead
            # end). "No warnings or errors" is itself useful information.
            log.save_excel(os.path.join(job_folder, "generation_log.xlsx"))
            zf.write(os.path.join(job_folder, "generation_log.xlsx"), arcname="generation_log.xlsx")

        _final_download_name = _build_download_zip_name("ar_azr_circuits.zip", hut_name, section)
        os.rename(zip_path, os.path.join(job_folder, _final_download_name))
        _set_status(
            job_id,
            state="done",
            progress=100,
            download_name=_final_download_name,
            hut_name=hut_name,
            section=section,
            sheet_count=len(generated_files),
        )

    except OdaConversionError as e:
        _set_status(job_id, state="error", error=str(e))
    except Exception as e:
        _set_status(job_id, state="error", error=str(e))


@app.route("/login", methods=["GET", "POST"])
def login_page():
    if request.method == "GET":
        return render_template("loginpage.html")
    userid = request.form.get("userid", "").strip()
    password = request.form.get("password", "")
    user = auth_core.authenticate(userid, password)
    if not user:
        return render_template("loginpage.html", error="Invalid User ID/password, or your account is still pending admin approval.")
    session["userid"] = user["userid"]
    session["username"] = user["username"]
    session["is_admin"] = bool(user["is_admin"])
    return redirect(url_for("index"))


USERID_PREFIX = "Usts_"
EMAIL_DOMAIN = "@unified-sts.com"


def _validate_account_format(userid: str, email: str):
    """
    CONFIRMED: organizational convention - every account's User ID must
    start with "Usts_" and every email must end with "@unified-sts.com".
    Returns an error message if either is violated, or None if both are
    fine. Checked at signup (so no account can ever violate this) and
    again at forgot-password (as a fast format check before even
    touching the database).
    """
    if not userid.startswith(USERID_PREFIX):
        return f"User ID must start with '{USERID_PREFIX}'."
    if not email.lower().endswith(EMAIL_DOMAIN):
        return f"Email must end with '{EMAIL_DOMAIN}'."
    return None


@app.route("/signup", methods=["GET", "POST"])
def signup_page():
    if request.method == "GET":
        return render_template("signup.html")
    username = request.form.get("username", "").strip()
    userid = request.form.get("userid", "").strip()
    email = request.form.get("email", "").strip()
    password = request.form.get("password", "")
    if not (username and userid and email and password):
        return render_template("signup.html", error="All fields are required.")
    format_error = _validate_account_format(userid, email)
    if format_error:
        return render_template("signup.html", error=format_error)
    try:
        code = auth_core.create_pending_verification(username, userid, email, password)
    except ValueError as e:
        return render_template("signup.html", error=str(e))
    try:
        email_core.send_verification_code(email, code, username)
    except email_core.EmailNotConfigured as e:
        return render_template("signup.html", error=str(e))
    except Exception as e:
        return render_template("signup.html", error=f"Could not send verification email: {e}")
    return redirect(url_for("verify_email_page", userid=userid))


@app.route("/verify-email", methods=["GET", "POST"])
def verify_email_page():
    if request.method == "GET":
        return render_template("verify_email.html", userid=request.args.get("userid", ""))
    userid = request.form.get("userid", "").strip()
    code = request.form.get("code", "").strip()
    if not (userid and code):
        return render_template("verify_email.html", userid=userid, error="Enter the code from your email.")
    if auth_core.verify_email_code(userid, code):
        return render_template("verify_email.html", userid=userid, message="Email verified! Your account is now active - you can log in.")
    return render_template("verify_email.html", userid=userid, error="Incorrect or expired code. Please try again.")


@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password_page():
    if request.method == "GET":
        return render_template("forgot_password.html")
    username = request.form.get("username", "").strip()
    userid = request.form.get("userid", "").strip()
    email = request.form.get("email", "").strip()
    new_password = request.form.get("password", "")
    confirm_password = request.form.get("confirm_password", "")
    if not (username and userid and email and new_password and confirm_password):
        return render_template("forgot_password.html", error="All fields are required.")
    if new_password != confirm_password:
        return render_template("forgot_password.html", error="New password and confirmation don't match.")
    format_error = _validate_account_format(userid, email)
    if format_error:
        return render_template("forgot_password.html", error=format_error)
    try:
        code, sent_to_email, sent_to_username = auth_core.create_password_reset_code(userid, username, email, new_password)
    except ValueError as e:
        return render_template("forgot_password.html", error=str(e))
    try:
        email_core.send_verification_code(sent_to_email, code, sent_to_username)
    except email_core.EmailNotConfigured as e:
        return render_template("forgot_password.html", error=str(e))
    except Exception as e:
        return render_template("forgot_password.html", error=f"Could not send verification email: {e}")
    return redirect(url_for("verify_password_reset_page", userid=userid))


@app.route("/verify-password-reset", methods=["GET", "POST"])
def verify_password_reset_page():
    if request.method == "GET":
        return render_template("verify_password_reset.html", userid=request.args.get("userid", ""))
    userid = request.form.get("userid", "").strip()
    code = request.form.get("code", "").strip()
    if not (userid and code):
        return render_template("verify_password_reset.html", userid=userid, error="Enter the code from your email.")
    if auth_core.verify_password_reset_code(userid, code):
        return render_template(
            "verify_password_reset.html",
            userid=userid,
            message="Email verified. Password reset requested - an admin must approve it before your new password takes effect.",
        )
    return render_template("verify_password_reset.html", userid=userid, error="Incorrect or expired code. Please try again.")


@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password_page():
    if request.method == "GET":
        return render_template("change_password.html")
    current_password = request.form.get("current_password", "")
    new_password = request.form.get("new_password", "")
    confirm_new_password = request.form.get("confirm_new_password", "")
    if not (current_password and new_password and confirm_new_password):
        return render_template("change_password.html", error="All fields are required.")
    if new_password != confirm_new_password:
        return render_template("change_password.html", error="New password and confirmation don't match.")
    if auth_core.change_password(session["userid"], current_password, new_password):
        return render_template("change_password.html", message="Password changed successfully.")
    return render_template("change_password.html", error="Current password is incorrect.")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login_page"))


@app.route("/admin")
@admin_required
def admin_panel():
    pending_resets = auth_core.get_pending_password_resets()
    all_history = auth_core.get_all_job_history()
    grouped_all_history = _group_job_history(all_history)
    all_users = auth_core.get_all_users()
    online_userids = auth_core.get_online_userids()
    return render_template(
        "admin_panel.html",
        pending_resets=pending_resets,
        grouped_all_history=grouped_all_history,
        all_users=all_users,
        online_userids=online_userids,
    )


@app.route("/api/admin/pause-user/<int:user_id>", methods=["POST"])
@admin_required
def api_pause_user(user_id):
    auth_core.pause_user(user_id)
    return redirect(url_for("admin_panel"))


@app.route("/api/admin/unpause-user/<int:user_id>", methods=["POST"])
@admin_required
def api_unpause_user(user_id):
    auth_core.unpause_user(user_id)
    return redirect(url_for("admin_panel"))


@app.route("/api/admin/delete-user/<int:user_id>", methods=["POST"])
@admin_required
def api_delete_user(user_id):
    auth_core.delete_user(user_id)
    return redirect(url_for("admin_panel"))


@app.route("/api/admin/export-excel")
@admin_required
def api_export_excel():
    export_path = os.path.join(JOBS_DIR, "msdac_users_export.xlsx")
    auth_core.export_all_data_to_excel(export_path)
    return send_from_directory(JOBS_DIR, "msdac_users_export.xlsx", as_attachment=True)


@app.route("/api/admin/approve-reset/<int:request_id>", methods=["POST"])
@admin_required
def api_approve_reset(request_id):
    auth_core.approve_password_reset(request_id)
    return redirect(url_for("admin_panel"))


@app.route("/api/admin/reject-reset/<int:request_id>", methods=["POST"])
@admin_required
def api_reject_reset(request_id):
    auth_core.reject_password_reset(request_id)
    return redirect(url_for("admin_panel"))


@app.route("/history")
@login_required
def history_page():
    my_history = auth_core.get_job_history_for_user(session["userid"])
    grouped_history = _group_job_history(my_history)
    return render_template("history.html", grouped_history=grouped_history, is_admin=session.get("is_admin", False))


@app.route("/")
@login_required
def index():
    return render_template("msdac.html")


@app.route("/retemplate")
@login_required
def retemplate_page():
    return render_template("retemplate.html")


@app.route("/api/upload", methods=["POST"])
@login_required
def upload():
    if "file" not in request.files:
        return jsonify({"error": "No Excel file uploaded"}), 400
    if "template" not in request.files:
        return jsonify({"error": "No border/title-block template uploaded"}), 400

    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "Please upload an .xlsx or .xlsm file"}), 400

    tmpl = request.files["template"]
    if not tmpl.filename.lower().endswith((".dxf", ".dwg")):
        return jsonify({"error": "Template must be a .dxf or .dwg file"}), 400

    job_id = str(uuid.uuid4())
    job_folder = os.path.join(JOBS_DIR, job_id)
    os.makedirs(job_folder, exist_ok=True)

    xlsx_path = os.path.join(job_folder, "source.xlsx")
    f.save(xlsx_path)

    template_ext = os.path.splitext(tmpl.filename)[1].lower()
    template_path = os.path.join(job_folder, f"border_template{template_ext}")
    tmpl.save(template_path)

    filename_pattern = request.form.get("filename_pattern", "")
    hut_name = request.form.get("hut_name", "")
    section = request.form.get("section", "")
    relay_position_enabled = request.form.get("relay_position", "0") == "1"
    script_type = request.form.get("script_type", "both")

    with jobs_lock:
        jobs[job_id] = {"state": "queued", "progress": 0, "userid": session.get("userid"), "job_type": "SIGNAL"}

    executor.submit(_run_generation_job, job_id, xlsx_path, template_path, filename_pattern, relay_position_enabled, hut_name, section, script_type)

    return jsonify({"job_id": job_id})


@app.route("/api/upload-track", methods=["POST"])
@login_required
def upload_track():
    if "file" not in request.files:
        return jsonify({"error": "No Excel file uploaded"}), 400
    if "template" not in request.files:
        return jsonify({"error": "No border/title-block template uploaded"}), 400

    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "Please upload an .xlsx or .xlsm file"}), 400

    tmpl = request.files["template"]
    if not tmpl.filename.lower().endswith((".dxf", ".dwg")):
        return jsonify({"error": "Template must be a .dxf or .dwg file"}), 400

    job_id = str(uuid.uuid4())
    job_folder = os.path.join(JOBS_DIR, job_id)
    os.makedirs(job_folder, exist_ok=True)

    xlsx_path = os.path.join(job_folder, "source.xlsx")
    f.save(xlsx_path)

    template_ext = os.path.splitext(tmpl.filename)[1].lower()
    template_path = os.path.join(job_folder, f"border_template{template_ext}")
    tmpl.save(template_path)

    filename_pattern = request.form.get("filename_pattern", "")
    hut_name = request.form.get("hut_name", "")
    section = request.form.get("section", "")
    relay_position_enabled = request.form.get("relay_position", "0") == "1"

    with jobs_lock:
        jobs[job_id] = {"state": "queued", "progress": 0, "userid": session.get("userid"), "job_type": "TRACK"}

    executor.submit(_run_track_generation_job, job_id, xlsx_path, template_path, filename_pattern, relay_position_enabled, hut_name, section, request.form.get("script_type", "both"))

    return jsonify({"job_id": job_id})


@app.route("/api/upload-ar-azr", methods=["POST"])
@login_required
def upload_ar_azr():
    if "file" not in request.files:
        return jsonify({"error": "No Excel file uploaded"}), 400
    if "template" not in request.files:
        return jsonify({"error": "No border/title-block template uploaded"}), 400

    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "Please upload an .xlsx or .xlsm file"}), 400

    tmpl = request.files["template"]
    if not tmpl.filename.lower().endswith((".dxf", ".dwg")):
        return jsonify({"error": "Template must be a .dxf or .dwg file"}), 400

    job_id = str(uuid.uuid4())
    job_folder = os.path.join(JOBS_DIR, job_id)
    os.makedirs(job_folder, exist_ok=True)

    xlsx_path = os.path.join(job_folder, "source.xlsx")
    f.save(xlsx_path)

    template_ext = os.path.splitext(tmpl.filename)[1].lower()
    template_path = os.path.join(job_folder, f"border_template{template_ext}")
    tmpl.save(template_path)

    filename_pattern = request.form.get("filename_pattern", "")
    hut_name = request.form.get("hut_name", "")
    section = request.form.get("section", "")
    relay_position_enabled = request.form.get("relay_position", "0") == "1"

    with jobs_lock:
        jobs[job_id] = {"state": "queued", "progress": 0, "userid": session.get("userid"), "job_type": "AR & AZR"}

    executor.submit(_run_ar_azr_generation_job, job_id, xlsx_path, template_path, filename_pattern, relay_position_enabled, hut_name, section, request.form.get("script_type", "both"))

    return jsonify({"job_id": job_id})


def _run_relay_rack_generation_job(job_id, xlsx_path, template_path, filename_pattern, hut_name="", section="", script_type="both"):
    job_folder = os.path.join(JOBS_DIR, job_id)
    dxf_folder = os.path.join(job_folder, "dxf")
    dwg_folder = os.path.join(job_folder, "dwg")
    os.makedirs(dxf_folder, exist_ok=True)
    os.makedirs(dwg_folder, exist_ok=True)

    try:
        if template_path.lower().endswith(".dwg"):
            _set_status(job_id, state="preparing_template", progress=2)
            template_path = convert_single_file_to_dxf(template_path)

        _set_status(job_id, state="reading_excel", progress=10)
        positions = read_relay_rack_position_config(xlsx_path)

        _set_status(job_id, state="generating_dxf", progress=30)
        try:
            results = generate_all_relay_rack_drawings(xlsx_path, positions, template_path)
        except Exception as e:
            _set_status(job_id, state="error", error=str(e))
            return

        _set_status(job_id, progress=60, signals_done=len(results))

        for out_name, doc, sht in results:
            name_no_ext, ext = os.path.splitext(out_name)
            final_name = _build_output_filename(name_no_ext, sht, filename_pattern) + ext
            out_path = os.path.join(dxf_folder, final_name)
            doc.saveas(out_path)

        _set_status(job_id, state="converting_dwg", progress=65)
        convert_folder_to_dwg(dxf_folder, dwg_folder)

        try:
            if script_type in ("arescad", "both"):
                scr_text = attsync_core.generate_attsync_scr(dxf_folder, target="ares")
                with open(os.path.join(dwg_folder, "attsync.scr"), "w", encoding="utf-8") as f:
                    f.write(scr_text)
            if script_type in ("autocad", "both"):
                scr_text_autocad = attsync_core.generate_attsync_scr(dxf_folder, target="autocad")
                with open(os.path.join(dwg_folder, "attsync_autocad.scr"), "w", encoding="utf-8") as f:
                    f.write(scr_text_autocad)
        except Exception:
            pass

        _set_status(job_id, state="zipping", progress=90)
        zip_path = os.path.join(job_folder, "relay_rack.zip")
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            _add_source_files_to_zip(zf, job_folder)
            for fname in os.listdir(dwg_folder):
                zf.write(os.path.join(dwg_folder, fname), arcname=fname)

        _final_download_name = _build_download_zip_name("relay_rack.zip", hut_name, section)
        os.rename(zip_path, os.path.join(job_folder, _final_download_name))
        _set_status(
            job_id,
            state="done",
            progress=100,
            download_name=_final_download_name,
            hut_name=hut_name,
            section=section,
            sheet_count=len(results),
        )

    except OdaConversionError as e:
        _set_status(job_id, state="error", error=str(e))
    except Exception as e:
        _set_status(job_id, state="error", error=str(e))


@app.route("/api/upload-relay-rack", methods=["POST"])
@login_required
def upload_relay_rack():
    if "file" not in request.files:
        return jsonify({"error": "No Excel file uploaded"}), 400
    if "template" not in request.files:
        return jsonify({"error": "No border/title-block template uploaded"}), 400

    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "Please upload an .xlsx or .xlsm file"}), 400

    tmpl = request.files["template"]
    if not tmpl.filename.lower().endswith((".dxf", ".dwg")):
        return jsonify({"error": "Template must be a .dxf or .dwg file"}), 400

    job_id = str(uuid.uuid4())
    job_folder = os.path.join(JOBS_DIR, job_id)
    os.makedirs(job_folder, exist_ok=True)

    xlsx_path = os.path.join(job_folder, "source.xlsx")
    f.save(xlsx_path)

    template_ext = os.path.splitext(tmpl.filename)[1].lower()
    template_path = os.path.join(job_folder, f"border_template{template_ext}")
    tmpl.save(template_path)

    filename_pattern = request.form.get("filename_pattern", "")
    hut_name = request.form.get("hut_name", "")
    section = request.form.get("section", "")

    with jobs_lock:
        jobs[job_id] = {"state": "queued", "progress": 0, "userid": session.get("userid"), "job_type": "RELAY RACK"}

    executor.submit(_run_relay_rack_generation_job, job_id, xlsx_path, template_path, filename_pattern, hut_name, section, request.form.get("script_type", "both"))

    return jsonify({"job_id": job_id})


def _run_communication_generation_job(job_id, xlsx_path, template_path, filename_pattern, relay_position_enabled=False, hut_name="", section="", script_type="both"):
    job_folder = os.path.join(JOBS_DIR, job_id)
    dxf_folder = os.path.join(job_folder, "dxf")
    dwg_folder = os.path.join(job_folder, "dwg")
    os.makedirs(dxf_folder, exist_ok=True)
    os.makedirs(dwg_folder, exist_ok=True)

    try:
        if template_path.lower().endswith(".dwg"):
            _set_status(job_id, state="preparing_template", progress=2)
            template_path = convert_single_file_to_dxf(template_path)

        _set_status(job_id, state="reading_excel", progress=5)
        start_sheet_number = get_communication_start_sheet_number(xlsx_path)
        next_circuit_start = get_next_circuit_after_communication(xlsx_path)

        _set_status(job_id, state="generating_dxf", progress=10)
        try:
            results, _ = generate_all_communication_sheets(xlsx_path, start_sheet_number, template_path)
        except Exception as e:
            _set_status(job_id, state="error", error=str(e))
            return

        _set_status(job_id, progress=60, signals_done=len(results), total_signals=len(results))

        if results and next_circuit_start:
            _, last_doc, _ = results[-1]
            set_cont_value(last_doc, next_circuit_start)

        named = []
        for out_name, doc, sht in results:
            name_no_ext, ext = os.path.splitext(out_name)
            final_name = _build_output_filename(name_no_ext, sht, filename_pattern) + ext
            named.append((final_name, doc))

        log = GenerationLog()
        # CONFIRMED FIX: Communication's own Relay Position check needs
        # filename_to_circuit_type to say "COMMUNICATION" specifically,
        # so its LOC1/LOC2 location-skip actually activates (a relay
        # from a different hut must not be validated against THIS
        # workbook's own Relay Rack). Without this, circuit_type falls
        # back to the raw filename, the "COMMUNICATION" check never
        # matches, and every relay gets validated unconditionally -
        # wrongly failing genuinely cross-location relays. Every other
        # individual endpoint is unaffected (no such circuit-type-
        # conditional check exists for them), so this fix is scoped to
        # Communication only.
        communication_filename_to_circuit_type = {final_name: "COMMUNICATION" for final_name, _doc in named}
        _apply_relay_position_check(named, xlsx_path, relay_position_enabled, log, communication_filename_to_circuit_type)
        clear_placeholder_values(named)
        apply_filename_attribute(named)
        apply_hut_name_keyword(named, hut_name)
        if _stop_if_vital_errors(job_id, job_folder, log):
            return

        for final_name, doc in named:
            out_path = os.path.join(dxf_folder, final_name)
            doc.saveas(out_path)

        _set_status(job_id, state="converting_dwg", progress=65)
        convert_folder_to_dwg(dxf_folder, dwg_folder)

        try:
            if script_type in ("arescad", "both"):
                scr_text = attsync_core.generate_attsync_scr(dxf_folder, target="ares")
                with open(os.path.join(dwg_folder, "attsync.scr"), "w", encoding="utf-8") as f:
                    f.write(scr_text)
            if script_type in ("autocad", "both"):
                scr_text_autocad = attsync_core.generate_attsync_scr(dxf_folder, target="autocad")
                with open(os.path.join(dwg_folder, "attsync_autocad.scr"), "w", encoding="utf-8") as f:
                    f.write(scr_text_autocad)
        except Exception:
            pass

        _set_status(job_id, state="zipping", progress=90)
        zip_path = os.path.join(job_folder, "communication_circuits.zip")
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            _add_source_files_to_zip(zf, job_folder)
            for fname in os.listdir(dwg_folder):
                zf.write(os.path.join(dwg_folder, fname), arcname=fname)
            # CONFIRMED: always include the log, not just when something
            # went wrong - so its absence is never ambiguous ("did this
            # run cleanly, or not run at all?" was a real debugging dead
            # end). "No warnings or errors" is itself useful information.
            log.save_excel(os.path.join(job_folder, "generation_log.xlsx"))
            zf.write(os.path.join(job_folder, "generation_log.xlsx"), arcname="generation_log.xlsx")

        _final_download_name = _build_download_zip_name("communication_circuits.zip", hut_name, section)
        os.rename(zip_path, os.path.join(job_folder, _final_download_name))
        _set_status(
            job_id,
            state="done",
            progress=100,
            download_name=_final_download_name,
            hut_name=hut_name,
            section=section,
            sheet_count=len(results),
        )

    except OdaConversionError as e:
        _set_status(job_id, state="error", error=str(e))
    except Exception as e:
        _set_status(job_id, state="error", error=str(e))


@app.route("/api/upload-communication", methods=["POST"])
@login_required
def upload_communication():
    if "file" not in request.files:
        return jsonify({"error": "No Excel file uploaded"}), 400
    if "template" not in request.files:
        return jsonify({"error": "No border/title-block template uploaded"}), 400

    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "Please upload an .xlsx or .xlsm file"}), 400

    tmpl = request.files["template"]
    if not tmpl.filename.lower().endswith((".dxf", ".dwg")):
        return jsonify({"error": "Template must be a .dxf or .dwg file"}), 400

    job_id = str(uuid.uuid4())
    job_folder = os.path.join(JOBS_DIR, job_id)
    os.makedirs(job_folder, exist_ok=True)

    xlsx_path = os.path.join(job_folder, "source.xlsx")
    f.save(xlsx_path)

    template_ext = os.path.splitext(tmpl.filename)[1].lower()
    template_path = os.path.join(job_folder, f"border_template{template_ext}")
    tmpl.save(template_path)

    filename_pattern = request.form.get("filename_pattern", "")
    hut_name = request.form.get("hut_name", "")
    section = request.form.get("section", "")
    relay_position_enabled = request.form.get("relay_position", "0") == "1"

    with jobs_lock:
        jobs[job_id] = {"state": "queued", "progress": 0, "userid": session.get("userid"), "job_type": "COMMUNICATION"}

    executor.submit(_run_communication_generation_job, job_id, xlsx_path, template_path, filename_pattern, relay_position_enabled, hut_name, section, request.form.get("script_type", "both"))

    return jsonify({"job_id": job_id})


def _run_sdf_generation_job(job_id, xlsx_path, template_path, filename_pattern, relay_position_enabled=False, hut_name="", section="", script_type="both"):
    job_folder = os.path.join(JOBS_DIR, job_id)
    dxf_folder = os.path.join(job_folder, "dxf")
    dwg_folder = os.path.join(job_folder, "dwg")
    os.makedirs(dxf_folder, exist_ok=True)
    os.makedirs(dwg_folder, exist_ok=True)

    try:
        if template_path.lower().endswith(".dwg"):
            _set_status(job_id, state="preparing_template", progress=2)
            template_path = convert_single_file_to_dxf(template_path)

        _set_status(job_id, state="reading_excel", progress=5)
        start_sheet_number = get_sdf_start_sheet_number(xlsx_path)
        next_circuit_start = get_next_circuit_after_sdf(xlsx_path)

        _set_status(job_id, state="generating_dxf", progress=10)
        try:
            results, _ = generate_all_sdf_sheets(xlsx_path, start_sheet_number, template_path)
        except Exception as e:
            _set_status(job_id, state="error", error=str(e))
            return

        _set_status(job_id, progress=60, signals_done=len(results), total_signals=len(results))

        if results and next_circuit_start:
            _, last_doc, _ = results[-1]
            set_cont_value(last_doc, next_circuit_start)

        named = []
        for out_name, doc, sht in results:
            name_no_ext, ext = os.path.splitext(out_name)
            final_name = _build_output_filename(name_no_ext, sht, filename_pattern) + ext
            named.append((final_name, doc))

        log = GenerationLog()
        _apply_relay_position_check(named, xlsx_path, relay_position_enabled, log)
        clear_placeholder_values(named)
        apply_filename_attribute(named)
        apply_hut_name_keyword(named, hut_name)
        if _stop_if_vital_errors(job_id, job_folder, log):
            return

        for final_name, doc in named:
            out_path = os.path.join(dxf_folder, final_name)
            doc.saveas(out_path)

        _set_status(job_id, state="converting_dwg", progress=65)
        convert_folder_to_dwg(dxf_folder, dwg_folder)

        try:
            if script_type in ("arescad", "both"):
                scr_text = attsync_core.generate_attsync_scr(dxf_folder, target="ares")
                with open(os.path.join(dwg_folder, "attsync.scr"), "w", encoding="utf-8") as f:
                    f.write(scr_text)
            if script_type in ("autocad", "both"):
                scr_text_autocad = attsync_core.generate_attsync_scr(dxf_folder, target="autocad")
                with open(os.path.join(dwg_folder, "attsync_autocad.scr"), "w", encoding="utf-8") as f:
                    f.write(scr_text_autocad)
        except Exception:
            pass

        _set_status(job_id, state="zipping", progress=90)
        zip_path = os.path.join(job_folder, "sdf_circuits.zip")
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            _add_source_files_to_zip(zf, job_folder)
            for fname in os.listdir(dwg_folder):
                zf.write(os.path.join(dwg_folder, fname), arcname=fname)
            # CONFIRMED: always include the log, not just when something
            # went wrong - so its absence is never ambiguous ("did this
            # run cleanly, or not run at all?" was a real debugging dead
            # end). "No warnings or errors" is itself useful information.
            log.save_excel(os.path.join(job_folder, "generation_log.xlsx"))
            zf.write(os.path.join(job_folder, "generation_log.xlsx"), arcname="generation_log.xlsx")

        _final_download_name = _build_download_zip_name("sdf_circuits.zip", hut_name, section)
        os.rename(zip_path, os.path.join(job_folder, _final_download_name))
        _set_status(
            job_id,
            state="done",
            progress=100,
            download_name=_final_download_name,
            hut_name=hut_name,
            section=section,
            sheet_count=len(results),
        )

    except OdaConversionError as e:
        _set_status(job_id, state="error", error=str(e))
    except Exception as e:
        _set_status(job_id, state="error", error=str(e))


@app.route("/api/upload-sdf", methods=["POST"])
@login_required
def upload_sdf():
    if "file" not in request.files:
        return jsonify({"error": "No Excel file uploaded"}), 400
    if "template" not in request.files:
        return jsonify({"error": "No border/title-block template uploaded"}), 400

    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "Please upload an .xlsx or .xlsm file"}), 400

    tmpl = request.files["template"]
    if not tmpl.filename.lower().endswith((".dxf", ".dwg")):
        return jsonify({"error": "Template must be a .dxf or .dwg file"}), 400

    job_id = str(uuid.uuid4())
    job_folder = os.path.join(JOBS_DIR, job_id)
    os.makedirs(job_folder, exist_ok=True)

    xlsx_path = os.path.join(job_folder, "source.xlsx")
    f.save(xlsx_path)

    template_ext = os.path.splitext(tmpl.filename)[1].lower()
    template_path = os.path.join(job_folder, f"border_template{template_ext}")
    tmpl.save(template_path)

    filename_pattern = request.form.get("filename_pattern", "")
    hut_name = request.form.get("hut_name", "")
    section = request.form.get("section", "")
    relay_position_enabled = request.form.get("relay_position", "0") == "1"

    with jobs_lock:
        jobs[job_id] = {"state": "queued", "progress": 0, "userid": session.get("userid"), "job_type": "SDF"}

    executor.submit(_run_sdf_generation_job, job_id, xlsx_path, template_path, filename_pattern, relay_position_enabled, hut_name, section, request.form.get("script_type", "both"))

    return jsonify({"job_id": job_id})


def _run_datalogger_generation_job(job_id, xlsx_path, template_path, filename_pattern, relay_position_enabled=False, hut_name="", section="", script_type="both"):
    job_folder = os.path.join(JOBS_DIR, job_id)
    dxf_folder = os.path.join(job_folder, "dxf")
    dwg_folder = os.path.join(job_folder, "dwg")
    os.makedirs(dxf_folder, exist_ok=True)
    os.makedirs(dwg_folder, exist_ok=True)

    try:
        if template_path.lower().endswith(".dwg"):
            _set_status(job_id, state="preparing_template", progress=2)
            template_path = convert_single_file_to_dxf(template_path)

        _set_status(job_id, state="reading_excel", progress=5)
        start_sheet_number = get_datalogger_start_sheet_number(xlsx_path)
        next_circuit_start = get_next_circuit_after_datalogger(xlsx_path)

        _set_status(job_id, state="generating_dxf", progress=10)
        try:
            results, _ = generate_all_datalogger_sheets(xlsx_path, start_sheet_number, template_path)
        except Exception as e:
            _set_status(job_id, state="error", error=str(e))
            return

        _set_status(job_id, progress=60, signals_done=len(results), total_signals=len(results))

        if results and next_circuit_start:
            _, last_doc, _ = results[-1]
            set_cont_value(last_doc, next_circuit_start)

        named = []
        for out_name, doc, sht in results:
            name_no_ext, ext = os.path.splitext(out_name)
            final_name = _build_output_filename(name_no_ext, sht, filename_pattern) + ext
            named.append((final_name, doc))

        log = GenerationLog()
        _apply_relay_position_check(named, xlsx_path, relay_position_enabled, log)
        clear_placeholder_values(named)
        apply_filename_attribute(named)
        apply_hut_name_keyword(named, hut_name)
        if _stop_if_vital_errors(job_id, job_folder, log):
            return

        for final_name, doc in named:
            out_path = os.path.join(dxf_folder, final_name)
            doc.saveas(out_path)

        _set_status(job_id, state="converting_dwg", progress=65)
        convert_folder_to_dwg(dxf_folder, dwg_folder)

        try:
            if script_type in ("arescad", "both"):
                scr_text = attsync_core.generate_attsync_scr(dxf_folder, target="ares")
                with open(os.path.join(dwg_folder, "attsync.scr"), "w", encoding="utf-8") as f:
                    f.write(scr_text)
            if script_type in ("autocad", "both"):
                scr_text_autocad = attsync_core.generate_attsync_scr(dxf_folder, target="autocad")
                with open(os.path.join(dwg_folder, "attsync_autocad.scr"), "w", encoding="utf-8") as f:
                    f.write(scr_text_autocad)
        except Exception:
            pass

        _set_status(job_id, state="zipping", progress=90)
        zip_path = os.path.join(job_folder, "datalogger_circuits.zip")
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            _add_source_files_to_zip(zf, job_folder)
            for fname in os.listdir(dwg_folder):
                zf.write(os.path.join(dwg_folder, fname), arcname=fname)
            # CONFIRMED: always include the log, not just when something
            # went wrong - so its absence is never ambiguous ("did this
            # run cleanly, or not run at all?" was a real debugging dead
            # end). "No warnings or errors" is itself useful information.
            log.save_excel(os.path.join(job_folder, "generation_log.xlsx"))
            zf.write(os.path.join(job_folder, "generation_log.xlsx"), arcname="generation_log.xlsx")

        _final_download_name = _build_download_zip_name("datalogger_circuits.zip", hut_name, section)
        os.rename(zip_path, os.path.join(job_folder, _final_download_name))
        _set_status(
            job_id,
            state="done",
            progress=100,
            download_name=_final_download_name,
            hut_name=hut_name,
            section=section,
            sheet_count=len(results),
        )

    except OdaConversionError as e:
        _set_status(job_id, state="error", error=str(e))
    except Exception as e:
        _set_status(job_id, state="error", error=str(e))


@app.route("/api/upload-datalogger", methods=["POST"])
@login_required
def upload_datalogger():
    if "file" not in request.files:
        return jsonify({"error": "No Excel file uploaded"}), 400
    if "template" not in request.files:
        return jsonify({"error": "No border/title-block template uploaded"}), 400

    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "Please upload an .xlsx or .xlsm file"}), 400

    tmpl = request.files["template"]
    if not tmpl.filename.lower().endswith((".dxf", ".dwg")):
        return jsonify({"error": "Template must be a .dxf or .dwg file"}), 400

    job_id = str(uuid.uuid4())
    job_folder = os.path.join(JOBS_DIR, job_id)
    os.makedirs(job_folder, exist_ok=True)

    xlsx_path = os.path.join(job_folder, "source.xlsx")
    f.save(xlsx_path)

    template_ext = os.path.splitext(tmpl.filename)[1].lower()
    template_path = os.path.join(job_folder, f"border_template{template_ext}")
    tmpl.save(template_path)

    filename_pattern = request.form.get("filename_pattern", "")
    hut_name = request.form.get("hut_name", "")
    section = request.form.get("section", "")
    relay_position_enabled = request.form.get("relay_position", "0") == "1"

    with jobs_lock:
        jobs[job_id] = {"state": "queued", "progress": 0, "userid": session.get("userid"), "job_type": "DATA LOGGER"}

    executor.submit(_run_datalogger_generation_job, job_id, xlsx_path, template_path, filename_pattern, relay_position_enabled, hut_name, section, request.form.get("script_type", "both"))

    return jsonify({"job_id": job_id})


def _run_custom_circuits_generation_job(job_id, xlsx_path, zip_path, template_path, filename_pattern, relay_position_enabled=False, hut_name="", section="", script_type="both"):
    job_folder = os.path.join(JOBS_DIR, job_id)
    dxf_folder = os.path.join(job_folder, "dxf")
    dwg_folder = os.path.join(job_folder, "dwg")
    os.makedirs(dxf_folder, exist_ok=True)
    os.makedirs(dwg_folder, exist_ok=True)

    try:
        if template_path.lower().endswith(".dwg"):
            _set_status(job_id, state="preparing_template", progress=2)
            template_path = convert_single_file_to_dxf(template_path)

        _set_status(job_id, state="reading_excel", progress=5)
        start_sheet_number = get_custom_circuits_start_sheet_number(xlsx_path)
        next_circuit_start = get_next_circuit_after_custom_circuits(xlsx_path)

        _set_status(job_id, state="generating_dxf", progress=10)
        try:
            results, _ = generate_all_custom_circuit_sheets(xlsx_path, zip_path, start_sheet_number, template_path)
        except Exception as e:
            _set_status(job_id, state="error", error=str(e))
            return

        _set_status(job_id, progress=60, signals_done=len(results), total_signals=len(results))

        if results and next_circuit_start:
            _, last_doc, _ = results[-1]
            set_cont_value(last_doc, next_circuit_start)

        named = []
        for out_name, doc, sht in results:
            name_no_ext, ext = os.path.splitext(out_name)
            final_name = _build_output_filename(name_no_ext, sht, filename_pattern) + ext
            named.append((final_name, doc))

        log = GenerationLog()
        _apply_relay_position_check(named, xlsx_path, relay_position_enabled, log)
        clear_placeholder_values(named)
        apply_filename_attribute(named)
        apply_hut_name_keyword(named, hut_name)
        if _stop_if_vital_errors(job_id, job_folder, log):
            return

        for final_name, doc in named:
            out_path = os.path.join(dxf_folder, final_name)
            doc.saveas(out_path)

        _set_status(job_id, state="converting_dwg", progress=65)
        convert_folder_to_dwg(dxf_folder, dwg_folder)

        try:
            if script_type in ("arescad", "both"):
                scr_text = attsync_core.generate_attsync_scr(dxf_folder, target="ares")
                with open(os.path.join(dwg_folder, "attsync.scr"), "w", encoding="utf-8") as f:
                    f.write(scr_text)
            if script_type in ("autocad", "both"):
                scr_text_autocad = attsync_core.generate_attsync_scr(dxf_folder, target="autocad")
                with open(os.path.join(dwg_folder, "attsync_autocad.scr"), "w", encoding="utf-8") as f:
                    f.write(scr_text_autocad)
        except Exception:
            pass

        _set_status(job_id, state="zipping", progress=90)
        zip_out_path = os.path.join(job_folder, "custom_circuits.zip")
        with zipfile.ZipFile(zip_out_path, "w", zipfile.ZIP_DEFLATED) as zf:
            _add_source_files_to_zip(zf, job_folder)
            for fname in os.listdir(dwg_folder):
                zf.write(os.path.join(dwg_folder, fname), arcname=fname)
            # CONFIRMED: always include the log, not just when something
            # went wrong - so its absence is never ambiguous ("did this
            # run cleanly, or not run at all?" was a real debugging dead
            # end). "No warnings or errors" is itself useful information.
            log.save_excel(os.path.join(job_folder, "generation_log.xlsx"))
            zf.write(os.path.join(job_folder, "generation_log.xlsx"), arcname="generation_log.xlsx")

        _final_download_name = _build_download_zip_name("custom_circuits.zip", hut_name, section)
        os.rename(zip_path, os.path.join(job_folder, _final_download_name))
        _set_status(
            job_id,
            state="done",
            progress=100,
            download_name=_final_download_name,
            hut_name=hut_name,
            section=section,
            sheet_count=len(results),
        )

    except OdaConversionError as e:
        _set_status(job_id, state="error", error=str(e))
    except Exception as e:
        _set_status(job_id, state="error", error=str(e))


@app.route("/api/upload-custom-circuits", methods=["POST"])
@login_required
def upload_custom_circuits():
    if "file" not in request.files:
        return jsonify({"error": "No Excel file uploaded"}), 400
    if "template" not in request.files:
        return jsonify({"error": "No border/title-block template uploaded"}), 400
    if "circuits_zip" not in request.files:
        return jsonify({"error": "No custom circuits zip uploaded"}), 400

    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "Please upload an .xlsx or .xlsm file"}), 400

    tmpl = request.files["template"]
    if not tmpl.filename.lower().endswith((".dxf", ".dwg")):
        return jsonify({"error": "Template must be a .dxf or .dwg file"}), 400

    circuits_zip = request.files["circuits_zip"]
    if not circuits_zip.filename.lower().endswith(".zip"):
        return jsonify({"error": "Custom circuits file must be a .zip"}), 400

    job_id = str(uuid.uuid4())
    job_folder = os.path.join(JOBS_DIR, job_id)
    os.makedirs(job_folder, exist_ok=True)

    xlsx_path = os.path.join(job_folder, "source.xlsx")
    f.save(xlsx_path)

    template_ext = os.path.splitext(tmpl.filename)[1].lower()
    template_path = os.path.join(job_folder, f"border_template{template_ext}")
    tmpl.save(template_path)

    zip_path = os.path.join(job_folder, "circuits.zip")
    circuits_zip.save(zip_path)

    filename_pattern = request.form.get("filename_pattern", "")
    hut_name = request.form.get("hut_name", "")
    section = request.form.get("section", "")
    relay_position_enabled = request.form.get("relay_position", "0") == "1"

    with jobs_lock:
        jobs[job_id] = {"state": "queued", "progress": 0, "userid": session.get("userid"), "job_type": "CUSTOM CIRCUITS"}

    executor.submit(_run_custom_circuits_generation_job, job_id, xlsx_path, zip_path, template_path, filename_pattern, relay_position_enabled, hut_name, section, request.form.get("script_type", "both"))

    return jsonify({"job_id": job_id})


def _run_full_ifc_generation_job(job_id, xlsx_path, template_path, zip_path, filename_pattern, relay_position_enabled=False, hut_name="", section="", script_type="both"):
    """
    Runs every circuit type in sequence, collecting all generated sheets
    into ONE combined output. Each circuit type's own generate_all_*
    function already reads its own correct starting sheet number and
    CONT-to-next-circuit value from FIELD PG.NO, so no manual chaining is
    needed here - just call each one and merge results.

    CONFIRMED behavior: if a circuit type's required Excel data is empty
    (e.g. no rows in a sheet), that circuit type is skipped gracefully
    rather than failing the whole job - not every project uses every
    circuit type. Custom Circuits only runs if a zip was provided.
    """
    job_folder = os.path.join(JOBS_DIR, job_id)
    dxf_folder = os.path.join(job_folder, "dxf")
    dwg_folder = os.path.join(job_folder, "dwg")
    os.makedirs(dxf_folder, exist_ok=True)
    os.makedirs(dwg_folder, exist_ok=True)

    skipped = []
    pending = []  # (final_name, doc)
    log = GenerationLog()
    registry = ContactRegistry(log)

    # CONFIRMED: errors should be categorized by circuit type (SIGNAL,
    # TRACK, RELAY RACK, etc.), not just a generic message. _add_results'
    # `prefix` argument was previously unused dead code - now it feeds
    # this filename -> circuit-type-label map, so Contact Analysis (and
    # anything else re-parsing saved files later) can label its errors
    # correctly even though _build_output_filename may have already
    # rewritten the actual filename per the user's own pattern.
    CIRCUIT_TYPE_LABELS = {
        "COVERSHEET": "COVER SHEET",
        "RELAYRACK": "RELAY RACK",
        "STATIONLAYOUT": "STATION LAYOUT",
        "FLOORPLAN": "FLOOR PLAN",
        "SIGNAL": "SIGNAL",
        "TRACK": "TRACK",
        "ARAZR": "AR & AZR",
        "COMM": "COMMUNICATION",
        "SDF": "SDF",
        "DL": "DATA LOGGER",
        "CUSTOM": "CUSTOM CIRCUITS",
    }
    filename_to_circuit_type = {}
    index_source_entries = []  # (title_text, sht_number, final_name_no_ext) for every real sheet generated

    def _sheet_title_text(doc):
        for e in doc.modelspace():
            if e.dxftype() == "INSERT" and e.dxf.name in ("TITLE", "TITLEBLOCK"):
                for att in e.attribs:
                    if att.dxf.tag == "TITLE":
                        return (att.dxf.text or "").strip()
        return ""

    def _add_results(results, prefix):
        for out_name, doc, sht in results:
            name_no_ext, ext = os.path.splitext(out_name)
            final_name = _build_output_filename(name_no_ext, sht, filename_pattern) + ext
            pending.append((final_name, doc))
            filename_to_circuit_type[final_name] = CIRCUIT_TYPE_LABELS.get(prefix, prefix)
            final_name_no_ext, _ext2 = os.path.splitext(final_name)
            index_source_entries.append((_sheet_title_text(doc), str(sht), final_name_no_ext))

    def _skip(message):
        skipped.append(message)
        log.warning(message)

    try:
        if template_path.lower().endswith(".dwg"):
            _set_status(job_id, state="preparing_template", progress=2)
            template_path = convert_single_file_to_dxf(template_path)

        # CONFIRMED: proactively checks whether enough repeater relay
        # instances exist in Relay Rack for every signal's LCPR
        # requirement, BEFORE any generation runs at all - stops here
        # with a clear message naming exactly what's missing, rather
        # than only discovering a shortfall via a generic "relay
        # position not found" error partway through generation.
        try:
            if not validate_lcpr_relay_availability(xlsx_path, log):
                if _stop_if_vital_errors(job_id, job_folder, log):
                    return
        except Exception as e:
            _skip(f"LCPR Availability Check: {e}")

        steps = [
            "Cover Sheet", "Station Layout", "Floor Plan", "Relay Rack", "Signal", "Track", "AR & AZR",
            "Communication", "SDF", "Data Logger", "Custom Circuits",
        ]
        total_steps = len(steps)

        # -1. Cover Sheet - reference content pick-and-placed at (0,0),
        # always exactly one sheet.
        _set_status(job_id, state="generating_dxf", progress=2, signals_done=0, total_signals=total_steps)
        try:
            results, _next_sht = generate_cover_sheet(xlsx_path, template_path)
            _add_results(results, "COVERSHEET")
        except Exception as e:
            _skip(f"Cover Sheet: {e}")

        # 0. Station Layout - blank placeholder sheets, count comes from
        # FIELD PG.NO's own SPARE SHEETS column - no real content.
        _set_status(job_id, state="generating_dxf", progress=3, signals_done=1, total_signals=total_steps)
        try:
            results, _next_sht = generate_all_station_layout_sheets(xlsx_path, template_path)
            _add_results(results, "STATIONLAYOUT")
        except Exception as e:
            _skip(f"Station Layout: {e}")

        # 0.5. Floor Plan - same convention as Station Layout, its own
        # SPARE SHEETS count from FIELD PG.NO.
        _set_status(job_id, state="generating_dxf", progress=4, signals_done=2, total_signals=total_steps)
        try:
            results, _next_sht = generate_all_floor_plan_sheets(xlsx_path, template_path)
            _add_results(results, "FLOORPLAN")
        except Exception as e:
            _skip(f"Floor Plan: {e}")

        # 1. Relay Rack - reads its own sheet numbering internally.
        _set_status(job_id, state="generating_dxf", progress=5, signals_done=3, total_signals=total_steps)
        relay_rack_final_names = []
        try:
            positions = read_relay_rack_position_config(xlsx_path)
            results = generate_all_relay_rack_drawings(xlsx_path, positions, template_path)
            before_count = len(pending)
            _add_results(results, "RELAYRACK")
            # CONFIRMED: Contact Analysis needs to tell Relay Rack sheets
            # (source of the relay roster) apart from every other sheet
            # (source of contact usage) - _build_output_filename may
            # rewrite names per the user's pattern, so track by position
            # in `pending` right here rather than guessing from filename.
            relay_rack_final_names = [name for name, _ in pending[before_count:]]
        except Exception as e:
            _skip(f"Relay Rack: {e}")

        # 2. Signal - per-row loop, own CONT-to-next-circuit fixup.
        _set_status(job_id, progress=15, signals_done=4)
        try:
            rows = _read_signal_rows(xlsx_path)
            if rows:
                start_sheet_number = get_signal_start_sheet_number(xlsx_path)
                next_circuit_start = get_next_circuit_start_sheet_number(xlsx_path)
                sheet_sets_override, title_map_override, lcpr_config = read_typical_config(xlsx_path)
                sheet_num = start_sheet_number
                signal_pending = []
                for row in rows:
                    sheets, sheet_num = generate_signal_drawing_set(
                        row, start_sheet_number=sheet_num, border_template_path=template_path,
                        sheet_sets_override=sheet_sets_override, title_map_override=title_map_override,
                        lcpr_config=lcpr_config, hut_name_input=hut_name, registry=registry,
                    )
                    for out_name, doc, sht in sheets:
                        name_no_ext, ext = os.path.splitext(out_name)
                        final_name = _build_output_filename(name_no_ext, sht, filename_pattern) + ext
                        signal_pending.append((final_name, doc))
                        filename_to_circuit_type[final_name] = "SIGNAL"
                        final_name_no_ext, _ext2 = os.path.splitext(final_name)
                        index_source_entries.append((_sheet_title_text(doc), str(sht), final_name_no_ext))
                if signal_pending and next_circuit_start:
                    set_cont_value(signal_pending[-1][1], next_circuit_start)
                pending.extend(signal_pending)
            else:
                _skip("Signal: no rows found in SIGNAL sheet")
        except Exception as e:
            _skip(f"Signal: {e}")

        # 3. Track
        _set_status(job_id, progress=25, signals_done=5)
        try:
            rows = _read_track_rows(xlsx_path)
            if rows:
                start_sheet_number = get_track_start_sheet_number(xlsx_path)
                next_circuit_start = get_next_circuit_after_track(xlsx_path)
                results, _ = generate_all_track_sheets(rows, start_sheet_number, template_path, hut_name_input=hut_name)
                if results and next_circuit_start:
                    set_cont_value(results[-1][1], next_circuit_start)
                _add_results(results, "TRACK")
            else:
                _skip("Track: no rows found in TRACK sheet")
        except Exception as e:
            _skip(f"Track: {e}")

        # 4. AR & AZR
        _set_status(job_id, progress=35, signals_done=6)
        try:
            rows = _read_ar_azr_rows(xlsx_path)
            if rows:
                start_sheet_number = get_ar_azr_start_sheet_number(xlsx_path)
                next_circuit_start = get_next_circuit_after_ar_azr(xlsx_path)
                contact_groups_h3 = get_ar_azr_contact_groups_h3(xlsx_path)
                contact_groups_g3 = get_ar_azr_contact_groups_g3(xlsx_path)
                positions = read_ar_azr_position_config(xlsx_path)
                name2_whitelist = read_name2_whitelist(xlsx_path)
                results, _ = generate_all_ar_azr_sheets(
                    rows, contact_groups_h3, contact_groups_g3, start_sheet_number, template_path, positions,
                    registry=registry, name2_whitelist=name2_whitelist,
                )
                if results and next_circuit_start:
                    set_cont_value(results[-1][1], next_circuit_start)
                _add_results(results, "ARAZR")
            else:
                _skip("AR & AZR: no rows found in AR & AZR sheet")
        except Exception as e:
            _skip(f"AR & AZR: {e}")

        # 5. Communication
        _set_status(job_id, progress=45, signals_done=7)
        try:
            start_sheet_number = get_communication_start_sheet_number(xlsx_path)
            next_circuit_start = get_next_circuit_after_communication(xlsx_path)
            results, _ = generate_all_communication_sheets(xlsx_path, start_sheet_number, template_path)
            if results and next_circuit_start:
                set_cont_value(results[-1][1], next_circuit_start)
            _add_results(results, "COMM")
        except Exception as e:
            _skip(f"Communication: {e}")

        # 6. SDF
        _set_status(job_id, progress=55, signals_done=8)
        try:
            start_sheet_number = get_sdf_start_sheet_number(xlsx_path)
            next_circuit_start = get_next_circuit_after_sdf(xlsx_path)
            results, _ = generate_all_sdf_sheets(xlsx_path, start_sheet_number, template_path, registry=registry)
            if results and next_circuit_start:
                set_cont_value(results[-1][1], next_circuit_start)
            _add_results(results, "SDF")
        except Exception as e:
            _skip(f"SDF: {e}")

        # 7. Data Logger
        _set_status(job_id, progress=62, signals_done=9)
        try:
            start_sheet_number = get_datalogger_start_sheet_number(xlsx_path)
            next_circuit_start = get_next_circuit_after_datalogger(xlsx_path)
            # CONFIRMED: when FIELD PG.NO!I3 is blank, Data Logger picks a
            # genuinely FREE contact per relay - "free" means not already
            # used by Signal/Track/AR & AZR/Communication/SDF, all of
            # which have already been generated by this point.
            used_contacts = build_used_contacts_registry([doc for _name, doc in pending])
            results, _ = generate_all_datalogger_sheets(
                xlsx_path, start_sheet_number, template_path, used_contacts, log,
            )
            if results and next_circuit_start:
                set_cont_value(results[-1][1], next_circuit_start)
            _add_results(results, "DL")
        except Exception as e:
            _skip(f"Data Logger: {e}")

        # 8. Custom Circuits - only if a zip was provided.
        _set_status(job_id, progress=68, signals_done=10)
        if zip_path:
            try:
                start_sheet_number = get_custom_circuits_start_sheet_number(xlsx_path)
                next_circuit_start = get_next_circuit_after_custom_circuits(xlsx_path)
                results, _ = generate_all_custom_circuit_sheets(xlsx_path, zip_path, start_sheet_number, template_path)
                if results and next_circuit_start:
                    set_cont_value(results[-1][1], next_circuit_start)
                _add_results(results, "CUSTOM")
            except Exception as e:
                _skip(f"Custom Circuits: {e}")
        else:
            _skip("Custom Circuits: no Custom Typicals zip uploaded")

        if not pending:
            _set_status(job_id, state="error", error="No sheets were generated at all - " + "; ".join(skipped))
            return

        # CONFIRMED FIX: LCPR allocation must run BEFORE Relay Position -
        # it's what resolves which LCPR/LCPR1/LCPR2 instance each block
        # actually belongs to (writing the real R_NAME). Relay
        # Position's own lookup needs that ALREADY resolved to find the
        # correct R(POS); running it first meant every LCPR block's
        # position got looked up using the raw, unresolved "LCPR"
        # literal regardless of which signal it really belonged to -
        # producing a wrong R(POS) without raising an error, since the
        # lookup still technically found *some* position (just not
        # accounting for the LCPR1/LCPR2 distinction). Runs regardless
        # of whether Relay Position was ticked - LCPR contact
        # allocation is its own separate feature, purely driven by
        # whether the SIGNAL sheet's own "LCPR" column is filled for a
        # given signal, not tied to the Relay Position checkbox.
        try:
            apply_lcpr_allocations(pending, xlsx_path, log)
        except Exception as e:
            _skip(f"LCPR Allocation: {e}")

        # CONFIRMED: Relay Position is OPTIONAL (only runs if the "Relay
        # Position" checkbox was ticked). Runs on the IN-MEMORY docs,
        # before anything is saved, so an unmatched contact/coil (vital
        # error) stops the job via the SAME has_errors gate below as
        # contact repetition - no separate error path needed.
        if relay_position_enabled:
            try:
                position_lookup = build_relay_position_lookup(xlsx_path, log)
                apply_relay_positions(pending, position_lookup, log, filename_to_circuit_type)
            except Exception as e:
                _skip(f"Relay Position: {e}")

        # CONFIRMED: runs regardless of whether Relay Position was ticked -
        # clears any leftover 'XXXXX'-style template placeholder that
        # never got a real value (e.g. R(POS) when Relay Position wasn't
        # run, or SDF_RELAY's S_NAME which circuit generation leaves
        # untouched by design either way).
        clear_placeholder_values(pending)
        apply_filename_attribute(pending)
        apply_hut_name_keyword(pending, hut_name)

        # CONFIRMED: contact repetition (and any other critical error) must
        # STOP generation entirely before any output is produced - not just
        # get logged while the run continues. Save an Excel copy of the log
        # so the user can see exactly what went wrong.
        if log.has_errors:
            log_xlsx_path = os.path.join(job_folder, "generation_log.xlsx")
            log.save_excel(log_xlsx_path)
            error_messages = [msg for level, msg in log.entries if level == "ERROR"]
            _set_status(
                job_id,
                state="error",
                error="Critical error(s) found - generation stopped before producing output:\n"
                + "\n".join(error_messages),
                log_errors=len(error_messages),
            )
            return

        _set_status(job_id, progress=75, signals_done=total_steps)

        for final_name, doc in pending:
            out_path = os.path.join(dxf_folder, final_name)
            doc.saveas(out_path)

        # CONFIRMED: Contact Analysis runs LAST, after everything else is
        # on disk - it re-parses those saved DXFs (Relay Rack for the
        # relay roster, every other sheet for contact usage) rather than
        # needing any Excel input of its own. Only runs if Relay Rack
        # actually produced something to build a roster from.
        if relay_rack_final_names:
            try:
                _set_status(job_id, state="generating_contact_analysis", progress=78)
                next_sht = get_contact_analysis_start_sheet_number(xlsx_path)
                ca_results = generate_all_contact_analysis_sheets(
                    dxf_folder, relay_rack_final_names, template_path, next_sht, log,
                    filename_to_circuit_type,
                    confirm_callback=lambda msg: ask_user_confirmation(job_id, msg),
                    xlsx_path=xlsx_path,
                    on_conflict=lambda relay_key, contact_code, pairs, sheet_nums, circuit_types, occurrence_choices: wait_for_contact_choice(job_id, relay_key, pairs, sheet_nums, circuit_types, occurrence_choices, contact_code),
                )
                for out_name, ca_doc, ca_sht in ca_results:
                    name_no_ext, ext = os.path.splitext(out_name)
                    final_name = _build_output_filename(name_no_ext, ca_sht, filename_pattern) + ext
                    ca_doc.saveas(os.path.join(dxf_folder, final_name))
                    pending.append((final_name, ca_doc))
                    filename_to_circuit_type[final_name] = "CONTACT ANALYSIS"
                    final_name_no_ext, _ext2 = os.path.splitext(final_name)
                    index_source_entries.append((_sheet_title_text(ca_doc), str(ca_sht), final_name_no_ext))
            except ContactAnalysisCancelled as e:
                log_xlsx_path = os.path.join(job_folder, "generation_log.xlsx")
                log.save_excel(log_xlsx_path)
                _set_status(job_id, state="cancelled", error=str(e))
                return
            except Exception as e:
                _skip(f"Contact Analysis: {e}")
        else:
            _skip("Contact Analysis: no Relay Rack sheets were generated, so there's no relay roster to chart - skipped.")

        # CONFIRMED: Contact Analysis's own vital errors (duplicate contact
        # usage, relay not found in Relay Rack) stop the job the same way
        # Relay Position's do - just later, since Contact Analysis needs
        # files already on disk to re-parse. Files already saved to
        # dxf_folder at this point are harmless leftovers since the job
        # never reaches DWG conversion/zipping, so nothing gets exposed
        # to the user.
        if _stop_if_vital_errors(job_id, job_folder, log):
            return

        # CONFIRMED: INDEX is generated LAST (needs the complete list of
        # every sheet actually produced, including Contact Analysis), but
        # its OWN sheet number is early (e.g. "00A" from FIELD PG.NO) -
        # that's fine, since nothing here depends on generation order,
        # only on each sheet's own number being correct in its filename.
        try:
            index_start_sht = get_index_start_sheet_number(xlsx_path)
            ordered_sheets = sorted(index_source_entries, key=lambda t: _sheet_number_sort_key(t[1]))
            ordered_sheets = add_index_self_listing(ordered_sheets, index_start_sht, filename_pattern)
            ordered_sheets.sort(key=lambda t: _sheet_number_sort_key(t[1]))
            index_entries = build_index_entries_with_spare_gaps(ordered_sheets, filename_pattern)
            index_results, _next_index_sht = generate_all_index_sheets(index_entries, index_start_sht, template_path)
            for out_name, index_doc, index_sht in index_results:
                name_no_ext, ext = os.path.splitext(out_name)
                final_name = _build_output_filename(name_no_ext, index_sht, filename_pattern) + ext
                index_doc.saveas(os.path.join(dxf_folder, final_name))
                pending.append((final_name, index_doc))
                filename_to_circuit_type[final_name] = "INDEX"
        except Exception as e:
            _skip(f"Index: {e}")

        _set_status(job_id, state="converting_dwg", progress=80)
        convert_folder_to_dwg(dxf_folder, dwg_folder)

        # CONFIRMED: generates the ATTSYNC .scr script(s) covering every
        # block actually used across every generated sheet, saved
        # alongside the DWG files so it can be run directly from
        # wherever the user extracts the zip. CONFIRMED FIX: respects
        # the user's script_type choice (AutoCAD/Both/ArES CAD) from the
        # frontend slider instead of always generating both regardless.
        # ARES syntax is empirically confirmed working (from Krish's own
        # tested ARES Commander tool); the AutoCAD variant is generated
        # from AutoCAD's documented command syntax but not yet verified
        # against real AutoCAD.
        try:
            if script_type in ("arescad", "both"):
                scr_text = attsync_core.generate_attsync_scr(dxf_folder, target="ares")
                scr_path = os.path.join(dwg_folder, "attsync.scr")
                with open(scr_path, "w", encoding="utf-8") as f:
                    f.write(scr_text)
            if script_type in ("autocad", "both"):
                scr_text_autocad = attsync_core.generate_attsync_scr(dxf_folder, target="autocad")
                scr_path_autocad = os.path.join(dwg_folder, "attsync_autocad.scr")
                with open(scr_path_autocad, "w", encoding="utf-8") as f:
                    f.write(scr_text_autocad)
        except Exception as e:
            _skip(f"ATTSYNC script: {e}")

        _set_status(job_id, state="zipping", progress=95)
        zip_out_path = os.path.join(job_folder, "full_ifc.zip")
        log_xlsx_path = os.path.join(job_folder, "generation_log.xlsx")
        log.save_excel(log_xlsx_path)
        with zipfile.ZipFile(zip_out_path, "w", zipfile.ZIP_DEFLATED) as zf:
            _add_source_files_to_zip(zf, job_folder)
            for fname in os.listdir(dwg_folder):
                zf.write(os.path.join(dwg_folder, fname), arcname=fname)
            # CONFIRMED: every generation run's log (errors, warnings, and
            # contact-repetition flags) gets included in the output so the
            # user can review it after the fact - both plain text and Excel.
            zf.write(log_xlsx_path, arcname="generation_log.xlsx")

        warnings_count, errors_count = log.summary_counts()
        _final_download_name = _build_download_zip_name("full_ifc.zip", hut_name, section)
        os.rename(zip_out_path, os.path.join(job_folder, _final_download_name))
        _set_status(
            job_id,
            state="done",
            progress=100,
            download_name=_final_download_name,
            hut_name=hut_name,
            section=section,
            sheet_count=len(pending),
            skipped=skipped,
            log_warnings=warnings_count,
            log_errors=errors_count,
        )

    except OdaConversionError as e:
        _set_status(job_id, state="error", error=str(e))
    except Exception as e:
        _set_status(job_id, state="error", error=str(e))


@app.route("/api/upload-ifc", methods=["POST"])
@login_required
def upload_ifc():
    if "file" not in request.files:
        return jsonify({"error": "No Excel file uploaded"}), 400
    if "template" not in request.files:
        return jsonify({"error": "No border/title-block template uploaded"}), 400

    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "Please upload an .xlsx or .xlsm file"}), 400

    tmpl = request.files["template"]
    if not tmpl.filename.lower().endswith((".dxf", ".dwg")):
        return jsonify({"error": "Template must be a .dxf or .dwg file"}), 400

    job_id = str(uuid.uuid4())
    job_folder = os.path.join(JOBS_DIR, job_id)
    os.makedirs(job_folder, exist_ok=True)

    xlsx_path = os.path.join(job_folder, "source.xlsx")
    f.save(xlsx_path)

    template_ext = os.path.splitext(tmpl.filename)[1].lower()
    template_path = os.path.join(job_folder, f"border_template{template_ext}")
    tmpl.save(template_path)

    # Custom Typicals zip is OPTIONAL for the full IFC run - Custom
    # Circuits just gets skipped gracefully if it's not provided.
    zip_path = None
    if "circuits_zip" in request.files and request.files["circuits_zip"].filename:
        circuits_zip = request.files["circuits_zip"]
        if circuits_zip.filename.lower().endswith(".zip"):
            zip_path = os.path.join(job_folder, "circuits.zip")
            circuits_zip.save(zip_path)

    filename_pattern = request.form.get("filename_pattern", "")
    hut_name = request.form.get("hut_name", "")
    section = request.form.get("section", "")
    relay_position_enabled = request.form.get("relay_position", "0") == "1"
    # CONFIRMED FIX: the frontend's script-output slider (AutoCAD/Both/
    # ArES CAD) was already sending this, but nothing on the backend
    # read it - the attsync script(s) were always generated
    # unconditionally regardless of the user's choice.
    script_type = request.form.get("script_type", "both")

    with jobs_lock:
        jobs[job_id] = {"state": "queued", "progress": 0, "userid": session.get("userid"), "job_type": "FULL IFC"}

    executor.submit(
        _run_full_ifc_generation_job, job_id, xlsx_path, template_path, zip_path, filename_pattern,
        relay_position_enabled, hut_name, section, script_type,
    )

    return jsonify({"job_id": job_id})


@app.route("/api/status/<job_id>")
@login_required
def status(job_id):
    with jobs_lock:
        job = jobs.get(job_id)
    if job is None:
        return jsonify({"error": "Unknown job id"}), 404
    return jsonify(job)


@app.route("/api/resolve-contact/<job_id>", methods=["POST"])
@login_required
def resolve_contact(job_id):
    """
    Accepts the user's chosen alternative contact pair (e.g. "D3/D4")
    for a conflict the job is currently paused on, and wakes up that
    job's waiting thread to continue with it.
    """
    data = request.get_json(silent=True) or {}
    cancelled = bool(data.get("cancelled"))
    chosen = data.get("chosen_contact_pair", "").strip()
    chosen_index = data.get("chosen_occurrence_index")
    if not cancelled and (not chosen or chosen_index is None):
        return jsonify({"error": "chosen_contact_pair and chosen_occurrence_index are required (or set cancelled: true)"}), 400

    with jobs_lock:
        job = jobs.get(job_id)
        if job is None:
            return jsonify({"error": "Unknown job id"}), 404
        if job.get("state") != "awaiting_contact_choice":
            return jsonify({"error": "This job isn't currently awaiting a contact choice"}), 400
        if cancelled:
            job["contact_choice_cancelled"] = True
        else:
            job["chosen_contact_pair"] = chosen
            job["chosen_occurrence_index"] = chosen_index

    with contact_choice_events_lock:
        event = contact_choice_events.get(job_id)
    if event is None:
        return jsonify({"error": "No pending conflict found for this job"}), 400
    event.set()
    return jsonify({"ok": True})


@app.route("/api/confirm/<job_id>", methods=["POST"])
@login_required
def confirm(job_id):
    """
    Answers a pause-and-ask prompt raised by ask_user_confirmation()
    (CONFIRMED: e.g. Track/SDF relays consistently missing from Relay
    Rack in pairs). Body: {"proceed": true/false}. Unblocks the job
    thread that's waiting on this job's Event.
    """
    data = request.get_json(silent=True) or {}
    proceed = bool(data.get("proceed", False))
    with jobs_lock:
        event = pending_confirmations.get(job_id)
        if event is None:
            return jsonify({"error": "No confirmation pending for this job"}), 404
        confirmation_answers[job_id] = proceed
    event.set()
    return jsonify({"ok": True})


def _can_access_job(job_id):
    """A job's owner or an admin can access it - anyone else (even if
    they somehow know/guess the job_id UUID) cannot.
    CONFIRMED FIX: checks the in-memory `jobs` dict FIRST (covers a job
    that's still running, or one that stopped early due to a critical
    error and never reached job_history - which only gets a record once
    a job reaches "done"). Falls back to persistent job_history for a
    job from a previous server run that's no longer in memory."""
    if session.get("is_admin"):
        return True
    with jobs_lock:
        job = jobs.get(job_id)
    if job is not None:
        return job.get("userid") == session.get("userid")
    conn = auth_core._get_conn()
    try:
        row = conn.execute("SELECT userid FROM job_history WHERE job_id = ?", (job_id,)).fetchone()
    finally:
        conn.close()
    return row is not None and row["userid"] == session.get("userid")


@app.route("/api/download/<job_id>")
@login_required
def download(job_id):
    with jobs_lock:
        job = jobs.get(job_id)
    if job is not None and job.get("state") == "done":
        job_folder = os.path.join(JOBS_DIR, job_id)
        return send_from_directory(job_folder, job["download_name"], as_attachment=True)

    # CONFIRMED: falls back to persistent history + the file still being
    # on disk - the in-memory `jobs` dict resets on every server
    # restart, but a user's past files should still be downloadable
    # ("access any time") as long as the actual output wasn't deleted.
    if not _can_access_job(job_id):
        return jsonify({"error": "Job not found"}), 404
    conn = auth_core._get_conn()
    try:
        row = conn.execute("SELECT download_name FROM job_history WHERE job_id = ?", (job_id,)).fetchone()
    finally:
        conn.close()
    if row is None:
        return jsonify({"error": "Job not ready"}), 404
    job_folder = os.path.join(JOBS_DIR, job_id)
    if not os.path.exists(os.path.join(job_folder, row["download_name"])):
        return jsonify({"error": "Output file no longer available on the server"}), 404
    return send_from_directory(job_folder, row["download_name"], as_attachment=True)


def _run_retemplate_job(job_id, target_paths, target_original_names, new_template_path, new_template_original_name, old_block_name, generate_index_enabled, xlsx_path, job_folder, targets_folder):
    """
    Background job (submitted to the shared executor, same pattern as
    every other generation job) - runs the actual erase/purge/insert
    work, updating job status per file so the frontend can show a real
    progress bar (current file name + percentage), and converts each
    result back to .dwg (with its ORIGINAL filename) if that's the
    format it was originally uploaded in.
    """
    log = GenerationLog()
    _set_status(job_id, state="processing", progress=0, current_file="", total=len(target_paths))

    try:
        if new_template_path.lower().endswith(".dwg"):
            new_template_path = convert_single_file_to_dxf(new_template_path)

        # CONFIRMED: convert any .dwg targets to .dxf first (erase/purge/
        # insert operates via ezdxf, which only works on .dxf), but
        # remember each one's ORIGINAL extension/name so the final output
        # can be converted back and keep its original filename/format -
        # re-keyed here by the CONVERTED path's basename, since that's
        # what apply_new_template's updated_paths will actually contain.
        converted_target_paths = []
        converted_to_original_name = {}
        for path in target_paths:
            original_name = target_original_names.get(path, os.path.basename(path))
            converted_path = path
            if path.lower().endswith(".dwg"):
                converted_path = convert_single_file_to_dxf(path)
            converted_target_paths.append(converted_path)
            converted_to_original_name[os.path.basename(converted_path)] = original_name
        target_original_names = converted_to_original_name

        def _on_progress(idx, total, filename):
            display_name = target_original_names.get(filename, filename)
            _set_status(job_id, state="processing", progress=int((idx - 1) / total * 100), current_file=display_name, total=total)

        updated_paths = apply_new_template(converted_target_paths, new_template_path, old_block_name, log, on_progress=_on_progress, original_names=target_original_names, new_block_name_override=new_template_original_name)
    except Exception as e:
        _set_status(job_id, state="error", error=f"Re-template failed: {e}")
        return

    if not updated_paths:
        _set_status(job_id, state="error", error="No target file contained the specified old block name - nothing was updated.", log_errors=True)
        log_xlsx_path = os.path.join(job_folder, "generation_log.xlsx")
        log.save_excel(log_xlsx_path)
        return

    _set_status(job_id, state="processing", progress=90, current_file="Finalizing...", total=len(target_paths))

    if generate_index_enabled:
        pending = []
        for path in updated_paths:
            doc = ezdxf.readfile(path)
            pending.append((os.path.basename(path), doc))
        _generate_index_for_pending(pending, xlsx_path, new_template_path, "[]", targets_folder, log, True)
        for final_name, doc in pending[len(updated_paths):]:
            doc.saveas(os.path.join(targets_folder, final_name))
            updated_paths.append(os.path.join(targets_folder, final_name))

    # CONFIRMED: convert each updated file back to its ORIGINAL format
    # (.dwg if that's what was uploaded), keeping the EXACT original
    # filename - the output should look identical in name/format to
    # what was uploaded, just with the template swapped inside.
    final_output_paths = []
    dwg_convert_folder = os.path.join(job_folder, "dwg_out")
    os.makedirs(dwg_convert_folder, exist_ok=True)
    dwg_needed = any(name.lower().endswith(".dwg") for name in target_original_names.values())
    if dwg_needed:
        convert_folder_to_dwg(targets_folder, dwg_convert_folder)

    for path in updated_paths:
        base_name = os.path.basename(path)
        original_name = target_original_names.get(base_name)
        if original_name and original_name.lower().endswith(".dwg"):
            converted_dwg = os.path.join(dwg_convert_folder, os.path.splitext(base_name)[0] + ".dwg")
            if os.path.exists(converted_dwg):
                final_path = os.path.join(job_folder, original_name)
                shutil.copy(converted_dwg, final_path)
                final_output_paths.append(final_path)
                continue
        # Not a DWG target (or conversion didn't produce a match) - use
        # as-is, with its original name (Index sheets have no "original"
        # name to preserve, so they just keep their own generated name).
        final_path = os.path.join(job_folder, original_name or base_name)
        if not os.path.exists(final_path):
            shutil.copy(path, final_path)
        final_output_paths.append(final_path)

    zip_path = os.path.join(job_folder, "retemplated_output.zip")
    with zipfile.ZipFile(zip_path, "w") as zf:
        _add_source_files_to_zip(zf, job_folder)
        for path in final_output_paths:
            zf.write(path, os.path.basename(path))

    log_xlsx_path = os.path.join(job_folder, "generation_log.xlsx")
    log.save_excel(log_xlsx_path)

    _set_status(
        job_id, state="done", progress=100,
        download_path=zip_path, download_name="retemplated_output.zip",
        updated_count=len(final_output_paths), log=log.entries,
    )


@app.route("/api/retemplate", methods=["POST"])
@login_required
def retemplate():
    """
    Saves the uploads and submits the actual work to the background
    executor (same pattern as every other generation endpoint) -
    returns immediately with a job_id for the frontend to poll via
    /api/status/<job_id> for a live progress bar.
    """
    if "target_files" not in request.files:
        return jsonify({"error": "No target files uploaded"}), 400
    if "new_template" not in request.files:
        return jsonify({"error": "No new template file uploaded"}), 400
    old_block_name = request.form.get("old_block_name", "").strip()
    if not old_block_name:
        return jsonify({"error": "Old title-block name is required"}), 400

    job_id = str(uuid.uuid4())
    job_folder = os.path.join(JOBS_DIR, job_id)
    targets_folder = os.path.join(job_folder, "targets")
    os.makedirs(targets_folder, exist_ok=True)

    target_files = request.files.getlist("target_files")
    target_paths = []
    target_original_names = {}  # original PATH (pre-conversion) -> original uploaded filename
    for f in target_files:
        if not f.filename.lower().endswith((".dxf", ".dwg")):
            continue
        original_name = os.path.basename(f.filename)
        path = os.path.join(targets_folder, original_name)
        f.save(path)
        target_paths.append(path)
        target_original_names[path] = original_name

    if not target_paths:
        return jsonify({"error": "No .dxf/.dwg target files found in upload"}), 400

    new_template_file = request.files["new_template"]
    new_template_original_name = os.path.splitext(os.path.basename(new_template_file.filename))[0]
    template_ext = os.path.splitext(new_template_file.filename)[1].lower()
    new_template_path = os.path.join(job_folder, f"new_template{template_ext}")
    new_template_file.save(new_template_path)

    generate_index_enabled = request.form.get("generate_index", "0") == "1"
    xlsx_saved_path = None
    if generate_index_enabled:
        xlsx_file = request.files.get("xlsx_for_index")
        if xlsx_file is not None:
            xlsx_saved_path = os.path.join(job_folder, "index_source.xlsx")
            xlsx_file.save(xlsx_saved_path)

    with jobs_lock:
        jobs[job_id] = {"state": "queued", "userid": session.get("userid"), "job_type": "RE-TEMPLATE DRAWINGS"}

    executor.submit(
        _run_retemplate_job, job_id, target_paths, target_original_names,
        new_template_path, new_template_original_name, old_block_name, generate_index_enabled, xlsx_saved_path,
        job_folder, targets_folder,
    )

    return jsonify({"job_id": job_id})

    return jsonify({"job_id": job_id, "log": log.entries, "updated_count": len(updated_paths)})



@app.route("/api/download-log/<job_id>")
@login_required
def download_log(job_id):
    """
    Downloads the generation log (Excel) for a job - works regardless of
    whether the job finished successfully or stopped on a critical error,
    since the log itself is what explains what went wrong.
    """
    if not _can_access_job(job_id):
        return jsonify({"error": "Job not found"}), 404
    job_folder = os.path.join(JOBS_DIR, job_id)
    log_path = os.path.join(job_folder, "generation_log.xlsx")
    if not os.path.exists(log_path):
        return jsonify({"error": "No log available for this job"}), 404
    return send_from_directory(job_folder, "generation_log.xlsx", as_attachment=True)


if __name__ == "__main__":
    # CONFIRMED: runs via waitress (production-grade WSGI server), not
    # Flask's built-in dev server - that one is explicitly unsuitable
    # for real concurrent/multi-user use, and debug=True is a genuine
    # security risk once more than one person can reach this server
    # (it exposes an interactive debugger that can execute arbitrary
    # code if triggered by an unhandled error).
    # threads=32 comfortably covers ~20 simultaneous users' page loads/
    # status polling on top of the 8 generation-job workers below -
    # HTTP request handling and job processing are separate pools, so
    # this doesn't compete with ThreadPoolExecutor's own worker count.
    from waitress import serve
    serve(app, host="0.0.0.0", port=5000, threads=32)
