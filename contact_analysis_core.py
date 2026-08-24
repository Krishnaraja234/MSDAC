"""
MSDAC Contact Analysis Chart Generator - core logic.

Runs AFTER every other circuit type has been generated (Relay Rack,
Signal, Track, AR & AZR, Communication, SDF, Data Logger, Custom
Circuits) - it re-parses the just-generated DXF files with ezdxf to find
out which relay contacts got used where, then produces one or more
"CONTACT ANALYSIS SHEET" DXFs (same pattern as every other circuit type:
returns (out_name, doc, sht) tuples the caller merges into the combined
output).

WHERE THE DATA COMES FROM (CONFIRMED - no separate Excel input needed):

  1. Relay roster (which relays exist, their type, their contact config)
     -> re-parsed from the generated RELAY RACK DXF(s). Each populated
     rack cell is a 'SingleRelay_20' block insert with attributes:
         NAME1 + NAME2  -> relay position name (e.g. "F54" + "TSPR")
         RELTYP         -> relay type code (e.g. "QN1", "QECX61")
         CNTCNF         -> contact configuration (e.g. "8F-8B")
         NAME3          -> "SPARE" for blank cells (skipped)

  2. Contact usage (which contact of which relay got used, on which
     sheet, by which circuit) -> re-parsed from EVERY other generated
     DXF. Each contact placement is a 'Front_Contact'/'Back_Contact'
     block insert (ezdxf's importer may suffix duplicates, e.g.
     'Front_Contact0', 'Front_Contact1'...) with attributes:
         R_NAME  -> the relay this contact belongs to
         S_NAME  -> the circuit/function consuming it
         C       -> contact letter (e.g. "D")
         F       -> first contact number of the pair (e.g. "1")
         A       -> second contact number, always F+1 (e.g. "2")
     Coil usage (R1/R2) comes the same way from 'Relay_Coil' inserts
     (R_NAME/S_NAME only - no C/F/A, since a coil isn't a lettered
     contact pair). The sheet number for any of this comes from that
     same DXF's own TITLE block's SHT attribute - no filename parsing.

CHART LAYOUT (CONFIRMED: modular "typical" files, one per relay contact
configuration - same pattern as AR & AZR's per-type template files,
rather than one bundled file):

  - Shared header - 'CONTACT_ANALYSIS_HEADER.dxf' ('INV_8F_8B_HEADER'
    block) - CONFIRMED: placed once above EACH relay-type section within
    a sheet (once before the QN1-8F-8B group, again before the QECX61
    group), not once per whole sheet.

  - Generic relays (any CNTCNF == "8F-8B") - 'QN1-8F-8B_CONFIGURATION.dxf'
    ('INV_8F_8B_ROW' block), one per relay. Each row has per-contact
    fields (A1, A2 ... D8) - CONFIRMED these get filled with the
    consuming circuit's name when used, left blank when spare - plus
    per-pair 'SHT_xxYY' fields (e.g. SHT_A1A2) which get the sheet
    number where that pair is used.

  - QECX61-type relays - 'QECX-4F-4B_CONFIGURATION.dxf'
    ('QECX-61-4F-4B' block, different tag names: 'R1/R2/SH', 'A1/A2/SH'
    etc. instead of 'SHT_R1R2'/'SHT_A1A2').

  - Any OTHER CNTCNF value has no known typical file yet. Rather than
    guess at unseen geometry, that relay's row is SKIPPED with a warning
    in the generation log. CONFIRMED: new configs get added as their own
    '<TYPE>_CONFIGURATION.dxf' file (Krish will say how/when to wire a
    new one in - not auto-discovered yet).

  - CONFIRMED exact placement coordinates (Krish's measured values):
    QN1 header (37.687, 255.171); 1st QN1 row (39.6610, 259.3041);
    QECX61 header (39.6610, 115.1378); 1st QECX61 row (39.6610, 115.137).
    Each subsequent row of the SAME section steps down by ROW_SPACING=6.0
    from the previous row of that section.

GROUPING (CONFIRMED - this is the "AR & AZR approach" restructure):
one Contact Analysis sheet per DISTINCT consuming circuit, not one
global roster paginated by row count. A relay is a member of a group if
ANY of its contacts/coil was used by a circuit sheet whose own TITLE
equals that group's name (so a relay CAN appear on more than one sheet
if its contacts are used by more than one circuit). Each relay's row
still shows its FULL usage across every circuit, not just the usage
under that group. Sheets are titled with a simple incrementing counter
("CONTACT ANALYSIS - 1", "CONTACT ANALYSIS - 2"...) - CONFIRMED no
fancy category label ("DN LINE SIGNAL CIRCUITS" etc.) is computed.

ASSUMPTIONS (not yet confirmed against a real production run - flagged
for Krish to double check/tune):
  - No overflow handling yet if a single group's relay count would run
    off the bottom of the sheet or collide with the other section's
    fixed header position - the placement coordinates above are used
    exactly as given, with no dynamic pagination within a group.
  - Sheet numbering for Contact Analysis continues from
    (highest sheet number used by any other circuit type so far) + 1,
    since there's no FIELD PG.NO row feeding it its own start number.
"""

import os
import re

import ezdxf
from ezdxf.addons import importer

from signal_core import TEMPLATES_DIR, resolve_template_path, _safe_str, insert_border_title


def compute_available_contact_pairs(letter: str, cntcnf: str, used_codes: set, contact_type: str = None) -> list:
    """
    CONFIRMED (from Krish's CONTACTS_CONFIGURATION.xlsx): alternatives
    are shown as FRONT-CONTACT PAIRS (e.g. "D1/D2"), and are searched
    ACROSS ALL FOUR LETTERS (A, B, C, D) - not restricted to the same
    letter as the conflicting code. A relay's usable contacts span all
    four letter groups, each following the SAME position-to-role
    mapping:

      8F-8B (per letter): pairs (1,2) and (3,4) are FRONT pairs,
        (5,6) and (7,8) are BOTH BACK pairs. So across A/B/C/D: 8
        possible front pairs, 8 possible back pairs total.

      4F-4B (per letter): pair (1,2) is the FRONT pair, (3,4) is the
        BACK pair. Across A/B/C/D: 4 possible front pairs, 4 possible
        back pairs total.

    `letter` is now unused for building the candidate list (kept as a
    parameter for backward compatibility / logging) - candidates are
    generated from ALL FOUR letters, not just the conflicting one.

    used_codes: set of already-used contact codes (upper-cased, e.g.
        {"D1", "D2", "D7"}) for this relay - only codes of the SAME
        contact_type matter here (caller already filters this).
    contact_type: "F" or "B" - which set of pairs to offer. If None
        (unknown), falls back to offering ALL pairs (front+back) from
        every letter - old behavior, kept as a safety net.

    Returns a list of (first_code, second_code) tuples for every PAIR
    that's entirely free (BOTH halves unused) - a pair with only one
    half free isn't offered, since the pairing convention needs both.
    """
    m = re.match(r"(\d+)F-(\d+)B", (cntcnf or "").strip().upper())
    total_positions = int(m.group(1)) if m else 8  # CONFIRMED default: 8F-8B is the common case

    if total_positions == 8:
        # CONFIRMED CORRECTED mapping (from Krish's config table): per
        # letter, positions are 1=FRONT,2=ARM,3=FRONT,4=ARM,5=ARM,
        # 6=BACK,7=ARM,8=BACK - so pairs (1,2) and (3,4) are FRONT
        # pairs, and pairs (5,6) and (7,8) are BOTH back pairs (earlier
        # assumption that (5,6) was unusable was wrong - it has one ARM
        # + one BACK, same shape as (7,8)).
        front_pair_starts = [1, 3]
        back_pair_starts = [5, 7]
    elif total_positions == 4:
        front_pair_starts = [1]
        back_pair_starts = [3]
    else:
        # CONFIRMED FALLBACK for any other configuration not covered by
        # the known 8F-8B/4F-4B table: split the range in half, same as
        # the earlier (less precise) rule - front gets the first half's
        # pairs, back gets the second half's.
        half = total_positions // 2
        front_pair_starts = list(range(1, half, 2))
        back_pair_starts = list(range(half + 1, total_positions, 2))

    if contact_type == "F":
        pair_starts = front_pair_starts
    elif contact_type == "B":
        pair_starts = back_pair_starts
    else:
        pair_starts = front_pair_starts + back_pair_starts

    ALL_LETTERS = ["A", "B", "C", "D"]
    pairs = []
    for candidate_letter in ALL_LETTERS:
        for start in pair_starts:
            first_code = f"{candidate_letter}{start}"
            second_code = f"{candidate_letter}{start + 1}"
            if first_code.upper() not in used_codes and second_code.upper() not in used_codes:
                pairs.append((first_code, second_code))
    return pairs



# CONFIRMED restructure: Contact Analysis now loads its blocks from
# separate "typical" files - one per relay contact-configuration, plus a
# shared header - matching how AR & AZR picks a separate template file
# per relay type, rather than one bundled CONTACT_ANALYSIS.dxf. CONFIRMED
# the QECX61 sub-header (QECX61-TITLE) is dropped - every row now sits
# under the one shared header.
HEADER_TEMPLATE = resolve_template_path("CONTACT_ANALYSIS_HEADER.dxf")
HEADER_BLOCK_NAME = "INV_8F_8B_HEADER"

# CONFIRMED: only these two typical files exist right now - Krish will
# say how to wire up new ones later (not auto-discovered yet).
ROW_TEMPLATES = {
    "8F-8B": {
        "file": resolve_template_path("QN1-8F-8B_CONFIGURATION.dxf"),
        "block": "INV_8F_8B_ROW",
    },
    "QECX61": {
        "file": resolve_template_path("QECX-4F-4B_CONFIGURATION.dxf"),
        "block": "QECX-61-4F-4B",
    },
}

TITLE_TEXT_PREFIX = "CONTACT ANALYSIS - "

# --- CONFIRMED exact placement coordinates (Krish's measured values) -----
# QN1-8F-8B section: header once, then rows stepping down by ROW_SPACING.
HEADER_QN1_DROP = (39.661, 258.7041)
FIRST_QN1_ROW_POINT = (39.661, 259.0041)

# QECX61 section: its own header instance (same block, reused), then rows.
HEADER_QECX_DROP = (39.661, 115.137)
FIRST_QECX_ROW_POINT = (39.6610, 115.137)

# CONFIRMED BUG FIX: unlike the row blocks (INV_8F_8B_ROW/QECX-61-4F-4B,
# which insert correctly with no correction needed - verified directly),
# the shared header block's internal geometry was extracted from the
# original whole-drawing reference file using large, absolute-style
# local coordinates rather than being centered on its own origin. Its
# own "pick point" (37.687, 255.171) - CONFIRMED same value for both the
# QN1 and QECX placements, since it's the same block reused - is exactly
# that baked-in offset. Subtracting it from the desired drop point
# before calling add_blockref() cancels it out; verified directly
# (inserting at DROP - PICK lands the geometry's own reference corner
# exactly at DROP, matching Krish's reported actual position otherwise).
HEADER_PICK_POINT = (37.687, 255.171)
HEADER_QN1_POINT = (HEADER_QN1_DROP[0] - HEADER_PICK_POINT[0], HEADER_QN1_DROP[1] - HEADER_PICK_POINT[1])
HEADER_QECX_POINT = (HEADER_QECX_DROP[0] - HEADER_PICK_POINT[0], HEADER_QECX_DROP[1] - HEADER_PICK_POINT[1])

ROW_SPACING = 6.0

CONTACT_BLOCK_NAMES = ("Front_Contact", "Back_Contact", "SDF_CONTACT", "LCPR_FRONT", "LCPR_BACK CONTACT")
# CONFIRMED FIX: LCPR_FRONT/LCPR_BACK CONTACT were missing from this
# list (relay_position_core.py's own CONTACT_BLOCK_NAMES already
# included them) - same bug pattern as the TIMER_RELAY/SDF_RELAY gap
# below, just on the contact side. Meant LCPR contacts' R(POS), even
# though correctly written by the Relay Position feature, was never
# shown in Contact Analysis output, and LCPR contact usage was never
# checked for repetition at all.
# CONFIRMED: TIMER_RELAY/SDF_RELAY are SDF-specific coil-equivalent
# blocks for timer relays (e.g. "10-60 SEC JR") - same S_NAME/R_NAME
# shape as Relay_Coil, just a different block name. Missing these meant
# timer relay coils were invisible to Contact Analysis entirely.
COIL_BLOCK_NAMES = ("Relay_Coil", "TIMER_RELAY", "SDF_RELAY")

# Generic 8F-8B row layout: (attr1, attr2, sht_tag)
_GENERIC_8F8B_PAIRS = [("R1", "R2", "SHT_R1R2")]
for letter in "ABCD":
    for n in (1, 3, 5, 7):
        _GENERIC_8F8B_PAIRS.append((f"{letter}{n}", f"{letter}{n + 1}", f"SHT_{letter}{n}{letter}{n + 1}"))

# QECX61 bespoke 4F-4B row layout: (attr1, attr2, sht_tag)
_QECX61_4F4B_PAIRS = [("R1", "R2", "R1/R2/SH")]
for letter in "ABCD":
    for n in (1, 3):
        _QECX61_4F4B_PAIRS.append((f"{letter}{n}", f"{letter}{n + 1}", f"{letter}{n}/{letter}{n + 1}/SH"))


def _is_qecx61(relay_type: str) -> bool:
    # CONFIRMED: the actual RELTYP value in Relay Rack is "QECX" (not
    # "QECX61" - that was a wrong assumption). Matching on "QECX" as a
    # prefix so it still works if a more specific variant ever appears.
    return bool(relay_type) and relay_type.strip().upper().startswith("QECX")


def _is_timer_relay(relay_type: str) -> bool:
    # CONFIRMED: RELTYP="TIMER" (with CNTCNF holding a duration label
    # like "10-60 SEC" instead of a normal 8F-8B/4F-4B contact
    # configuration) - these relays don't have a chartable contact bank,
    # so they're intentionally excluded from Contact Analysis.
    return bool(relay_type) and relay_type.strip().upper() == "TIMER"


def _format_sheet_number(raw: str) -> str:
    """CONFIRMED: SHT_xxYY fields show 'SH001'/'SH700' style (SH prefix,
    zero-padded to at least 3 digits), not the raw sheet number alone."""
    raw = (raw or "").strip()
    if not raw:
        return ""
    digits = re.sub(r"\D", "", raw)
    if not digits:
        return _safe_str(raw)
    return f"SH{int(digits):03d}"


# CONFIRMED: TS(A)/TS(B) are alternate names sometimes used instead of
# TS(M)/TS(R) for the same Track/SDF relay pair - treated as identical
# when matching against Relay Rack (both directions, everywhere a
# relay's NAME2 gets used as part of a matching key).
_TRACK_TYPE_ALIASES = {"TS(A)": "TS(M)", "TS(B)": "TS(R)"}
# CONFIRMED pairing: if one half of a TS(M)/TS(R) pair (in either naming
# convention) is missing from Relay Rack, that's only "expected" (ask the
# user, don't hard-fail) if its PARTNER is ALSO missing - a partial
# mismatch (one present, one absent) stays a hard vital error.
_TRACK_PAIR_PARTNER = {"TS(M)": "TS(R)", "TS(R)": "TS(M)"}


def _normalize_relay_type(name2: str) -> str:
    if not name2:
        return name2
    upper = name2.strip().upper()
    return _TRACK_TYPE_ALIASES.get(upper, upper)


def _inserts_named(entities, *prefixes):
    """Every INSERT whose block name starts with any of the given prefixes
    (handles ezdxf importer's 'Front_Contact0'/'Front_Contact1' suffixing)."""
    return [
        e for e in entities
        if e.dxftype() == "INSERT" and any(e.dxf.name.startswith(p) for p in prefixes)
    ]


def _attrib_map(insert_entity) -> dict:
    return {a.dxf.tag: (a.dxf.text or "").strip() for a in insert_entity.attribs}


def extract_relay_roster(relay_rack_dxf_paths: list) -> list:
    """
    Re-parses the already-generated Relay Rack DXF(s) and returns one
    dict per populated (non-SPARE) relay position:
        {"bit_name": "F54 TSPR", "name1": "F54", "name2": "TSPR",
         "relay_type": "QN1", "cntcnf": "8F-8B", "rack_location": "A2"}
    "bit_name" (NAME1+" "+NAME2) is for DISPLAY only. CONFIRMED: matching
    a relay to its contact/coil usage elsewhere uses the (NAME1, NAME2)
    pair directly against each instance's (S_NAME, R_NAME) - the same
    convention every circuit-type module already writes - NOT a single
    concatenated name.
    "rack_location" is the rack's own identifier (CONFIRMED: each rack
    corresponds to a hut/location, e.g. "A2") - parsed from that Relay
    Rack sheet's own TITLE text ("RELAY RACK - {rack_num}", set in
    relay_rack_core.py) - used by Communication's contact-location check.
    Order is preserved (rack position order), duplicates across multiple
    Relay Rack sheets are kept as-is (each physical relay should only
    appear once across the rack's sheets).
    """
    roster = []
    seen = set()
    for path in relay_rack_dxf_paths:
        try:
            doc = ezdxf.readfile(path)
        except Exception:
            continue
        msp = doc.modelspace()

        rack_location = ""
        title_inserts = _inserts_named(msp, "TITLE", "TITLEBLOCK")
        if title_inserts:
            title_text = _attrib_map(title_inserts[0]).get("TITLE", "")
            m = re.match(r"^RELAY RACK\s*-\s*(.+)$", title_text.strip())
            if m:
                rack_location = m.group(1).strip()

        for e in _inserts_named(msp, "SingleRelay_20"):
            att = _attrib_map(e)
            if att.get("NAME3", "").strip().upper() == "SPARE":
                continue
            name1 = att.get("NAME1", "").strip()
            name2 = att.get("NAME2", "").strip()
            if not name2:
                continue  # CONFIRMED: NAME1 can legitimately be blank (e.g. "DN SDF", "DN JR" alone) - only NAME2 is required
            name2_norm = _normalize_relay_type(name2)
            key = (name1.upper(), name2_norm)
            if key in seen:
                continue
            seen.add(key)
            roster.append({
                "bit_name": f"{name1} {name2}".strip(),
                "name1": name1,
                "name2": name2,
                "name2_norm": name2_norm,
                "relay_type": att.get("RELTYP", "").strip(),
                "cntcnf": att.get("CNTCNF", "").strip(),
                "rack_location": rack_location,
            })
    return roster


def _wire_values_from_wiring_doc(wiring: str, doc_value: str):
    """
    CONFIRMED: WIRING can be a single digit (1 or 2) OR a 2-digit string
    (each digit 1 or 2 - max is 2). DOC's first letter is the direction
    (L=left, R=right; e.g. 'LF'/'RF'/'LB'/'RB' for contacts, 'LR1'/'RR1'
    for coils).
      - Single digit (e.g. "2"): CONFIRMED both positions (e.g. A1 AND
        A2) get that SAME value - the arm's wiring count applies
        uniformly, no left/right split needed.
      - Two digits (e.g. "21"): split between the two positions -
          Left:  use as-is (1st digit -> "F" position, 2nd -> "A" position).
          Right: REVERSE them (2nd digit -> "F" position, 1st -> "A" position).
    Returns (w_f, w_a) - the value to show at the "F" position (e.g. A1)
    and the "A" position (e.g. A2) respectively. Confirmed against given
    examples: WIRING="21"/DOC="LF" -> (2,1); WIRING="21"/DOC="RF" -> (1,2);
    WIRING="2"/DOC="LF" -> (2,2). Returns ("","") if WIRING/DOC missing
    or WIRING isn't a clean 1- or 2-digit string.
    """
    wiring = (wiring or "").strip()
    doc_value = (doc_value or "").strip().upper()
    if not wiring or not wiring.isdigit() or not doc_value:
        return "", ""
    if len(wiring) == 1:
        return wiring, wiring
    if len(wiring) == 2:
        d1, d2 = wiring[0], wiring[1]
        if doc_value.startswith("R"):
            return d2, d1
        return d1, d2  # CONFIRMED: "L" is the default/normal order; unrecognized direction also falls here
    return "", ""  # unexpected length (3+ digits) - not a recognized format


def extract_contact_usage(circuit_dxf_paths: list, log, filename_to_circuit_type: dict = None, roster_by_key: dict = None, on_conflict=None) -> dict:
    """
    Re-parses every OTHER already-generated DXF (Signal, Track, AR & AZR,
    Communication, SDF, Data Logger, Custom Circuits - anything that
    isn't the Relay Rack itself) and returns:
        {(s_name_upper, r_name_upper): {contact_code: (sheet_number, consuming_circuit, wire_value)}}
    contact_code is e.g. "D1", or "R1"/"R2" for the relay's own coil.
    wire_value (CONFIRMED) is what actually gets shown in the chart's
    individual A1/A2-style fields - computed from that instance's own
    WIRING + DOC attributes via _wire_values_from_wiring_doc() - NOT the
    consuming circuit's name. sheet_number is still what SHT_xxYY shows.

    CORRECTED matching convention (CONFIRMED): (S_NAME, R_NAME) together
    identify the RELAY itself - matching a Relay Rack cell's
    (NAME1, NAME2) split - NOT "S_NAME = consuming circuit". The best
    available stand-in for "which circuit consumes this contact" is that
    generated sheet's own TITLE block text, used for grouping/title
    lookups elsewhere (not shown directly in the chart's contact cells
    anymore, now that wire_value is confirmed to belong there instead).

    CONFIRMED VITAL ERROR: if the SAME relay's SAME contact (e.g. F54
    TSPR's D1) is used more than once anywhere, that's an error - NOT
    silently resolved by keeping whichever occurrence was found last.
    Logged via log.error(), which the caller treats as a stop-before-
    output condition, same as every other vital error in this app.
    CONFIRMED: errors are simple and labeled by circuit type.

    CONFIRMED Communication-specific filter: a Communication (UFSBI)
    sheet's own TITLE encodes its two locations ("UFSBI {id}- {loc1} -
    {loc2} CIRCUITS", set in communication_core.py). A CONTACT wired on
    that sheet is only counted if the relay's own true rack location
    (roster_by_key[key]["rack_location"], from Relay Rack) matches
    LOC1 - the panel's INPUT side. A COIL is only counted if its
    relay's true rack location matches LOC2 instead - the panel's
    OUTPUT/receiving side. Either way, a mismatch is silently skipped
    (not an error, just not counted).
    """
    filename_to_circuit_type = filename_to_circuit_type or {}
    roster_by_key = roster_by_key or {}
    raw = {}

    def _record(s_name, r_name, contact_code, sheet_number, consuming_circuit, source_file, wire_value, owning_signal, entity=None, tag_name=None, r_pos=""):
        # CONFIRMED: S_NAME can legitimately be blank (e.g. "DN SDF"/
        # "UP SDF" standalone entries, matching the same blank-NAME1
        # convention already handled in the roster) - only R_NAME and
        # contact_code are actually required to record a usage.
        if not r_name or not contact_code:
            return
        # CONFIRMED: "SPARE" is Data Logger's own padding marker (fills
        # out empty slots on its last sheet purely for visual layout) -
        # not a real relay reference, so it's never checked against
        # Relay Rack or shown in Contact Analysis.
        if r_name.strip().upper() == "SPARE":
            return
        s_name = s_name or ""
        key = (s_name.strip().upper(), _normalize_relay_type(r_name))
        raw.setdefault(key, {}).setdefault(contact_code.strip().upper(), []).append(
            (sheet_number, consuming_circuit, source_file, wire_value, owning_signal, entity, tag_name, r_pos)
        )

    docs_by_source_file = {}
    paths_by_source_file = {}
    for path in circuit_dxf_paths:
        try:
            doc = ezdxf.readfile(path)
        except Exception:
            continue
        docs_by_source_file[os.path.basename(path)] = doc
        paths_by_source_file[os.path.basename(path)] = path
        msp = doc.modelspace()
        source_file = os.path.basename(path)
        circuit_type = filename_to_circuit_type.get(source_file, "")

        sht_number = ""
        consuming_circuit = ""
        group_location = ""
        owning_signal = ""
        title_inserts = _inserts_named(msp, "TITLE", "TITLEBLOCK")
        if title_inserts:
            title_att = _attrib_map(title_inserts[0])
            sht_number = title_att.get("SHT", "")
            consuming_circuit = title_att.get("TITLE", "")
            group_location = title_att.get("GROUP_LOC", "").strip().upper()
            owning_signal = title_att.get("OWNING_SIGNAL", "").strip()

        communication_loc1 = None
        communication_loc2 = None
        if circuit_type == "COMMUNICATION" and title_inserts:
            # CONFIRMED FIX: read LOC1/LOC2 directly from their own
            # hidden attributes rather than regex-parsing the TITLE
            # text - a hut name containing its own hyphen (e.g. "MSDAC
            # HUT-1") broke the old regex-based split.
            title_att = _attrib_map(title_inserts[0])
            if "LOC1" in title_att:
                communication_loc1 = title_att.get("LOC1", "").strip().upper()
                communication_loc2 = title_att.get("LOC2", "").strip().upper()

        for e in _inserts_named(msp, *CONTACT_BLOCK_NAMES):
            att = _attrib_map(e)
            r_name = att.get("R_NAME", "")
            s_name = att.get("S_NAME", "")
            letter = att.get("C", "")
            # CONFIRMED: Back_Contact uses tags A (higher number, e.g. "8")
            # + B (lower number, e.g. "7") - NOT F/A like Front_Contact/
            # SDF_CONTACT. Detected by tag presence (robust regardless of
            # block name) rather than assuming from the block name alone.
            if "B" in att:
                first_val = att.get("B", "")   # "first"/lower position - same role as Front_Contact's "F"
            else:
                first_val = att.get("F", "")
            second_val = att.get("A", "")       # "second"/higher position - same tag name in both schemes
            wiring = att.get("WIRING", "")
            doc_value = att.get("DOC", "")
            r_pos = att.get("R(POS)", "")
            wire_first, wire_second = _wire_values_from_wiring_doc(wiring, doc_value)

            if communication_loc1 is not None:
                relay = roster_by_key.get((s_name.strip().upper(), _normalize_relay_type(r_name)))
                if relay is None:
                    continue  # not in Relay Rack at all - never a "match", regardless of what LOC1 is
                # CONFIRMED FIX (reverted): rack_location (e.g. "A2", a
                # rack cabinet ID) and LOC1/LOC2 (e.g. "MSDAC HUT-2", a
                # hut/room name from the Communication sheet's own text)
                # are DIFFERENT naming schemes that can never textually
                # match - comparing them excluded every single
                # Communication relay from CA, not just cross-location
                # ones. group_location uses the SAME hut/room-name
                # format as LOC1/LOC2 (both come from the Communication
                # sheet's own text), so it's the comparable value.
                if group_location != communication_loc1:
                    continue  # this row-group's home hut isn't LOC1 - skip, not an error

            if letter and first_val:
                first_tag = "B" if "B" in att else "F"
                _record(s_name, r_name, f"{letter}{first_val}", sht_number, consuming_circuit, source_file, wire_first, owning_signal, e, first_tag, r_pos)
            if letter and second_val:
                _record(s_name, r_name, f"{letter}{second_val}", sht_number, consuming_circuit, source_file, wire_second, owning_signal, e, "A", r_pos)

        for e in _inserts_named(msp, *COIL_BLOCK_NAMES):
            att = _attrib_map(e)
            r_name = att.get("R_NAME", "")
            s_name = att.get("S_NAME", "")
            wiring = att.get("WIRING", "")
            doc_value = att.get("DOC", "")
            r_pos = att.get("R(POS)", "")
            wire_r1, wire_r2 = _wire_values_from_wiring_doc(wiring, doc_value)

            if communication_loc2 is not None:
                relay = roster_by_key.get((s_name.strip().upper(), _normalize_relay_type(r_name)))
                if relay is None:
                    continue  # not in Relay Rack at all - never a "match", regardless of what LOC2 is
                # CONFIRMED FIX (reverted): same correction as LOC1 above.
                if group_location != communication_loc2:
                    continue  # this row-group's home hut isn't LOC2 - skip, not an error

            _record(s_name, r_name, "R1", sht_number, consuming_circuit, source_file, wire_r1, owning_signal, r_pos=r_pos)
            _record(s_name, r_name, "R2", sht_number, consuming_circuit, source_file, wire_r2, owning_signal, r_pos=r_pos)

    usage = {}
    title_to_circuit_type = {}
    def _pair_partner_code(contact_code):
        """Given 'D1', returns 'D2' (and vice versa) - the OTHER half of
        the same physical contact pair, regardless of CNTCNF, since
        pairs are always consecutive (odd, even)."""
        m = re.match(r"^([A-Za-z]+)(\d+)$", contact_code.strip())
        if not m:
            return None
        letter, num_str = m.group(1), m.group(2)
        num = int(num_str)
        partner_num = num + 1 if num % 2 == 1 else num - 1
        return f"{letter}{partner_num}"

    def _contact_type(occs):
        for occ in occs:
            tag = occ[6]
            if tag in ("F", "B"):
                return tag
            if tag == "A":
                # CONFIRMED FIX: "A" (the second half) is used by BOTH
                # Front_Contact and Back_Contact - the tag name alone
                # can't tell them apart. Check the entity itself for
                # which FIRST-half tag it actually has instead.
                entity = occ[5]
                if entity is not None:
                    entity_tags = {a.dxf.tag for a in entity.attribs}
                    if "B" in entity_tags:
                        return "B"
                    if "F" in entity_tags:
                        return "F"
        return None

    for key, contacts in raw.items():
        usage[key] = {}
        already_combined = set()  # contact_codes already handled as part of a pair
        # CONFIRMED FIX: codes the picker has ALREADY assigned as a
        # resolution for an earlier conflict on THIS SAME relay, within
        # this same run - must be excluded from the available pool for
        # any later conflict on the same relay (e.g. Communication's
        # "main" and "redundant" both conflicting on one relay: once
        # the user picks an alternative for "main", that same
        # alternative must not be offered again when resolving
        # "redundant"), since the original scan alone doesn't know
        # about picks made mid-run.
        already_reassigned_codes = set()
        for contact_code, occurrences in contacts.items():
            if contact_code in already_combined:
                continue
            for _sht, consumer, file_name, _wire, _owning, _ent, _tag, _rpos in occurrences:
                if consumer and consumer not in title_to_circuit_type:
                    title_to_circuit_type[consumer] = filename_to_circuit_type.get(file_name, file_name)
            sheet_nums = sorted({sht_num for sht_num, _c, _f, _w, _o, _e, _t, _rp in occurrences if sht_num})
            circuit_types = sorted({
                filename_to_circuit_type.get(file_name, file_name)
                for _s, _c, file_name, _w, _o, _e, _t, _rp in occurrences
            })
            # CONFIRMED: previously, repeats within the SAME sheet were
            # exempted (on the theory that SDF's DN SDF/UP SDF timer
            # wiring legitimately reuses one contact twice on one sheet) -
            # but Krish confirmed the picker should ALSO offer to fix
            # same-sheet repeats now, even though this may occasionally
            # also catch that SDF timer pattern. Uses the total number of
            # OCCURRENCES (not unique sheet numbers) so 2+ uses on the
            # exact same sheet number count as a conflict too.
            # CONFIRMED EXEMPTION: Data Logger's C7/C8 is its designated
            # last-resort fallback pair (used when nothing else is free)
            # - repeating across sheets for the same relay is expected
            # behavior when Relay Rack has multiple positions for it, not
            # a genuine data conflict, so it's excluded from this check.
            is_datalogger_fallback_pair = (
                circuit_types == ["DATA LOGGER"] and contact_code.strip().upper() in ("C7", "C8")
            )

            # CONFIRMED: if this contact's PAIR PARTNER (e.g. "D2" for
            # "D1") also exists for this relay and ALSO conflicts, both
            # halves are resolved together in ONE picker interaction,
            # not two separate ones - they're really one physical
            # conflict, not two.
            partner_code = _pair_partner_code(contact_code)
            partner_occurrences = contacts.get(partner_code) if partner_code else None
            combine_with_partner = (
                partner_occurrences is not None
                and len(occurrences) > 1
                and len(partner_occurrences) > 1
                and not is_datalogger_fallback_pair
            )

            reassigned_new_code = None
            reassigned_occurrence = None
            reassigned_partner_new_code = None
            reassigned_partner_occurrence = None

            if len(occurrences) > 1 and not is_datalogger_fallback_pair:
                resolved = False
                # CONFIRMED: R1/R2 are the relay's own COIL terminals,
                # not a regular switching contact - there's no
                # alternative position to reassign a coil to (unlike
                # front/back contacts, which can move to a different
                # free pair), so this is always a hard, unresolvable
                # conflict - stop generation immediately, never offer
                # the picker for this case.
                if contact_code.strip().upper() in ("R1", "R2"):
                    log.error(
                        f"[{'/'.join(circuit_types)}] Duplicate coil: "
                        f"{key[0]} {contact_code.strip().upper()} (sheets {', '.join(sheet_nums)}) "
                        "- a relay's coil can't be reassigned to an alternative position."
                    )
                    resolved = True  # error already logged above, don't double-log below
                elif on_conflict is not None and roster_by_key is not None:
                    relay = roster_by_key.get(key)
                    if relay is not None:
                        # CONFIRMED: front and back contacts are separate,
                        # independent sets of physical positions - a
                        # front-side "D7" and a back-side "D7" don't
                        # conflict with each other, even though they share
                        # the same letter/number label. When a FRONT
                        # contact repeats, only OTHER FRONT contacts count
                        # as "used" (and vice versa for back) - classified
                        # by which tag ("F" vs "B") each contact_code's own
                        # occurrences actually came from.
                        this_type = _contact_type(occurrences)
                        used_codes = {
                            c.strip().upper() for c, occs in contacts.items()
                            if _contact_type(occs) == this_type
                        }
                        used_codes |= {
                            c for c, t in already_reassigned_codes if t == this_type
                        }
                        # CONFIRMED: the letter prefix comes from the
                        # conflicting code itself (e.g. "D" from "D7") -
                        # alternatives use the SAME letter, just
                        # different numbers. Same logic covers Back_Contact
                        # too, since its letter also comes from the "C" tag.
                        m_letter = re.match(r"^([A-Za-z]+)(\d+)$", contact_code.strip())
                        if m_letter:
                            letter = m_letter.group(1)
                            available_pairs = compute_available_contact_pairs(letter, relay.get("cntcnf", ""), used_codes, this_type)
                            if not available_pairs:
                                # CONFIRMED: no free pairs at all for this
                                # relay - stop the process, nothing to
                                # offer the user, this is a hard vital error.
                                display_code = f"{contact_code}/{partner_code}" if combine_with_partner else contact_code
                                log.error(
                                    f"[{'/'.join(circuit_types)}] Duplicate contact: "
                                    f"{key[0]}/{key[1]} {display_code} (sheets {', '.join(sheet_nums)}) "
                                    "- no free alternative contact pairs available for this relay."
                                )
                                resolved = True  # error already logged above, don't double-log below
                            else:
                                # CONFIRMED: the user picks BOTH which
                                # alternative pair to use AND which specific
                                # conflicting occurrence (sheet) gets changed
                                # to it - other occurrences are left as-is.
                                occurrence_choices = [
                                    {"sheet_number": o[0], "circuit_type": filename_to_circuit_type.get(o[2], o[2])}
                                    for o in occurrences
                                ]
                                display_code = f"{contact_code}/{partner_code}" if combine_with_partner else contact_code
                                result = on_conflict(key, display_code, available_pairs, sheet_nums, circuit_types, occurrence_choices)
                                if result == "CANCELLED":
                                    # CONFIRMED: user chose to cancel generation
                                    # entirely rather than pick an alternative -
                                    # same cancellation path already used
                                    # elsewhere (confirm_callback), so the Full
                                    # IFC job stops cleanly and the generation
                                    # log up to this point is still saved.
                                    raise ContactAnalysisCancelled(
                                        f"User cancelled generation while resolving a repeated contact "
                                        f"conflict on {key[0]}/{key[1]} {display_code}."
                                    )
                                if result is not None:
                                    chosen_pair, chosen_occurrence_index = result
                                    target = occurrences[chosen_occurrence_index]
                                    target_entity, target_tag = target[5], target[6]
                                    if target_entity is not None and target_tag is not None:
                                        # CONFIRMED: reassign the CHOSEN
                                        # occurrence's own block attributes -
                                        # both the number tag (F/B/A) AND the
                                        # letter tag (C), since the chosen
                                        # alternative can come from a
                                        # DIFFERENT letter than the original
                                        # (e.g. "D1" -> "A1") - the other half
                                        # of the pair, and every other
                                        # occurrence, are untouched, EXCEPT
                                        # when combined with its partner
                                        # below, where the SAME entity's
                                        # other tag also gets updated.
                                        full_code = chosen_pair[0] if target_tag in ("F", "B") else chosen_pair[1]
                                        m_new = re.match(r"^([A-Za-z]+)(\d+)$", full_code)
                                        new_letter, new_number = m_new.group(1), m_new.group(2)
                                        for att in target_entity.attribs:
                                            if att.dxf.tag == target_tag:
                                                att.dxf.text = new_number
                                            elif att.dxf.tag == "C":
                                                att.dxf.text = new_letter

                                        reassigned_new_code = f"{new_letter}{new_number}"
                                        reassigned_occurrence = target
                                        already_reassigned_codes.add((reassigned_new_code, target_tag))

                                        # CONFIRMED: combined-pair resolution -
                                        # find the SAME entity's occurrence
                                        # under the partner code (matched by
                                        # sheet number, not position, to avoid
                                        # any list-ordering assumptions) and
                                        # update ITS tag to the OTHER half of
                                        # the chosen pair too.
                                        if combine_with_partner:
                                            partner_target = next(
                                                (o for o in partner_occurrences if o[0] == target[0] and o[2] == target[2]),
                                                None,
                                            )
                                            if partner_target is not None:
                                                partner_entity, partner_tag = partner_target[5], partner_target[6]
                                                if partner_entity is not None and partner_tag is not None:
                                                    partner_full_code = chosen_pair[0] if partner_tag in ("F", "B") else chosen_pair[1]
                                                    m_p = re.match(r"^([A-Za-z]+)(\d+)$", partner_full_code)
                                                    p_letter, p_number = m_p.group(1), m_p.group(2)
                                                    for att in partner_entity.attribs:
                                                        if att.dxf.tag == partner_tag:
                                                            att.dxf.text = p_number
                                                        elif att.dxf.tag == "C":
                                                            att.dxf.text = p_letter
                                                    reassigned_partner_new_code = f"{p_letter}{p_number}"
                                                    reassigned_partner_occurrence = partner_target
                                                    already_reassigned_codes.add((reassigned_partner_new_code, partner_tag))
                                            already_combined.add(partner_code)

                                        # CONFIRMED: persist the change - this
                                        # entity came from a doc re-read from
                                        # disk (separate from the in-memory
                                        # `pending` docs used for the final
                                        # zip), so the edit is lost unless
                                        # saved back to that same file here.
                                        target_source_file = target[2]
                                        target_doc = docs_by_source_file.get(target_source_file)
                                        target_path = paths_by_source_file.get(target_source_file)
                                        if target_doc is not None and target_path is not None:
                                            target_doc.saveas(target_path)

                                        log.warning(
                                            f"[{'/'.join(circuit_types)}] Duplicate contact {key[0]}/{key[1]} "
                                            f"{display_code} on sheet {target[0]} reassigned by user -> "
                                            f"{chosen_pair[0]}/{chosen_pair[1]}"
                                        )
                                        resolved = True
                if not resolved:
                    display_code = f"{contact_code}/{partner_code}" if combine_with_partner else contact_code
                    log.error(
                        f"[{'/'.join(circuit_types)}] Duplicate contact: "
                        f"{key[0]}/{key[1]} {display_code} (sheets {', '.join(sheet_nums)})"
                    )

            # Finalize usage entry for THIS contact_code (D1 or whichever
            # half was iterated to)
            if reassigned_new_code is not None and reassigned_occurrence is not None:
                sht_num, consumer, _file_name, wire_value, owning_signal, _e, _t, r_pos = reassigned_occurrence
                usage[key][reassigned_new_code] = (sht_num, consumer, wire_value, owning_signal, r_pos)
                remaining = [o for o in occurrences if o is not reassigned_occurrence]
                if remaining:
                    sht_num, consumer, _file_name, wire_value, owning_signal, _e, _t, r_pos = remaining[0]
                    usage[key][contact_code] = (sht_num, consumer, wire_value, owning_signal, r_pos)
            else:
                sht_num, consumer, _file_name, wire_value, owning_signal, _e, _t, r_pos = occurrences[0]
                usage[key][contact_code] = (sht_num, consumer, wire_value, owning_signal, r_pos)

            # Finalize usage entry for the PARTNER contact_code (D2),
            # since it was combined into this same resolution and won't
            # be visited again by the outer loop.
            if combine_with_partner and partner_occurrences is not None:
                if reassigned_partner_new_code is not None and reassigned_partner_occurrence is not None:
                    sht_num, consumer, _file_name, wire_value, owning_signal, _e, _t, r_pos = reassigned_partner_occurrence
                    usage[key][reassigned_partner_new_code] = (sht_num, consumer, wire_value, owning_signal, r_pos)
                    remaining_p = [o for o in partner_occurrences if o is not reassigned_partner_occurrence]
                    if remaining_p:
                        sht_num, consumer, _file_name, wire_value, owning_signal, _e, _t, r_pos = remaining_p[0]
                        usage[key][partner_code] = (sht_num, consumer, wire_value, owning_signal, r_pos)
                elif partner_code not in usage[key]:
                    sht_num, consumer, _file_name, wire_value, owning_signal, _e, _t, r_pos = partner_occurrences[0]
                    usage[key][partner_code] = (sht_num, consumer, wire_value, owning_signal, r_pos)

    return usage, title_to_circuit_type


def _ensure_block_imported(doc, block_name, source_file):
    """Imports a block DEFINITION from the given typical file into `doc`,
    only if it isn't already there (safe to call every row/sheet)."""
    if block_name not in doc.blocks:
        source_doc = ezdxf.readfile(source_file)
        imp = importer.Importer(source_doc, doc)
        imp.import_block(block_name)
        imp.finalize()


def _place_header(doc, insert_point):
    _ensure_block_imported(doc, HEADER_BLOCK_NAME, HEADER_TEMPLATE)
    return doc.modelspace().add_blockref(HEADER_BLOCK_NAME, insert_point)


def _place_row(doc, insert_point, relay: dict, usage: dict, log):
    """Imports the correct row block for this relay, fills its attributes
    from the extracted usage data. Returns True if placed, False if this
    relay's CNTCNF has no known template (row skipped, warning logged)."""
    bit_name = relay["bit_name"]
    relay_type = relay["relay_type"]
    cntcnf = relay["cntcnf"]
    relay_usage = usage.get((relay["name1"].upper(), relay["name2_norm"]), {})

    if _is_qecx61(relay_type):
        row_key = "QECX61"
        pairs = _QECX61_4F4B_PAIRS
    elif cntcnf.upper() == "8F-8B":
        row_key = "8F-8B"
        pairs = _GENERIC_8F8B_PAIRS
    else:
        log.warning(f"No typical for config: {bit_name} ({cntcnf})")
        return False

    template = ROW_TEMPLATES[row_key]
    block_name = template["block"]
    _ensure_block_imported(doc, block_name, template["file"])
    blockref = doc.modelspace().add_blockref(block_name, insert_point)

    values = {}
    if block_name == "INV_8F_8B_ROW":
        values["BIT-NAME"] = _safe_str(bit_name)
        values["RELAY_TYPE"] = _safe_str(relay_type)
        pos_tag = "POS"
    else:
        values["NAME"] = _safe_str(bit_name)
        values["QECX61"] = _safe_str(relay_type)
        values["4F-4B"] = "4F-4B"
        pos_tag = "RELAYPOS"

    # CONFIRMED: POS/RELAYPOS reflects the SAME physical relay's actual
    # R(POS) - the same value Relay Position already wrote onto every
    # instance of this relay's own contact/coil blocks, so any single
    # occurrence's r_pos is representative of the whole relay. If Relay
    # Position wasn't run (or hasn't reached this relay), r_pos will be
    # blank - defaults to "R0A0" as a clear placeholder in that case,
    # rather than leaving it blank.
    relay_r_pos = ""
    for v in relay_usage.values():
        if len(v) > 4 and v[4]:
            relay_r_pos = v[4]
            break
    values[pos_tag] = _safe_str(relay_r_pos) if relay_r_pos else "R0A0"

    for attr1, attr2, sht_tag in pairs:
        v1 = relay_usage.get(attr1)
        v2 = relay_usage.get(attr2)
        if v1:
            sht1, consumer1, wire1, _owning1 = v1[:4]
            values[attr1] = _safe_str(wire1)
        if v2:
            sht2, consumer2, wire2, _owning2 = v2[:4]
            values[attr2] = _safe_str(wire2)
        # Whichever half is used tells us the sheet number for the pair.
        pair_sht = (v1[0] if v1 else "") or (v2[0] if v2 else "")
        if pair_sht:
            values[sht_tag] = _format_sheet_number(pair_sht)

    blockref.add_auto_attribs(values)
    return True


def get_contact_analysis_start_sheet_number(xlsx_path: str) -> int:
    """
    CONFIRMED: Contact Analysis has its own row in FIELD PG.NO (labeled
    'CONTACT ANALYSIS', e.g. row 13/column C = 700) - same pattern every
    other circuit type already uses for its own starting sheet number.
    Replaces the old approach of scanning already-generated DXFs for the
    highest SHT already used.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(row=r, column=2).value or "").strip().upper()
        if label == "CONTACT ANALYSIS":
            value = ws.cell(row=r, column=3).value
            if value is None or str(value).strip() == "":
                raise ValueError("FIELD PG.NO has no Sheet Number for the CONTACT ANALYSIS row")
            return int(str(value).strip())
    raise ValueError("Could not find a 'CONTACT ANALYSIS' row in FIELD PG.NO")


def get_next_sheet_number(dxf_folder: str) -> int:
    """
    Scans every already-saved .dxf in dxf_folder for its TITLE block's SHT
    attribute and returns (highest SHT found) + 1. Used as the Contact
    Analysis start sheet number - CONFIRMED assumption: Contact Analysis
    has no FIELD PG.NO row of its own (it isn't one of the Excel-driven
    circuit types), so it simply continues after whatever came last,
    rather than needing its own configured start number.
    """
    highest = 0
    for fname in os.listdir(dxf_folder):
        if not fname.lower().endswith(".dxf"):
            continue
        try:
            doc = ezdxf.readfile(os.path.join(dxf_folder, fname))
        except Exception:
            continue
        for e in _inserts_named(doc.modelspace(), "TITLE", "TITLEBLOCK"):
            sht_text = _attrib_map(e).get("SHT", "")
            m = re.search(r"\d+", sht_text)
            if m:
                highest = max(highest, int(m.group()))
    return highest + 1


def _read_track_direction_lookup(xlsx_path: str) -> dict:
    """
    CONFIRMED: TRACK sheet's own 'DIRECTION' column (DN/UP) keyed by
    'TRACK NAME' (matches a relay's NAME1) - used to split TSPR/TS(M)/
    TS(R) relays into DN vs UP Contact Analysis groups.
    Returns {track_name_upper: "DN"/"UP"}.
    """
    import openpyxl

    lookup = {}
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb["TRACK"]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not rows:
            return lookup
        headers = rows[0]
        for r in rows[1:]:
            row_dict = dict(zip(headers, r))
            name = row_dict.get("TRACK NAME")
            direction = row_dict.get("DIRECTION")
            if name in (None, "") or direction in (None, ""):
                continue
            lookup[str(name).strip().upper()] = str(direction).strip().upper()
    except Exception:
        pass
    return lookup


def _is_track_direction_relay(name2_norm: str) -> bool:
    # CONFIRMED category B: TSPR (and TSPR1/TSPR2 variants) + TS(M)/TS(R)
    # (TS(A)/TS(B) already normalized to these by _normalize_relay_type).
    return name2_norm.startswith("TSPR") or name2_norm in ("TS(M)", "TS(R)")


def _is_lcpr_relay(name2_norm: str) -> bool:
    # CONFIRMED: LCPR (and LCPR1/LCPR2 variants) explicitly excluded from
    # category A's per-signal grouping - falls through to category C.
    return name2_norm.startswith("LCPR")


ROWS_PER_SHEET_PAGINATED = 32


class ContactAnalysisCancelled(Exception):
    """Raised when the user declines to proceed past a
    consistently-missing track relay pair confirmation prompt."""
    pass


def _read_signal_typical_lookup(xlsx_path: str) -> dict:
    """{signal_name: typical_name} from the SIGNAL sheet's own SIG
    NAME/TYPICAL columns - used to decide whether two signals sharing
    an externally-referenced relay (e.g. both pointing to "TO36" via
    ZR) should be merged into one group."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["SIGNAL"]
    lookup = {}
    for r in range(2, ws.max_row + 1):
        sig_name = ws.cell(row=r, column=1).value
        if sig_name is None or str(sig_name).strip() == "":
            continue
        typical_name = ws.cell(row=r, column=4).value or ""
        lookup[str(sig_name).strip()] = str(typical_name).strip()
    return lookup


def merge_shared_signal_groups(category_a_groups: dict, category_a_order: list, signal_typical_lookup: dict) -> None:
    """
    CONFIRMED: if an externally-referenced relay (e.g. "TO36", pulled in
    via a signal's own ZR column) is shared by MORE THAN ONE signal,
    merges its usage into ONE group keyed by the relay's OWN name -
    rather than splitting it across each referencing signal's own
    group. If only one signal references it, leaves the existing
    per-signal grouping as-is.

    CONFIRMED FIX: merging no longer requires the referencing signals to
    share the same TYPICAL. The relay's physical identity comes from
    its Relay Rack (NAME1, NAME2) position, not from what typical the
    signal referencing it via ZR happens to use - confirmed directly
    against a real workbook, where a shared relay (e.g. "2034"'s own
    DZR/HZR/HHZR) was correctly merged across two STANDARD-typical
    signals, but stayed wrongly duplicated across a STANDARD /
    STANDARD+ZRP boundary even though it's the exact same physical
    relay. signal_typical_lookup is accepted for backward compatibility
    but no longer consulted here.
    Modifies category_a_groups/category_a_order in place.
    """
    name1_to_owning_signals = {}
    for owning_signal, relay_keys in category_a_groups.items():
        for relay_key in relay_keys:
            name1_to_owning_signals.setdefault(relay_key[0], set()).add(owning_signal)

    for name1, owning_signals in name1_to_owning_signals.items():
        if len(owning_signals) < 2:
            continue  # only referenced by one signal - keep existing behavior

        merged_keys = set()
        for owning_signal in owning_signals:
            keys_to_move = {rk for rk in category_a_groups[owning_signal] if rk[0] == name1}
            category_a_groups[owning_signal] -= keys_to_move
            merged_keys |= keys_to_move

        # CONFIRMED priority order for where the merged usage lands:
        #   1. If the relay's own name IS one of the signals referencing
        #      it (the common case - e.g. "2034" is itself a real
        #      STANDARD signal AND is shared via ZR by "2046"), merge
        #      into THAT signal's own existing group.
        #   2. Otherwise (a relay with no SIGNAL row of its own, e.g. a
        #      hypothetical external relay only ever pulled in via ZR)
        #      - NEVER create a new standalone/orphan-named CA sheet for
        #      it. Fold it into one of the EXISTING referencing signals'
        #      groups instead, so the relay never gets its own repeated
        #      sheet. Prefer a STANDARD-typical signal over a
        #      STANDARD+ZRP one as the host (STANDARD+ZRP signals draw a
        #      repeater-panel copy of the base relay, so the STANDARD
        #      signal is the more natural "home"); fall back to a stable
        #      alphabetical choice if that doesn't disambiguate.
        if name1 in owning_signals:
            target_group = name1
        else:
            def _host_priority(sig):
                typical = signal_typical_lookup.get(sig, "")
                is_zrp = "ZRP" in typical.upper()
                return (is_zrp, str(sig))
            target_group = sorted(owning_signals, key=_host_priority)[0]

        if target_group not in category_a_groups:
            category_a_groups[target_group] = set()
            category_a_order.append(target_group)
        category_a_groups[target_group] |= merged_keys


def sort_track_relays_by_type_order(relays: list) -> list:
    """
    CONFIRMED (from Krish's own reference sheet): within a direction,
    Track relays are grouped by their own track name (NAME1, e.g.
    "1994", "2004"), preserving the order those track names first
    appear - and WITHIN each track name's group, the types are always
    ordered TS(M) -> TS(R) -> TSPR -> TSPR1 (if present), regardless of
    whatever order Relay Rack happens to list them in.
    """
    TYPE_PRIORITY = {"TS(M)": 0, "TS(R)": 1, "TSPR": 2, "TSPR1": 3}
    groups = {}
    group_order = []
    for r in relays:
        name1 = r["name1"].upper()
        if name1 not in groups:
            groups[name1] = []
            group_order.append(name1)
        groups[name1].append(r)
    sorted_relays = []
    for name1 in group_order:
        sorted_relays.extend(
            sorted(groups[name1], key=lambda r: TYPE_PRIORITY.get(r["name2_norm"].upper(), 99))
        )
    return sorted_relays


def generate_all_contact_analysis_sheets(
    dxf_folder: str,
    relay_rack_filenames: list,
    border_template_path: str,
    start_sheet_number,
    log,
    filename_to_circuit_type: dict = None,
    confirm_callback=None,
    xlsx_path: str = None,
    on_conflict=None,
):
    """
    dxf_folder: folder already containing every OTHER circuit type's
        just-saved .dxf files for this job (Relay Rack included).
    relay_rack_filenames: the specific filenames (within dxf_folder) that
        are Relay Rack sheets - everything else in the folder is treated
        as contact-usage source.
    filename_to_circuit_type: optional {final_name: "SIGNAL"/"TRACK"/etc.}
        map, so errors are labeled by circuit type - CONFIRMED requirement.
    confirm_callback: optional callable(message: str) -> bool. CONFIRMED:
        Track/SDF relays sometimes legitimately aren't placed in the
        Relay Rack at all (e.g. TS(M)/TS(R) pair, or TS(A)/TS(B) - the
        same pair under an alternate naming convention). If BOTH halves
        of such a pair are consistently missing, that's not necessarily
        an error - the user is asked (via this callback, which should
        pause the job and return the user's real choice) whether to
        proceed anyway. If only ONE half is missing (a partial
        mismatch), that stays a hard vital error - CONFIRMED. If no
        callback is given (e.g. called outside the Full IFC job), this
        just logs a warning and proceeds rather than blocking forever.
    xlsx_path: needed to read the TRACK sheet's DIRECTION column (used
        for the DN/UP grouping below) - optional so this still works if
        no track relays end up needing it.
    Returns a list of (out_name, doc, sht) tuples, same shape every other
    generate_all_* function returns, ready to merge into `pending`.
    Raises ContactAnalysisCancelled if the user declines to proceed.

    CONFIRMED sheet grouping (THREE categories):
      A. Signal-sourced STANDARD relays (any type except TSPR/TS(M)/
         TS(R)/LCPR): one sheet per signal, grouped by that signal's own
         consuming-circuit TITLE - ONLY relays whose usage actually comes
         from that signal's own SIGNAL-type circuit (not shared usage
         from Track/AR & AZR/Communication/etc.).
      B. TSPR/TS(M)/TS(R) relays (any signal, any circuit): grouped by
         DN vs UP direction, read from the TRACK sheet's own DIRECTION
         column (keyed by TRACK NAME = relay's NAME1) - NOT grouped by
         consuming circuit at all. Paginated 32/sheet within each
         direction.
      B excludes LCPR - CONFIRMED LCPR falls through to category C
      instead, grouped with everything else.
      C. Everything else (LCPR, AR & AZR/Communication/SDF/Data Logger/
         Custom-sourced relays, relays with no Signal usage): paginated
         32 relays per sheet, continuing onto more sheets as needed.
    Each sheet is titled with a simple incrementing counter
    ("CONTACT ANALYSIS - 1", "CONTACT ANALYSIS - 2"...) across ALL
    categories combined, in the order sheets are actually produced
    (A, then B-DN, then B-UP, then C) - CONFIRMED no fancy category
    label is computed.
    """
    filename_to_circuit_type = filename_to_circuit_type or {}
    all_paths = [os.path.join(dxf_folder, f) for f in os.listdir(dxf_folder) if f.lower().endswith(".dxf")]
    relay_rack_paths = [os.path.join(dxf_folder, f) for f in relay_rack_filenames]
    other_paths = [p for p in all_paths if p not in relay_rack_paths]

    roster_list = extract_relay_roster(relay_rack_paths)
    if not roster_list:
        log.warning("[RELAY RACK] No relays found - Contact Analysis skipped")
        return []
    roster_by_key = {(r["name1"].upper(), r["name2_norm"]): r for r in roster_list}

    usage, title_to_circuit_type = extract_contact_usage(other_paths, log, filename_to_circuit_type, roster_by_key, on_conflict)

    # CONFIRMED: flag any relay whose own coil (R1/R2) is never
    # referenced by ANY circuit anywhere - not a specific circuit type
    # (no group to attribute it to), just a general data-quality warning.
    for relay in roster_list:
        if _is_timer_relay(relay["relay_type"]):
            continue  # timer relays have no normal coil to check
        key = (relay["name1"].upper(), relay["name2_norm"])
        relay_usage = usage.get(key, {})
        if "R1" not in relay_usage and "R2" not in relay_usage:
            log.warning(f"[RELAY RACK] Coil never used: {relay['bit_name']}")

    # Group relay keys by which consuming circuit (sheet TITLE) used them.
    # Preserves first-discovery order for both groups and relays-within-a-
    # group (falls back to original rack order for relay ordering, below).
    group_relay_keys = {}
    group_order = []
    for relay_key, contacts in usage.items():
        for contact_code, (sht_num, consuming_circuit, wire_value, owning_signal, _r_pos) in contacts.items():
            if not consuming_circuit:
                continue
            if consuming_circuit not in group_relay_keys:
                group_relay_keys[consuming_circuit] = set()
                group_order.append(consuming_circuit)
            group_relay_keys[consuming_circuit].add(relay_key)

    # --- PASS 1: classify every missing relay across ALL groups, before
    # building any sheet - hard errors vs "consistently missing pair,
    # ask the user once" (CONFIRMED).
    consistent_missing = set()
    for consuming_circuit in group_order:
        missing_keys = group_relay_keys[consuming_circuit] - set(roster_by_key.keys())
        for key in missing_keys:
            name1, name2_norm = key
            partner_type = _TRACK_PAIR_PARTNER.get(name2_norm)
            if partner_type is not None and (name1, partner_type) not in roster_by_key:
                consistent_missing.add(key)
                continue
            circuit_type = title_to_circuit_type.get(consuming_circuit, consuming_circuit)
            log.error(f"[{circuit_type}] Relay not in Relay Rack: {key[0]}/{key[1]}")

    if consistent_missing and not log.has_errors:
        track_names = sorted(f"{k[0]} ({k[1]})" for k in consistent_missing)
        message = (
            "These track relays are not placed in the Relay Rack (both halves of "
            "each pair are consistently absent, not a partial mismatch):\n\n"
            + "\n".join(track_names)
            + "\n\nThis can be normal if these tracks intentionally live off-rack. "
            "Proceed with generation anyway? (These relays will be left off Contact Analysis.)"
        )
        if confirm_callback is not None:
            if not confirm_callback(message):
                raise ContactAnalysisCancelled(
                    "User chose not to proceed past missing track relay(s): "
                    + ", ".join(track_names)
                )
        else:
            for name in track_names:
                log.warning(f"Track relay not in Relay Rack (proceeding): {name}")

    # --- PASS 2: classify every relay-with-usage into one of 3
    # categories (CONFIRMED), then build sheets category by category.
    track_direction_lookup = _read_track_direction_lookup(xlsx_path) if xlsx_path else {}

    category_a_groups = {}   # signal_title -> set of relay_keys
    category_a_order = []
    category_b_dn = set()
    category_b_up = set()
    category_b_unknown = set()   # track-type relay, but no DIRECTION found
    category_c_keys = set()

    for relay_key in usage.keys():
        if relay_key not in roster_by_key:
            continue  # already logged as missing/consistent-missing above
        name1, name2_norm = relay_key

        if _is_track_direction_relay(name2_norm):
            direction = track_direction_lookup.get(name1.upper())
            if direction == "DN":
                category_b_dn.add(relay_key)
            elif direction == "UP":
                category_b_up.add(relay_key)
            else:
                category_b_unknown.add(relay_key)
            continue

        if _is_lcpr_relay(name2_norm):
            category_c_keys.add(relay_key)
            continue

        signal_owners = {
            owning_signal
            for _code, (_sht, consuming_circuit, _wire, owning_signal, _rp) in usage[relay_key].items()
            if consuming_circuit and title_to_circuit_type.get(consuming_circuit) == "SIGNAL" and owning_signal
        }
        if signal_owners:
            # CONFIRMED FIX: group by the OWNING signal (embedded per
            # sheet by signal_core.py) rather than the relay's own NAME1.
            # This correctly keeps an externally-referenced repeater
            # relay (e.g. "TO36" HZR/HHZR/DZR, drawn via the "$"/ZR
            # substitution on one of "F565"'s own STANDARD sheets) grouped
            # WITH F565's own relays, since it belongs to F565's typical
            # batch - even though the relay itself is named "TO36", not
            # "F565". A relay could in principle belong to more than one
            # signal's batch, so it can land in more than one group.
            for group_key in signal_owners:
                if group_key not in category_a_groups:
                    category_a_groups[group_key] = set()
                    category_a_order.append(group_key)
                category_a_groups[group_key].add(relay_key)
        else:
            category_c_keys.add(relay_key)

    if category_b_unknown:
        names = sorted(f"{k[0]} {k[1]}" for k in category_b_unknown)
        log.warning(f"[TRACK] No DIRECTION found, added to remaining: {', '.join(names)}")
        category_c_keys |= category_b_unknown

    # CONFIRMED: if an externally-referenced relay (e.g. "TO36", pulled
    # in via a signal's own ZR column) is shared by MORE THAN ONE
    # signal, and those signals use the SAME typical, merge its usage
    # into ONE group keyed by the relay's OWN name - rather than
    # splitting it across each referencing signal's own group. If the
    # signals use DIFFERENT typicals, or only one signal references it,
    # leave the existing per-signal grouping as-is.
    if xlsx_path:
        signal_typical_lookup = _read_signal_typical_lookup(xlsx_path)
        merge_shared_signal_groups(category_a_groups, category_a_order, signal_typical_lookup)

        # Drop any owning-signal groups that ended up empty after moving
        # their shared relays out into the relay's own group.
        category_a_order = [g for g in category_a_order if category_a_groups.get(g)]

    results = []
    sht = int(start_sheet_number)
    counter = 1

    def _relays_for(relay_keys):
        return [r for r in roster_list if (r["name1"].upper(), r["name2_norm"]) in relay_keys]

    def _build_sheet(relays, circuit_type_for_warnings):
        nonlocal sht, counter
        qn1_relays = [
            r for r in relays
            if not _is_qecx61(r["relay_type"]) and not _is_timer_relay(r["relay_type"])
            and r["cntcnf"].upper() == "8F-8B"
        ]
        qecx_relays = [r for r in relays if _is_qecx61(r["relay_type"]) and not _is_timer_relay(r["relay_type"])]

        doc = ezdxf.new()

        if qn1_relays:
            _place_header(doc, HEADER_QN1_POINT)
            for i, relay in enumerate(qn1_relays):
                y = FIRST_QN1_ROW_POINT[1] - i * ROW_SPACING
                _place_row(doc, (FIRST_QN1_ROW_POINT[0], y), relay, usage, log)

        if qecx_relays:
            _place_header(doc, HEADER_QECX_POINT)
            for i, relay in enumerate(qecx_relays):
                y = FIRST_QECX_ROW_POINT[1] - i * ROW_SPACING
                _place_row(doc, (FIRST_QECX_ROW_POINT[0], y), relay, usage, log)

        # CONFIRMED: TIMER relays (RELTYP="TIMER", e.g. CNTCNF="10-60 SEC")
        # don't have a normal front/back contact bank - they're
        # intentionally left OFF the Contact Analysis chart entirely, no
        # warning needed (this is a known, expected relay type, not an
        # unsupported one). Relay Position still updates their R(POS)
        # separately - that's unaffected by this exclusion.
        for relay in relays:
            if relay in qn1_relays or relay in qecx_relays or _is_timer_relay(relay["relay_type"]):
                continue
            log.warning(f"[{circuit_type_for_warnings}] No typical for config: {relay['bit_name']} ({relay['cntcnf']})")

        next_sht = sht + 1
        insert_border_title(
            doc,
            border_template_path,
            template_name="CONTACT_ANALYSIS",
            sig_name="", ahead_sig="", loc="",
            sht_number=str(sht),
            cont_number=str(next_sht),
            title_text_override=f"{TITLE_TEXT_PREFIX}{counter}",
        )

        out_name = f"CONTACT_ANALYSIS_SHT{sht}.dxf"
        results.append((out_name, doc, str(sht)))
        sht = next_sht
        counter += 1

    # Category A: one sheet per signal (uncapped - a signal's own relay
    # count is expected to fit).
    for signal_name in category_a_order:
        relays = _relays_for(category_a_groups[signal_name])
        if not relays:
            continue
        _build_sheet(relays, "SIGNAL")

    # Category B: DN then UP, each paginated 32/sheet.
    for direction_keys, direction_label in ((category_b_dn, "DN"), (category_b_up, "UP")):
        relays = sort_track_relays_by_type_order(_relays_for(direction_keys))
        for idx in range(0, len(relays), ROWS_PER_SHEET_PAGINATED):
            chunk = relays[idx: idx + ROWS_PER_SHEET_PAGINATED]
            _build_sheet(chunk, f"TRACK {direction_label}")

    # Category C: everything else, paginated 32/sheet.
    remaining_relays = _relays_for(category_c_keys)
    for idx in range(0, len(remaining_relays), ROWS_PER_SHEET_PAGINATED):
        chunk = remaining_relays[idx: idx + ROWS_PER_SHEET_PAGINATED]
        _build_sheet(chunk, "REMAINING")

    return results
