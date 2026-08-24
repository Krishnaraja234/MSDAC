"""
MSDAC Signal Circuit Generator - core substitution logic.

Reads one row from the SIGNAL sheet of MSDAC_FAR.xlsx and produces a set
of DXF sheets by substituting placeholder tokens in the matching typical
template files, then inserting a separately-uploaded border/title-block
template at (0,0) into each sheet.

CONFIRMED placeholder rules (per row, columns: A=SIG NAME, B=DIRECTION,
C=LOC, D=TYPICAL, E=LCPR, F=ZR):

  - '*' token (on S_NAME attributes, and inside TITLE/MTEXT strings)
        -> replaced with A2 (SIG NAME)
  - '$' token (on S_NAME attributes, and inside TITLE/MTEXT strings)
        -> replaced with F2 (ZR)
  - literal text "LOCATION" appearing in MTEXT/TEXT entities
        -> replaced with C2 (LOC)
  - '@' token (on S_NAME attributes, LCPR/AHPR blocks only)
        -> replaced with E2 (LCPR)
  - FUZEBLOCK / FUZEENDBLOCK VOLT attribute's literal '#'
        -> replaced with B2 (DIRECTION)

Border/title-block template (separate file, uploaded by user):
  - No longer baked into each typical. Instead, every block referenced by
    an INSERT in the border template's modelspace (e.g. TEMPHUT1, TITLE)
    is imported into the generated sheet and re-inserted at the SAME
    insertion point it had in the border template (normally (0,0)).
  - The TITLE block's SHT / CONT attributes carry the current/next sheet
    number - running numbers assigned sequentially by the caller.
  - The TITLE attribute's text still contains the '$'/'*' token and gets
    the same substitution as elsewhere.

Starting sheet number for Signal circuits is NOT hardcoded - it's read
from the FIELD PG.NO sheet, cell C6 (the "Sheet Number" value on the
SIGNAL row of that index sheet).
"""

import ezdxf
from ezdxf.addons import importer
import os
import re

TEMPLATES_DIR = os.path.join(os.path.dirname(__file__), "dxf_templates")

_TEMPLATE_PATH_CACHE = {}


def resolve_template_path(filename: str) -> str:
    """
    CONFIRMED: templates now live organized into subfolders by circuit
    type (dxf_templates/SIGNAL/, dxf_templates/AR_AZR/, dxf_templates/
    COMMUNICATION/, dxf_templates/"CONTACT ANALYSIS"/, dxf_templates/
    "DATA LOGGER"/, dxf_templates/"RELAY RACK"/, dxf_templates/SDF/,
    dxf_templates/TSPR/) rather than sitting flat directly in
    dxf_templates/. This searches for a given filename across TEMPLATES_DIR
    itself (flat, for backward compatibility) and every subfolder under
    it, and returns whichever path actually exists. Cached per filename
    since the folder layout doesn't change during a run.
    """
    if filename in _TEMPLATE_PATH_CACHE:
        return _TEMPLATE_PATH_CACHE[filename]

    flat_path = os.path.join(TEMPLATES_DIR, filename)
    if os.path.isfile(flat_path):
        _TEMPLATE_PATH_CACHE[filename] = flat_path
        return flat_path

    if os.path.isdir(TEMPLATES_DIR):
        filename_lower = filename.lower()
        for entry in sorted(os.listdir(TEMPLATES_DIR)):
            subdir = os.path.join(TEMPLATES_DIR, entry)
            if os.path.isdir(subdir):
                candidate = os.path.join(subdir, filename)
                if os.path.isfile(candidate):
                    _TEMPLATE_PATH_CACHE[filename] = candidate
                    return candidate
                # CONFIRMED: case-insensitive fallback, so this works
                # reliably even on case-sensitive filesystems (Windows is
                # already case-insensitive, but this keeps behavior
                # consistent everywhere).
                for actual_name in os.listdir(subdir):
                    if actual_name.lower() == filename_lower:
                        candidate = os.path.join(subdir, actual_name)
                        _TEMPLATE_PATH_CACHE[filename] = candidate
                        return candidate

    raise FileNotFoundError(
        f"Template file '{filename}' not found in {TEMPLATES_DIR} or any of its subfolders."
    )

TOKEN_STAR = "*"     # -> SIG NAME
TOKEN_DOLLAR = "$"   # -> ZR
TOKEN_AT = "@"       # -> LCPR
LITERAL_LOCATION = "LOCATION"  # -> LOC (matched as substring, case-sensitive)

# CONFIRMED: TYPICAL_SHEET_SETS and TITLE_TEXT_MAP are now driven by a
# dedicated Excel sheet ("TYPICAL CONFIG"), not hardcoded here - this lets
# the order, file names, and title text for every typical be controlled
# entirely from Excel. The dicts below remain as a FALLBACK only, used if
# the workbook has no TYPICAL CONFIG sheet at all (e.g. older workbooks).
TYPICAL_SHEET_SETS = {
    "HZR+HHZR+DZR": ["HZR_HHZR_DZR.dxf"],
    "STANDARD": ["STANDARD1.dxf", "STANDARD2.dxf", "STANDARD3.dxf", "STANDARD4.dxf"],
    "STANDARD+ZRP": ["STANDARD1.dxf", "STANDARD2_ZRP.dxf", "STANDARD3.dxf", "STANDARD4.dxf"],
}


def read_typical_config(xlsx_path: str):
    """
    Reads the 'TYPICAL CONFIG' sheet: columns
      TYPICAL | ORDER | FILE NAME | TITLE TEXT | LCPR FILE NAME | LCPR ONLY | DROP IF LCPR | LCPR TITLE TEXT

    Multiple rows share the same TYPICAL value, each describing ONE sheet
    in that typical's sequence (ordered by the ORDER column).

      LCPR FILE NAME: if filled, this row's file gets swapped to this
        filename when the LCPR column is filled for a given signal row
        (replaces the hardcoded LCPR_SWAP dict for this file).
      LCPR ONLY: 'YES'/TRUE - this row is INCLUDED ONLY when LCPR is
        filled (e.g. the extra 5th STANDARD sheet); omitted otherwise.
      DROP IF LCPR: 'YES'/TRUE - this row is EXCLUDED when LCPR is filled
        (e.g. STANDARD+ZRP's dropped 4th sheet).
      LCPR TITLE TEXT: optional - if the LCPR-swapped file needs its own
        different title (not the same as the original row's TITLE TEXT),
        put it here. Left blank, the swapped file just reuses the
        original row's TITLE TEXT.

    Returns (sheet_sets, title_map, lcpr_config):
      sheet_sets: {typical_name: [file1, file2, ...]} (ordered by ORDER,
        WITHOUT any LCPR-only/drop-if-lcpr filtering applied yet - that
        happens in resolve_sheet_set using lcpr_config)
      title_map: {file_name: title_text}
      lcpr_config: {typical_name: {"swap": {file: lcpr_file}, "lcpr_only": {file,...}, "drop_if_lcpr": {file,...}}}

    Falls back to the hardcoded TYPICAL_SHEET_SETS/TITLE_TEXT_MAP/LCPR
    rules (returning lcpr_config=None) if the workbook has no TYPICAL
    CONFIG sheet, or has none of the LCPR-related columns filled in.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "TYPICAL CONFIG" not in wb.sheetnames:
        return dict(TYPICAL_SHEET_SETS), dict(TITLE_TEXT_MAP), None

    ws = wb["TYPICAL CONFIG"]
    rows = []
    any_lcpr_column_filled = False
    for r in range(2, ws.max_row + 1):
        typical = ws.cell(row=r, column=1).value
        order = ws.cell(row=r, column=2).value
        file_name = ws.cell(row=r, column=3).value
        title_text = ws.cell(row=r, column=4).value
        lcpr_file_name = ws.cell(row=r, column=5).value
        lcpr_only = ws.cell(row=r, column=6).value
        drop_if_lcpr = ws.cell(row=r, column=7).value
        # CONFIRMED: optional 8th column - if the LCPR-swapped file needs
        # its own different title text (not just reusing the original
        # row's title), it goes here. Falls back to the same TITLE TEXT
        # as the original row if left blank, since often it's the same
        # circuit, just an LCPR variant.
        lcpr_title_text = ws.cell(row=r, column=8).value
        if typical is None or file_name is None:
            continue
        if lcpr_file_name or lcpr_only or drop_if_lcpr:
            any_lcpr_column_filled = True
        rows.append({
            "TYPICAL": str(typical).strip(),
            "ORDER": int(order) if order is not None else len(rows) + 1,
            "FILE_NAME": str(file_name).strip(),
            "TITLE_TEXT": str(title_text) if title_text is not None else "",
            "LCPR_FILE_NAME": str(lcpr_file_name).strip() if lcpr_file_name else None,
            "LCPR_ONLY": str(lcpr_only).strip().upper() in ("YES", "TRUE", "1") if lcpr_only else False,
            "DROP_IF_LCPR": str(drop_if_lcpr).strip().upper() in ("YES", "TRUE", "1") if drop_if_lcpr else False,
            "LCPR_TITLE_TEXT": str(lcpr_title_text) if lcpr_title_text is not None else None,
        })

    if not rows:
        return dict(TYPICAL_SHEET_SETS), dict(TITLE_TEXT_MAP), None

    # CONFIRMED: sort WITHIN each typical's own group, not globally - more
    # robust when different typicals have different numbers of rows.
    by_typical = {}
    for row in rows:
        by_typical.setdefault(row["TYPICAL"], []).append(row)

    sheet_sets = {}
    title_map = {}
    lcpr_config = {} if any_lcpr_column_filled else None
    for typical, typical_rows in by_typical.items():
        typical_rows.sort(key=lambda r: r["ORDER"])
        sheet_sets[typical] = [r["FILE_NAME"] for r in typical_rows]
        for r in typical_rows:
            title_map[r["FILE_NAME"]] = r["TITLE_TEXT"]
            # CONFIRMED FIX: the LCPR-swapped file gets its OWN title_map
            # entry too - previously it had none at all, so a swapped-in
            # LCPR file's title lookup would silently miss. Uses the
            # dedicated LCPR_TITLE_TEXT column if given, otherwise falls
            # back to the same title as the original (non-LCPR) row.
            if r["LCPR_FILE_NAME"]:
                title_map[r["LCPR_FILE_NAME"]] = r["LCPR_TITLE_TEXT"] if r["LCPR_TITLE_TEXT"] else r["TITLE_TEXT"]
        if lcpr_config is not None:
            lcpr_config[typical] = {
                "swap": {r["FILE_NAME"]: r["LCPR_FILE_NAME"] for r in typical_rows if r["LCPR_FILE_NAME"]},
                "lcpr_only": {r["FILE_NAME"] for r in typical_rows if r["LCPR_ONLY"]},
                "drop_if_lcpr": {r["FILE_NAME"] for r in typical_rows if r["DROP_IF_LCPR"]},
            }

    return sheet_sets, title_map, lcpr_config

# When the LCPR column is filled, these sheets get swapped for their _LCPR variant.
LCPR_SWAP = {
    "STANDARD2.dxf": "STANDARD2_LCPR.dxf",
    "STANDARD3.dxf": "STANDARD3_LCPR.dxf",
    "STANDARD4.dxf": "STANDARD4_LCPR.dxf",
}

# When LCPR is filled AND the typical is STANDARD-based, this extra sheet
# gets APPENDED on top of the (already-swapped) 4-sheet set - it's an
# additional 5th sheet, not a replacement of any existing one.
# CONFIRMED: STANDARD+ZRP does NOT get this extra 5th sheet, and also
# drops the 4th sheet (STANDARD4_LCPR) entirely when LCPR is filled.
LCPR_EXTRA_SHEET = "STANDARD5_LCPR.dxf"
TYPICALS_THAT_GET_LCPR_EXTRA_SHEET = {"STANDARD"}
TYPICALS_THAT_DROP_4TH_SHEET_WITH_LCPR = {"STANDARD+ZRP"}


def resolve_sheet_set(typical: str, lcpr_filled: bool, sheet_sets_override: dict = None, lcpr_config: dict = None) -> list[str]:
    """Return the ordered list of template filenames for a given TYPICAL value."""
    typical = (typical or "").strip()
    sheet_sets = sheet_sets_override if sheet_sets_override is not None else TYPICAL_SHEET_SETS
    base_set = sheet_sets.get(typical)
    if base_set is None:
        raise ValueError(f"Unknown TYPICAL value: {typical!r}")

    sheets = list(base_set)

    # CONFIRMED: if this typical has Excel-driven LCPR config, use it
    # entirely (swap/lcpr_only/drop_if_lcpr), replacing the hardcoded rules.
    if lcpr_config is not None and typical in lcpr_config:
        cfg = lcpr_config[typical]
        if lcpr_filled:
            sheets = [s for s in sheets if s not in cfg["drop_if_lcpr"]]
            sheets = [cfg["swap"].get(s, s) for s in sheets]
            extra_files = [f for f in cfg["lcpr_only"] if f not in sheets]
            sheets.extend(extra_files)
        else:
            sheets = [s for s in sheets if s not in cfg["lcpr_only"]]
        return sheets

    if lcpr_filled:
        sheets = [LCPR_SWAP.get(s, s) for s in sheets]
        if typical in TYPICALS_THAT_DROP_4TH_SHEET_WITH_LCPR:
            sheets = sheets[:3]  # CONFIRMED: drop the 4th sheet entirely, no 5th either
        elif typical in TYPICALS_THAT_GET_LCPR_EXTRA_SHEET:
            sheets.append(LCPR_EXTRA_SHEET)
    return sheets


def _safe_str(value) -> str:
    """None/blank Excel cells should substitute to empty string, not 'None'."""
    if value is None:
        return ""
    return str(value)


def _substitute_text_value(text: str, sig_name, ahead_sig, loc, hut_name_input=None) -> str:
    """Apply token/literal substitutions inside a string (TITLE, MTEXT, TEXT)."""
    if text is None:
        return text
    new_text = text
    if TOKEN_STAR in new_text:
        new_text = new_text.replace(TOKEN_STAR, _safe_str(sig_name))
    if TOKEN_DOLLAR in new_text:
        new_text = new_text.replace(TOKEN_DOLLAR, _safe_str(ahead_sig))

    # Specific static hut-location patterns found in existing typicals.
    if "MSDAC-HUT" in new_text:
        new_text = new_text.replace("MSDAC-HUT", f"{_safe_str(loc)} HUT")
    if "HUT NAME" in new_text:
        # CONFIRMED: this specific keyword now uses the user-entered
        # global HUT Name input field (from the main page), NOT the
        # per-row LOC column - a separate, distinct value from the rest
        # of this function's loc-based substitutions.
        new_text = new_text.replace("HUT NAME", _safe_str(hut_name_input if hut_name_input is not None else loc))

    if LITERAL_LOCATION in new_text:
        new_text = new_text.replace(LITERAL_LOCATION, _safe_str(loc))

    # Standalone word "LOC" (e.g. future "HUT LOC" template convention).
    # Word-bounded so it never matches inside "LOCATION" or other words.
    new_text = re.sub(r"\bLOC\b", _safe_str(loc), new_text)

    return new_text


def substitute_signal_sheet(
    template_path: str,
    sig_name,
    direction: str,
    loc,
    ahead_sig,
    lcpr_value,
    hut_name_input=None,
    is_dollar_relay_template=False,
):
    """
    Load a typical template DXF and substitute placeholders for one signal
    instance. The typicals still physically contain their OWN baked-in
    TITLE block (leftover from before the border-template workflow) - that
    block is explicitly removed here, since TITLE/SHT/CONT are now owned
    exclusively by insert_border_title().
    Returns an ezdxf Drawing object (modified in place, ready to save).
    """
    doc = ezdxf.readfile(template_path)
    msp = doc.modelspace()

    to_delete = []

    # CONFIRMED WORKAROUND: STANDARD4.dxf, HZR_HHZR_DZR.dxf, and
    # STANDARD5_LCPR.dxf (and any custom-typical equivalent, detected by
    # the caller via title_map_override) all draw the ZR-REFERENCED
    # relay's own circuit, not this signal's own - but their S_NAME
    # attdefs are literally "*" in the template file itself (should be
    # "$"), so the normal TOKEN_STAR/TOKEN_DOLLAR substitution below
    # would incorrectly tag every relay on this sheet with the CURRENT
    # signal's own name instead of the relay it's actually drawing. This
    # is the real root cause of the same relay (e.g. "2024"'s own
    # HZR/HHZR/DZR) getting a DIFFERENT, meaningless name (e.g. "2034")
    # depending on which signal's ZR happened to reference it - which
    # meant it could never be recognized as the same relay and merged.
    # Force S_NAME to use ahead_sig for these templates specifically,
    # regardless of the template's own (incorrect) literal placeholder
    # character. is_dollar_relay_template is determined by the caller
    # (generate_signal_drawing_set), which has access to
    # title_map_override for custom-typical robustness.

    for e in msp:
        if e.dxftype() == "TEXT":
            e.dxf.text = _substitute_text_value(e.dxf.text, sig_name, ahead_sig, loc, hut_name_input)
            continue
        if e.dxftype() == "MTEXT":
            e.text = _substitute_text_value(e.text, sig_name, ahead_sig, loc, hut_name_input)
            continue

        if e.dxftype() != "INSERT":
            continue

        if e.dxf.name == "TITLE":
            # Old baked-in title block - the border template supplies its
            # own TITLE block now, so drop this one to avoid duplicates.
            to_delete.append(e)
            continue

        # CONFIRMED FIX: the "$"-titled sheets (STANDARD1.dxf,
        # HZR_HHZR_DZR.dxf, STANDARD4.dxf, STANDARD5_LCPR.dxf) draw a
        # MIX of block types, not one uniform kind - the actual shared
        # coil (Relay_Coil/Relay_Coil_ACI/COIL_REPEATER blocks whose
        # R_NAME is the base type DZR/HZR/HHZR, or its "P"-suffixed
        # STANDARD+ZRP variant DZRP/HZRP/HHZRP, plus that coil's own
        # direct front-contacts) is genuinely the ZR-REFERENCED relay
        # and must be tagged with ahead_sig. But derived-contact blocks
        # on the SAME sheet (RECR, DECPR, HECPR, HHECPR, HRP1, HRP2,
        # HHRP1, HHRP2, DRP1, etc.) are each the SIGNAL'S OWN separate
        # relay, confirmed directly against the Relay Rack sheet (e.g.
        # "2024 RECR" is registered under 2024 itself, never under its
        # ZR target) - blanket-forcing those to ahead_sig was wrong and
        # caused their Relay Rack lookups to fail. Determine this once
        # per block from its own R_NAME, not from the sheet as a whole.
        _r_name_for_block = ""
        for _a in e.attribs:
            if _a.dxf.tag == "R_NAME":
                _r_name_for_block = (_a.dxf.text or "").strip().upper()
                break
        _block_is_shared_coil = _r_name_for_block in (
            "DZR", "HZR", "HHZR", "DZRP", "HZRP", "HHZRP"
        )

        for att in e.attribs:
            tag = att.dxf.tag
            val = att.dxf.text

            if tag == "S_NAME":
                if is_dollar_relay_template and _block_is_shared_coil:
                    ahead_sig_filled = ahead_sig is not None and str(ahead_sig).strip() != ""
                    att.dxf.text = _safe_str(ahead_sig) if ahead_sig_filled else _safe_str(sig_name)
                elif val == TOKEN_STAR:
                    att.dxf.text = _safe_str(sig_name)
                elif val == TOKEN_DOLLAR:
                    att.dxf.text = _safe_str(ahead_sig)
                elif val == TOKEN_AT:
                    att.dxf.text = _safe_str(lcpr_value)

            elif tag == "VOLT" and "#" in (val or ""):
                att.dxf.text = val.replace("#", _safe_str(direction))

            elif val and (
                LITERAL_LOCATION in val
                or TOKEN_STAR in val
                or TOKEN_DOLLAR in val
                or "MSDAC-HUT" in val
                or "HUT NAME" in val
                or re.search(r"\bLOC\b", val)
            ):
                att.dxf.text = _substitute_text_value(val, sig_name, ahead_sig, loc, hut_name_input)

    for e in to_delete:
        msp.delete_entity(e)

    return doc


# Per-typical TITLE text. This used to live inside each typical's own
# baked-in TITLE block; now that TITLE is supplied externally by the
# border template, the correct descriptive text has to be selected here
# based on which typical template file is being generated. The '$'/'*'
# token inside each string still gets substituted the same way as
# everywhere else (per CONFIRMED placeholder rules above).
TITLE_TEXT_MAP = {
    "HZR_HHZR_DZR.dxf": "$ - HZR, HHZR, DZR CIRCUITS ",
    "STANDARD1.dxf": "* - LIGHTING CIRCUITS ",
    "STANDARD2.dxf": "* - HR, HHR & DR CIRCUITS ",
    "STANDARD2_ZRP.dxf": "* - HR, HHR & DR CIRCUITS ",
    "STANDARD2_LCPR.dxf": "* - HR, HHR & DR CIRCUITS ",
    "STANDARD3.dxf": "* - DECPR, HHECPR & HECPR CIRCUITS & EXTERNAL REPEATER CIRCUITS ",
    "STANDARD3_LCPR.dxf": "* - DECPR, HHECPR & HECPR CIRCUITS & A LIGHT MARKER CIRCUITS ",
    "STANDARD4.dxf": "$ - HZR, HHZR, DZR CIRCUITS ",
    "STANDARD4_LCPR.dxf": "* - EXTERNAL REPEATER CIRCUITS ",
    "STANDARD5_LCPR.dxf": "$ - HZR, HHZR, DZR CIRCUITS ",
}


def insert_border_title(
    doc,
    border_template_path: str,
    template_name: str,
    sig_name,
    ahead_sig,
    loc,
    sht_number: str,
    cont_number: str,
    title_text_override: str = None,
):
    """
    Import the ENTIRE border/title-block template as a NEW block definition
    (border lines, hut label, title block - everything, including the
    SHT/CONT/TITLE ATTDEF entities that now sit directly in the template's
    modelspace rather than nested inside their own sub-block), then insert
    ONE block reference to it at (0,0) into the sheet's modelspace.

    ezdxf's add_auto_attribs() reads the ATTDEF entities baked into the new
    block definition and creates matching ATTRIB entities on the single
    INSERT, so the whole border stays one selectable block while still
    carrying this sheet's own SHT/CONT/TITLE values.

    title_text_override: if given, used verbatim as the title's source
    text (still passed through _substitute_text_value) instead of looking
    it up in TITLE_TEXT_MAP by template_name. Callers outside signal_core
    (e.g. track_core) use this to avoid mutating the shared TITLE_TEXT_MAP,
    which isn't safe under concurrent generation jobs.
    """
    border_doc = ezdxf.readfile(border_template_path)

    block_name = "TITLEBLOCK"
    new_block = doc.blocks.new(name=block_name, base_point=(0, 0))

    imp = importer.Importer(border_doc, doc)
    imp.import_entities(border_doc.modelspace(), target_layout=new_block)
    imp.finalize()

    if title_text_override is not None:
        title_text = title_text_override
    else:
        title_text = TITLE_TEXT_MAP.get(template_name)
        if title_text is None:
            raise ValueError(
                f"No TITLE text mapping found for typical template '{template_name}'. "
                "Add it to TITLE_TEXT_MAP in signal_core.py."
            )
    final_title = _substitute_text_value(title_text, sig_name, ahead_sig, loc)

    blockref = doc.modelspace().add_blockref(block_name, (0, 0))
    attrib_values = {
        "SHT": str(sht_number),
        "CONT": str(cont_number),
        "TITLE": final_title,
    }
    blockref.add_auto_attribs(attrib_values)

    # The TITLE (and possibly other) ATTDEF may be a MULTI-LINE attribute
    # (backed by an embedded MTEXT sub-object). Just setting .dxf.text
    # (which add_auto_attribs does) doesn't update that embedded MTEXT, so
    # BricsCAD/AutoCAD shows it blank and can even auto-rename the tag
    # (e.g. "TITLE" -> "TITLE_001") during their attribute audit. Converting
    # to a plain single-line attribute avoids this entirely.
    for att in blockref.attribs:
        if att.has_embedded_mtext_entity:
            att.discard_mtext()
        if att.dxf.tag in attrib_values:
            att.dxf.text = attrib_values[att.dxf.tag]

    return doc


def generate_signal_drawing_set(
    row: dict, start_sheet_number: int, border_template_path: str,
    sheet_sets_override: dict = None, title_map_override: dict = None, lcpr_config: dict = None,
    hut_name_input=None, registry=None,
):
    """
    row: dict with keys SIG NAME, DIRECTION, LOC, TYPICAL, LCPR, ZR
    start_sheet_number: int, the sheet number the first sheet in this set should carry
    border_template_path: path to the user-uploaded border/title-block DXF
    sheet_sets_override / title_map_override: from read_typical_config(xlsx_path),
        if the 'TYPICAL CONFIG' Excel sheet is present - gives full Excel-side
        control over typical order, file names, and title text.
    registry: optional ContactRegistry - CONFIRMED FIX: when provided
        (currently only Full IFC creates one), registers every contact/
        coil block's usage so a genuine repeated-contact conflict halts
        generation with a clear error, same protection AR&AZR/SDF
        already have. CONFIRMED semantics (from Krish, verified against
        real data): the same (S_NAME, R_NAME) pair, and even the same
        (S_NAME, R_NAME, letter), legitimately repeats many times
        across a signal's own sheets - that's normal template
        structure, not a conflict. The real uniqueness key is the FULL
        combination of S_NAME + R_NAME + letter (C) + number-pair
        (F/A on front contacts, A/B on back contacts) together - that
        exact combination must never repeat on any sheet. Coil blocks
        (Relay_Coil/Relay_Coil_ACI) have no letter/number tags, so
        they're registered by (S_NAME, R_NAME) alone.

    Returns: (list of (output_filename, ezdxf.Drawing, sht_number_str), next_sheet_number)
    """
    sig_name = row.get("SIG NAME")
    direction = row.get("DIRECTION") or ""
    loc = row.get("LOC") or ""
    typical = row.get("TYPICAL") or ""
    lcpr = row.get("LCPR")
    ahead_sig = row.get("ZR")
    lcpr_filled = bool(lcpr and str(lcpr).strip())

    if typical.strip().upper().startswith("STANDARD"):
        sig_name_str = _safe_str(sig_name).strip()
        ahead_sig_str = _safe_str(ahead_sig).strip()
        if sig_name_str and sig_name_str == ahead_sig_str:
            raise ValueError(
                f"SIG NAME ('{sig_name}') and ZR ('{ahead_sig}') are identical for this row "
                f"(TYPICAL={typical}). This is almost always a data-entry mistake - please "
                "check the SIGNAL sheet and correct the ZR value for this row."
            )

    sheet_files = resolve_sheet_set(typical, lcpr_filled, sheet_sets_override=sheet_sets_override, lcpr_config=lcpr_config)

    results = []
    sheet_num = start_sheet_number
    for template_name in sheet_files:
        template_path = resolve_template_path(template_name)
        sht = f"{sheet_num:03d}"
        cont = f"{sheet_num + 1:03d}"

        # CONFIRMED: detect "$"-based relay-circuit templates (which draw
        # the ZR-REFERENCED relay's own circuit, not this signal's own)
        # via BOTH the known default filenames AND the resolved title text
        # (from title_map_override, if the typical config customizes
        # which file is used for this slot) - covers custom typicals
        # like STANDARD+ZRP where a Excel-side override might substitute
        # a differently-named file for the same "$"-type sheet.
        title_for_template = (title_map_override or TITLE_TEXT_MAP).get(template_name, "")
        is_dollar_relay_template = (
            template_name in ("HZR_HHZR_DZR.dxf", "STANDARD4.dxf", "STANDARD5_LCPR.dxf")
            or title_for_template.strip().startswith("$")
        )

        doc = substitute_signal_sheet(
            template_path=template_path,
            sig_name=sig_name,
            direction=direction,
            loc=loc,
            ahead_sig=ahead_sig,
            lcpr_value=lcpr if lcpr_filled else "",
            hut_name_input=hut_name_input,
            is_dollar_relay_template=is_dollar_relay_template,
        )

        insert_border_title(
            doc=doc,
            border_template_path=border_template_path,
            template_name=template_name,
            sig_name=sig_name,
            ahead_sig=ahead_sig,
            loc=loc,
            sht_number=sht,
            cont_number=cont,
            title_text_override=title_map_override.get(template_name) if title_map_override else None,
        )

        # CONFIRMED FIX: carries the OWNING signal's own name through to
        # the generated sheet as a hidden extra attribute - Contact
        # Analysis groups a signal's own STANDARD-typical relays by THIS
        # value (the whole typical batch, e.g. all 4-5 STANDARD sheets
        # for one signal), not by each individual relay's own NAME1.
        # BUT: templates whose title starts with "$" (HZR_HHZR_DZR.dxf,
        # STANDARD4.dxf, STANDARD5_LCPR.dxf) draw the ZR-REFERENCED
        # relay's own circuit, not this signal's own content - so their
        # true owner is ahead_sig (the ZR value), not sig_name. Tagging
        # these with sig_name was the actual bug: it made the SAME
        # relay (e.g. signal "2024"'s own HZR/HHZR/DZR) look like it had
        # TWO different owners depending on which signal's row happened
        # to reference it via ZR, so Contact Analysis's merge-by-shared-
        # relay logic never recognized them as the same relay and listed
        # it on both signals' CA sheets instead of merging into one.
        for e in doc.modelspace():
            if e.dxftype() == "INSERT" and e.dxf.name in ("TITLE", "TITLEBLOCK"):
                e.add_attrib("OWNING_SIGNAL", _safe_str(sig_name), (0, 0), dxfattribs={"flags": 1})
                break

        if registry is not None:
            for e in doc.modelspace():
                if e.dxftype() != "INSERT":
                    continue
                atts = {a.dxf.tag: a.dxf.text for a in e.attribs}
                r_name = (atts.get("R_NAME") or "").strip()
                s_name = (atts.get("S_NAME") or "").strip()
                if not (r_name and s_name):
                    continue
                if e.dxf.name.startswith("Relay_Coil"):
                    registry.register(s_name, r_name, "SIGNAL", sht)
                elif e.dxf.name.startswith("Front_Contact") or e.dxf.name.startswith("Back_Contact"):
                    letter = (atts.get("C") or "").strip()
                    f_val = (atts.get("F") or "").strip()
                    a_val = (atts.get("A") or "").strip()
                    b_val = (atts.get("B") or "").strip()
                    # CONFIRMED: the full (R_NAME, letter, number-pair) is
                    # the real contact identity, not letter alone - the
                    # same R_NAME+letter combo legitimately repeats
                    # across different terminal pairs (F/A on front
                    # contacts, A/B on back contacts).
                    if letter:
                        registry.register(s_name, f"{r_name}{letter}{f_val}{a_val}{b_val}", "SIGNAL", sht)

        out_name = f"{sig_name}_{template_name.replace('.dxf', '')}_SHT{sht}.dxf"
        results.append((out_name, doc, sht))
        sheet_num += 1

    return results, sheet_num


def get_signal_start_sheet_number(xlsx_path: str) -> int:
    """
    Read the Signal circuit starting sheet number from the FIELD PG.NO
    sheet's own 'SIGNAL' row (searched by label, not a fixed cell -
    CONFIRMED FIX: a hardcoded 'C6' reference broke as soon as new rows
    like Cover Sheet/Station Layout/Floor Plan shifted the whole sheet's
    layout, since row positions aren't stable).
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(row=r, column=2).value or "").strip().upper()
        if label == "SIGNAL":
            value = ws.cell(row=r, column=3).value
            if value is None or str(value).strip() == "":
                raise ValueError("FIELD PG.NO has no Sheet Number for the SIGNAL row")
            return int(str(value).strip())
    raise ValueError("Could not find a 'SIGNAL' row in FIELD PG.NO")


def get_next_circuit_start_sheet_number(xlsx_path: str) -> str:
    """
    Read the NEXT circuit's starting sheet number from the FIELD PG.NO
    sheet - whatever row comes right after 'SIGNAL' (typically TRACK),
    searched by label rather than a fixed cell (CONFIRMED FIX: a
    hardcoded 'C7' reference broke as soon as new rows shifted the
    sheet's layout - see get_signal_start_sheet_number above). This is
    what the very last Signal sheet's CONT value should point to, since
    physically the next sheet in the document set is the first sheet of
    the next circuit type, not a simple +1 continuation of Signal's own
    numbering.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    found_signal = False
    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(row=r, column=2).value or "").strip().upper()
        if found_signal:
            value = ws.cell(row=r, column=3).value
            if value is not None and str(value).strip() != "":
                v = str(value).strip()
                return v.zfill(3) if v.isdigit() else v
            continue  # skip blank rows, keep looking for the next real one
        if label == "SIGNAL":
            found_signal = True
    raise ValueError("Could not find a row after 'SIGNAL' in FIELD PG.NO")


def set_cont_value(doc, new_cont_value: str) -> None:
    """Override the CONT attribute on a generated sheet's TITLEBLOCK insert."""
    msp = doc.modelspace()
    for e in msp:
        if e.dxftype() == "INSERT" and e.dxf.name == "TITLEBLOCK":
            for att in e.attribs:
                if att.dxf.tag == "CONT":
                    att.dxf.text = str(new_cont_value)
