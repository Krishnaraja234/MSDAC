"""
FLOOR PLAN sheet generation - same convention as STATION LAYOUT: blank
placeholder sheets (border/title-block only, no actual content), with
the sheet count read from FIELD PG.NO's 'SPARE SHEETS' column for the
'FLOOR PLAN' row.
"""

import ezdxf

from signal_core import insert_border_title

TITLE_TEXT_PREFIX = "FLOOR PLAN - "


def get_floor_plan_config(xlsx_path: str):
    """Reads FLOOR PLAN's starting sheet number and sheet count (from
    SPARE SHEETS) from FIELD PG.NO."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(row=r, column=2).value or "").strip().upper()
        if label == "FLOOR PLAN":
            sht_value = ws.cell(row=r, column=3).value
            count_value = ws.cell(row=r, column=4).value
            if sht_value is None or str(sht_value).strip() == "":
                raise ValueError("FIELD PG.NO has no Sheet Number for the FLOOR PLAN row")
            if count_value is None or str(count_value).strip() == "":
                raise ValueError("FIELD PG.NO has no SPARE SHEETS count for the FLOOR PLAN row")
            return str(sht_value).strip(), int(count_value)
    raise ValueError("Could not find a 'FLOOR PLAN' row in FIELD PG.NO")


def _increment_alpha_suffix(sht: str) -> str:
    """Same convention as station_layout_core.py/index_core.py's own
    version - kept local to avoid a cross-module dependency for one
    small shared helper."""
    m = 0
    while m < len(sht) and sht[len(sht) - 1 - m].isalpha():
        m += 1
    if m == 0:
        return str(int(sht) + 1)
    prefix, suffix = sht[:-m], sht[-m:]
    chars = list(suffix)
    i = len(chars) - 1
    while i >= 0:
        if chars[i] != "Z":
            chars[i] = chr(ord(chars[i]) + 1)
            break
        chars[i] = "A"
        i -= 1
    else:
        chars.insert(0, "A")
    return prefix + "".join(chars)


def generate_all_floor_plan_sheets(xlsx_path: str, border_template_path: str):
    """
    Returns (list of (out_name, doc, sht_number_str), next_sheet_number),
    same shape as every other generate_all_* function - one blank sheet
    per SPARE SHEETS count, titled "FLOOR PLAN - N".
    """
    start_sht, count = get_floor_plan_config(xlsx_path)

    results = []
    sht = start_sht
    for page_number in range(1, count + 1):
        next_sht = _increment_alpha_suffix(sht)
        doc = ezdxf.new()
        insert_border_title(
            doc=doc,
            border_template_path=border_template_path,
            template_name="FLOOR_PLAN",
            sig_name="", ahead_sig="", loc="",
            sht_number=str(sht),
            cont_number=str(next_sht),
            title_text_override=f"{TITLE_TEXT_PREFIX}{page_number}",
        )
        out_name = f"FLOOR_PLAN_SHT{sht}.dxf"
        results.append((out_name, doc, sht))
        sht = next_sht

    return results, sht
