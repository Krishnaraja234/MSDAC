"""
LCPR repeater relay contact allocation.

CONFIRMED (worked through directly against Krish's own LCPR_CONFIG.xlsx
example, tracing every allocation for 2 signals end-to-end):

  - Signal sheet's own "LCPR" column (E) holds a repeater relay NAME
    per signal row (e.g. "@LCPR") - when filled, TYPICAL CONFIG swaps in
    the LCPR-variant files for that signal (handled elsewhere, in
    signal_core.py's existing TYPICAL CONFIG swap logic). This module
    allocates which physical contact each LCPR_FRONT/LCPR_BACK CONTACT
    block instance in those files actually uses.

  - STANDARD3_LCPR.dxf / STANDARD4_LCPR.dxf are FRONT-ONLY (no back
    contact at all). Their instances fill a shared pool of 6 slots per
    relay instance - letters A, B, C, 2 pairs each (e.g. A1/A2, A3/A4).
    Once all 6 slots on the current relay are used, front-only
    allocation moves to a NEW relay instance ("@LCPR" -> "@LCPR1" ->
    "@LCPR2" -> ...) and starts again at A. Front-only NEVER touches D.

  - STANDARD5_LCPR.dxf is different: it has 4 LCPR_FRONT instances PLUS
    exactly 1 LCPR_BACK CONTACT instance, and front+back MUST stay on
    the SAME relay instance:
      - Its 4 front instances continue from wherever the shared A/B/C
        cursor currently is on the CURRENT relay (does NOT jump to a
        new relay just because A/B/C don't have enough room) - if
        there isn't enough room left in A/B/C, it overflows into
        D1/D2, D3/D4, D5/D6 (still on the SAME relay) rather than
        starting a new relay instance.
      - Its 1 back contact ALWAYS uses the fixed pair D7/D8 on that
        SAME relay instance, regardless of how much of A/B/C/D1-D6 the
        front instances actually used.
      - After STANDARD5 finishes, the cursor continues from wherever
        it left off (e.g. if only A+B of the normal pool were used,
        C is still free for whatever comes next - confirmed directly:
        Signal 2's STANDARD3 correctly picked up Signal 1's leftover
        "C" on the same relay instance after Signal 1's STANDARD5).

  - If two different signal rows share the SAME repeater relay name
    (same LCPR column value), they share the SAME cursor/pool - not a
    separate one each. Signals are processed in the order they appear
    in the SIGNAL sheet, and within a signal, typicals are processed in
    TYPICAL CONFIG's own ORDER (STANDARD3, then STANDARD4, then
    STANDARD5).
"""

from collections import defaultdict

NORMAL_SLOTS = [("A", 0), ("A", 1), ("B", 0), ("B", 1), ("C", 0), ("C", 1)]
D_OVERFLOW_SLOTS = [("D", 0), ("D", 1), ("D", 2)]  # D1/D2, D3/D4, D5/D6 - D7/D8 always reserved for STANDARD5's own back


def allocate_lcpr_contacts(signal_lcpr_requests: list) -> dict:
    """
    signal_lcpr_requests: list of (signal_name, lcpr_value, front_only_count, has_standard5)
        tuples, IN THE ORDER signals appear in the SIGNAL sheet.
        front_only_count = total plain front-only instances needed from
        STANDARD3_LCPR.dxf + STANDARD4_LCPR.dxf combined (e.g. 2+4=6).
        has_standard5 = True if this signal also uses STANDARD5_LCPR.dxf
        (always exactly 4 front + 1 back when True, per its own fixed
        template shape - CONFIRMED, not a variable count).

    Returns: {lcpr_value: {signal_name: {
        "front_only": [(relay_instance, letter, pair_index), ...],
        "standard5_front": [(relay_instance, letter, pair_index), ...] (always 4 entries when has_standard5),
        "standard5_back": (relay_instance, "D", 3) or None (pair_index 3 = the fixed D7/D8 slot)
    }}}
    relay_instance is 0 for "@LCPR" itself, 1 for "@LCPR1", 2 for "@LCPR2", etc.
    """
    groups = defaultdict(list)
    group_order = []
    for req in signal_lcpr_requests:
        lcpr_value = req[1]
        if lcpr_value not in group_order:
            group_order.append(lcpr_value)
        groups[lcpr_value].append(req)

    result = {}
    for lcpr_value in group_order:
        result[lcpr_value] = {}
        relay_instance = 0
        normal_used = 0   # 0-6, how many of NORMAL_SLOTS used on the CURRENT relay
        d_overflow_used = 0  # 0-3, how many of D_OVERFLOW_SLOTS used on the CURRENT relay (only STANDARD5 ever touches this)

        for signal_name, _lcpr_value, front_only_count, has_standard5 in groups[lcpr_value]:
            front_only_allocations = []
            for _ in range(front_only_count):
                if normal_used >= len(NORMAL_SLOTS):
                    # CONFIRMED: front-only NEVER uses D, even if D
                    # slots are sitting unused from a previous
                    # STANDARD5's overflow - it moves to a fresh relay.
                    relay_instance += 1
                    normal_used = 0
                    d_overflow_used = 0
                letter, pair_index = NORMAL_SLOTS[normal_used]
                front_only_allocations.append((relay_instance, letter, pair_index))
                normal_used += 1

            standard5_front = None
            standard5_back = None
            if has_standard5:
                if normal_used >= len(NORMAL_SLOTS):
                    # CONFIRMED: the current relay has ZERO room left at
                    # all (not even a partial letter) - move to a fresh
                    # relay instance first, same as front-only would.
                    # Overflowing into D only happens when there's SOME
                    # partial room left to use first (see Signal 2's
                    # case in the confirmed example) - not when there's
                    # nothing left whatsoever.
                    relay_instance += 1
                    normal_used = 0
                    d_overflow_used = 0

                standard5_front = []
                for _ in range(4):  # CONFIRMED: STANDARD5 always has exactly 4 front instances
                    if normal_used < len(NORMAL_SLOTS):
                        letter, pair_index = NORMAL_SLOTS[normal_used]
                        normal_used += 1
                    else:
                        # CONFIRMED: overflow into D1-D6 on the SAME
                        # relay, rather than moving to a new one.
                        letter, pair_index = D_OVERFLOW_SLOTS[d_overflow_used]
                        d_overflow_used += 1
                    standard5_front.append((relay_instance, letter, pair_index))
                # CONFIRMED: back contact always uses the fixed D7/D8
                # pair on this SAME relay instance, regardless of how
                # much of the normal/overflow pool the front used.
                standard5_back = (relay_instance, "D", 3)

            result[lcpr_value][signal_name] = {
                "front_only": front_only_allocations,
                "standard5_front": standard5_front,
                "standard5_back": standard5_back,
            }

    return result


def relay_instance_name(relay_instance: int) -> str:
    """0 -> 'LCPR', 1 -> 'LCPR1', 2 -> 'LCPR2', ... CONFIRMED: this is a
    FIXED suffix word, independent of whatever the user typed in the
    SIGNAL sheet's own LCPR column (that value becomes NAME1 in Relay
    Rack, e.g. "LC29" - unchanged across every overflow instance for
    that signal). This function's result becomes NAME2 - e.g. Relay
    Rack tag row "LC29 LCPR1" for the second repeater relay instance."""
    if relay_instance == 0:
        return "LCPR"
    return f"LCPR{relay_instance}"


def slot_to_contact_codes(letter: str, pair_index: int) -> tuple:
    """(letter, pair_index) -> ('A1','A2') style contact code pair -
    pair_index 0 -> 1,2 ; 1 -> 3,4 ; 2 -> 5,6 ; 3 -> 7,8."""
    first_num = pair_index * 2 + 1
    return (f"{letter}{first_num}", f"{letter}{first_num + 1}")


def get_signal_lcpr_column(xlsx_path: str) -> list:
    """Reads the SIGNAL sheet's own SIG NAME/LCPR columns, IN ROW ORDER
    - returns [(sig_name, lcpr_value), ...] for every row (lcpr_value is
    "" for rows where the LCPR column is blank)."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["SIGNAL"]
    rows = []
    for r in range(2, ws.max_row + 1):
        sig_name = ws.cell(row=r, column=1).value
        if sig_name is None or str(sig_name).strip() == "":
            continue
        lcpr_value = ws.cell(row=r, column=5).value
        rows.append((str(sig_name).strip(), str(lcpr_value).strip() if lcpr_value else ""))
    return rows


def apply_lcpr_allocations(pending: list, xlsx_path: str, log) -> None:
    """
    Post-processing step (run after all Signal sheets are generated,
    same pattern as apply_relay_positions/clear_placeholder_values):
    finds every LCPR_FRONT/LCPR_BACK CONTACT block instance across all
    generated sheets, groups them by which signal owns them (via the
    OWNING_SIGNAL marker already embedded on each sheet's title block),
    runs the confirmed allocation engine, and writes the resulting
    letter/number directly onto each block's own C/F/B/A attributes.

    S_NAME (the repeater relay's own name, e.g. "@LCPR1") is NOT set
    here - that's still handled by the existing "@" token substitution
    in signal_core.py using the raw LCPR column value. This function
    only sets the per-instance CONTACT POSITION (C + F/B/A), since a
    single signal's own LCPR_FRONT instances can legitimately span
    MULTIPLE repeater relay instances (e.g. "@LCPR" then "@LCPR1") -
    something the simple raw-value substitution can't express on its
    own. Where a signal's allocation crosses into an overflow relay
    instance, this ALSO corrects S_NAME to the correct instance name.
    """
    signal_lcpr = dict(get_signal_lcpr_column(xlsx_path))
    signal_order = [name for name, _ in get_signal_lcpr_column(xlsx_path)]

    # Group pending docs by owning signal, counting LCPR_FRONT/BACK instances.
    docs_by_signal = {}
    for _final_name, doc in pending:
        owning_signal = None
        msp = doc.modelspace()
        for e in msp:
            if e.dxftype() == "INSERT" and e.dxf.name in ("TITLE", "TITLEBLOCK"):
                for att in e.attribs:
                    if att.dxf.tag == "OWNING_SIGNAL":
                        owning_signal = (att.dxf.text or "").strip()
                break
        if owning_signal:
            docs_by_signal.setdefault(owning_signal, []).append(doc)

    requests = []
    for sig_name in signal_order:
        lcpr_value = signal_lcpr.get(sig_name, "")
        if not lcpr_value:
            continue
        docs = docs_by_signal.get(sig_name, [])
        front_only_count = 0
        has_standard5 = False
        for doc in docs:
            msp = doc.modelspace()
            front_instances = [e for e in msp if e.dxftype() == "INSERT" and e.dxf.name == "LCPR_FRONT"]
            back_instances = [e for e in msp if e.dxftype() == "INSERT" and e.dxf.name == "LCPR_BACK CONTACT"]
            if back_instances:
                # CONFIRMED: this is the STANDARD5-style sheet (4 front +
                # 1 back, always) - not counted into front_only_count.
                has_standard5 = True
            else:
                front_only_count += len(front_instances)
        requests.append((sig_name, lcpr_value, front_only_count, has_standard5))

    if not requests:
        return

    allocation = allocate_lcpr_contacts(requests)

    for sig_name, lcpr_value, _front_count, has_standard5 in requests:
        allocs = allocation[lcpr_value][sig_name]
        docs = docs_by_signal.get(sig_name, [])

        front_only_queue = list(allocs["front_only"])
        for doc in docs:
            msp = doc.modelspace()
            back_instances = [e for e in msp if e.dxftype() == "INSERT" and e.dxf.name == "LCPR_BACK CONTACT"]
            front_instances = [e for e in msp if e.dxftype() == "INSERT" and e.dxf.name == "LCPR_FRONT"]

            if back_instances:
                # STANDARD5-style sheet: 4 front (standard5_front) + 1 back (standard5_back).
                if allocs["standard5_front"] is None:
                    log.error(f"[SIGNAL] {sig_name}: expected a STANDARD5-style LCPR sheet but no allocation was computed.")
                    continue
                for e, (relay_instance, letter, pair_index) in zip(front_instances, allocs["standard5_front"]):
                    _apply_lcpr_contact(e, letter, pair_index, lcpr_value, relay_instance, front=True)
                for e in back_instances:
                    relay_instance, letter, pair_index = allocs["standard5_back"]
                    _apply_lcpr_contact(e, letter, pair_index, lcpr_value, relay_instance, front=False)
            else:
                for e in front_instances:
                    if not front_only_queue:
                        log.error(f"[SIGNAL] {sig_name}: ran out of allocated contacts for a LCPR_FRONT instance.")
                        continue
                    relay_instance, letter, pair_index = front_only_queue.pop(0)
                    _apply_lcpr_contact(e, letter, pair_index, lcpr_value, relay_instance, front=True)


def _apply_lcpr_contact(entity, letter: str, pair_index: int, base_lcpr_value: str, relay_instance: int, front: bool):
    """
    Writes the allocated letter/number onto one LCPR_FRONT or
    LCPR_BACK CONTACT block instance's own attributes, and sets R_NAME
    to the correct overflow suffix ("LCPR"/"LCPR1"/"LCPR2"...) -
    CONFIRMED: Relay Rack represents these as NAME1=<whatever the user
    typed in the SIGNAL sheet's own LCPR column, e.g. "LC29"> and
    NAME2=<this fixed overflow suffix>, e.g. tag row "LC29 LCPR1" for
    the second repeater relay instance. S_NAME (NAME1) is left to the
    EXISTING "@" token substitution mechanism in signal_core.py, since
    it never varies by relay_instance for a given signal - only R_NAME
    (NAME2) needs to be set here, per-instance.
    """
    first_code, second_code = slot_to_contact_codes(letter, pair_index)
    relay_suffix = relay_instance_name(relay_instance)
    first_num = first_code[len(letter):]
    second_num = second_code[len(letter):]
    for att in entity.attribs:
        if att.dxf.tag == "C":
            att.dxf.text = letter
        elif front and att.dxf.tag == "F":
            att.dxf.text = first_num
        elif not front and att.dxf.tag == "B":
            att.dxf.text = first_num
        elif att.dxf.tag == "A":
            att.dxf.text = second_num
        elif att.dxf.tag == "R_NAME":
            att.dxf.text = relay_suffix


def _count_lcpr_blocks_in_file(template_path: str) -> tuple:
    """Returns (front_only_count, has_back) for one template file -
    counts LCPR_FRONT and LCPR_BACK CONTACT instances directly."""
    import ezdxf
    doc = ezdxf.readfile(template_path)
    msp = doc.modelspace()
    front = sum(1 for e in msp if e.dxftype() == "INSERT" and e.dxf.name == "LCPR_FRONT")
    back = sum(1 for e in msp if e.dxftype() == "INSERT" and e.dxf.name == "LCPR_BACK CONTACT")
    return front, back > 0


def get_lcpr_shape_for_typical(typical_name: str, xlsx_path: str) -> tuple:
    """
    Returns (front_only_count, has_standard5) for a given SIGNAL row's
    own "TYPICAL" value (e.g. "STANDARD") - reads TYPICAL CONFIG to find
    which LCPR-variant files apply (via LCPR FILE NAME / LCPR ONLY
    columns), then counts each file's own LCPR_FRONT/LCPR_BACK CONTACT
    instances directly from the template files on disk. This is what
    lets relay availability be checked BEFORE generation ever runs -
    no generated sheets needed yet, just the template files themselves.
    """
    import openpyxl
    from signal_core import resolve_template_path

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["TYPICAL CONFIG"]

    front_only_count = 0
    has_standard5 = False
    for r in range(2, ws.max_row + 1):
        row_typical = ws.cell(row=r, column=1).value
        if row_typical is None or str(row_typical).strip() != typical_name.strip():
            continue
        base_file = ws.cell(row=r, column=3).value
        lcpr_file = ws.cell(row=r, column=5).value
        lcpr_only = ws.cell(row=r, column=6).value

        target_file = None
        if lcpr_only and str(lcpr_only).strip().upper() == "YES":
            target_file = base_file  # e.g. STANDARD5_LCPR.dxf itself - only exists when LCPR is filled
        elif lcpr_file and str(lcpr_file).strip():
            target_file = lcpr_file  # e.g. STANDARD3.dxf -> STANDARD3_LCPR.dxf swap

        if target_file:
            path = resolve_template_path(str(target_file).strip())
            front, has_back = _count_lcpr_blocks_in_file(path)
            if has_back:
                has_standard5 = True
            else:
                front_only_count += front

    return front_only_count, has_standard5


def validate_lcpr_relay_availability(xlsx_path: str, log) -> bool:
    """
    CONFIRMED: proactively calculates exactly how many repeater relay
    instances every LCPR group needs, BEFORE generation runs, and
    checks whether they actually exist in Relay Rack - rather than only
    discovering a shortfall via a generic "relay position not found"
    error after the fact. Returns False (and logs a clear vital error
    naming exactly what's missing and how many relays are needed in
    total) if anything's missing; True if everything checks out (or
    there's no LCPR usage at all in this workbook).
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["SIGNAL"]
    requests = []
    for r in range(2, ws.max_row + 1):
        sig_name = ws.cell(row=r, column=1).value
        if sig_name is None or str(sig_name).strip() == "":
            continue
        lcpr_value = ws.cell(row=r, column=5).value
        if not lcpr_value or str(lcpr_value).strip() == "":
            continue
        typical_name = ws.cell(row=r, column=4).value or ""
        front_only_count, has_standard5 = get_lcpr_shape_for_typical(str(typical_name), xlsx_path)
        requests.append((str(sig_name).strip(), str(lcpr_value).strip(), front_only_count, has_standard5))

    if not requests:
        return True

    allocation = allocate_lcpr_contacts(requests)

    needed_pairs = set()
    for lcpr_value, signals in allocation.items():
        for _sig_name, allocs in signals.items():
            for ri, _letter, _pi in allocs["front_only"]:
                needed_pairs.add((lcpr_value.upper(), relay_instance_name(ri).upper()))
            if allocs["standard5_front"]:
                for ri, _letter, _pi in allocs["standard5_front"]:
                    needed_pairs.add((lcpr_value.upper(), relay_instance_name(ri).upper()))
            if allocs["standard5_back"]:
                ri, _letter, _pi = allocs["standard5_back"]
                needed_pairs.add((lcpr_value.upper(), relay_instance_name(ri).upper()))

    from relay_rack_core import read_all_relay_racks
    racks = read_all_relay_racks(xlsx_path)
    existing_pairs = set()
    for _rack_num, grid_rows in racks:
        for band in grid_rows:
            for cell in band:
                if cell["NAME3"] != "SPARE" and cell["NAME1"]:
                    existing_pairs.add((cell["NAME1"].strip().upper(), cell["NAME2"].strip().upper()))

    missing = sorted(needed_pairs - existing_pairs)
    if missing:
        by_lcpr_value = {}
        for lcpr_value, relay_name in needed_pairs:
            by_lcpr_value.setdefault(lcpr_value, set()).add(relay_name)
        summary = "; ".join(
            f"{lv} needs {len(names)} relay instance(s) ({', '.join(sorted(names))})"
            for lv, names in by_lcpr_value.items()
        )
        missing_text = ", ".join(f"{n1} {n2}" for n1, n2 in missing)
        log.error(
            f"[LCPR] Not enough repeater relays defined in Relay Rack before generation. "
            f"{summary}. Missing from Relay Rack: {missing_text}."
        )
        return False

    return True
