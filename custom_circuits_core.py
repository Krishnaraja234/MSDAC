"""
MSDAC Custom Circuits Generator - core assembly logic.

CONFIRMED workflow: user uploads a ZIP of their own pre-made circuit
files (DXF and/or DWG - DWGs get converted to DXF via ODA first). Each
file is matched by name against the CUSTOM CIRCUITS sheet's B column
(CAD FILE NAME), and gets the border/title template applied - TITLE
comes from that row's C column, SHT/CONT assigned sequentially starting
from FIELD PG.NO's CUSTOM CIRCUITS row.

Reads start from row 2 (row 1 is the header: SI.No | CAD FILE NAME | TITLE).
Rows with a blank CAD FILE NAME are skipped.
"""

import os
import shutil
import zipfile
import tempfile

import ezdxf

from signal_core import insert_border_title as _insert_border_title_generic
from oda_convert import convert_folder_to_dxf


def read_custom_circuits_rows(xlsx_path: str):
    """Reads the CUSTOM CIRCUITS sheet: list of dicts {CAD_FILE_NAME, TITLE}, skipping blank rows."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["CUSTOM CIRCUITS"]
    rows = []
    for r in range(2, ws.max_row + 1):
        cad_file_name = ws.cell(row=r, column=2).value
        title = ws.cell(row=r, column=3).value
        if cad_file_name is None or str(cad_file_name).strip() == "":
            continue
        rows.append({
            "CAD_FILE_NAME": str(cad_file_name).strip(),
            "TITLE": str(title).strip() if title else "",
        })
    return rows


def _extract_and_normalize_zip(zip_path: str, work_dir: str) -> str:
    """
    Extracts the uploaded zip, converts any .dwg files to .dxf (via ODA),
    and returns the folder path containing only .dxf files ready to match.
    """
    extract_dir = os.path.join(work_dir, "extracted")
    os.makedirs(extract_dir, exist_ok=True)
    with zipfile.ZipFile(zip_path, "r") as zf:
        zf.extractall(extract_dir)

    # Flatten - some zips have files nested in a subfolder.
    flat_dir = os.path.join(work_dir, "flat")
    os.makedirs(flat_dir, exist_ok=True)
    for root, _dirs, files in os.walk(extract_dir):
        for fname in files:
            if fname.lower().endswith((".dxf", ".dwg")):
                shutil.copy(os.path.join(root, fname), os.path.join(flat_dir, fname))

    dwg_files = [f for f in os.listdir(flat_dir) if f.lower().endswith(".dwg")]
    if dwg_files:
        dxf_out_dir = os.path.join(work_dir, "converted_dxf")
        os.makedirs(dxf_out_dir, exist_ok=True)
        convert_folder_to_dxf(flat_dir, dxf_out_dir)
        # Copy converted DXFs alongside any DXFs that were already present.
        for fname in os.listdir(dxf_out_dir):
            shutil.copy(os.path.join(dxf_out_dir, fname), os.path.join(flat_dir, fname))

    return flat_dir


def _find_matching_file(folder: str, cad_file_name: str):
    """Matches a CAD_FILE_NAME (with or without extension) against files in folder, case-insensitively."""
    target_stem = os.path.splitext(cad_file_name)[0].strip().lower()
    for fname in os.listdir(folder):
        if not fname.lower().endswith(".dxf"):
            continue
        stem = os.path.splitext(fname)[0].strip().lower()
        if stem == target_stem:
            return os.path.join(folder, fname)
    return None


def generate_all_custom_circuit_sheets(
    xlsx_path: str, zip_path: str, start_sheet_number: int, border_template_path: str
):
    """
    Generates all Custom Circuit sheets by matching each CUSTOM CIRCUITS
    row's CAD_FILE_NAME against the uploaded zip's files, applying the
    border/title template.
    Returns (list of (output_filename, ezdxf.Drawing, sht_number_str), next_sheet_number)
    """
    rows = read_custom_circuits_rows(xlsx_path)
    if not rows:
        raise ValueError("No rows found in CUSTOM CIRCUITS sheet (column B, CAD FILE NAME, is empty for all rows)")

    work_dir = tempfile.mkdtemp(prefix="custom_circuits_")
    try:
        folder = _extract_and_normalize_zip(zip_path, work_dir)

        results = []
        sheet_num = start_sheet_number
        for row in rows:
            cad_file_name = row["CAD_FILE_NAME"]
            matched_path = _find_matching_file(folder, cad_file_name)
            if matched_path is None:
                raise ValueError(
                    f"CUSTOM CIRCUITS row references CAD FILE NAME {cad_file_name!r}, "
                    f"but no matching .dxf/.dwg file was found in the uploaded zip"
                )

            doc = ezdxf.readfile(matched_path)
            sht = f"{sheet_num:03d}"
            cont = f"{sheet_num + 1:03d}"

            _insert_border_title_generic(
                doc=doc,
                border_template_path=border_template_path,
                template_name="CUSTOM",
                sig_name=None,
                ahead_sig=None,
                loc=None,
                sht_number=sht,
                cont_number=cont,
                title_text_override=row["TITLE"],
            )

            out_name = f"CUSTOM_{cad_file_name}_SHT{sht}.dxf"
            results.append((out_name, doc, sht))
            sheet_num += 1

        return results, sheet_num
    finally:
        shutil.rmtree(work_dir, ignore_errors=True)


def get_custom_circuits_start_sheet_number(xlsx_path: str) -> int:
    """Reads CUSTOM CIRCUITS's starting sheet number from FIELD PG.NO."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(row=r, column=2).value or "").strip().upper()
        if label == "CUSTOM CIRCUITS":
            value = ws.cell(row=r, column=3).value
            if value is None or str(value).strip() == "":
                raise ValueError("FIELD PG.NO has no Sheet Number for the CUSTOM CIRCUITS row")
            return int(str(value).strip())
    raise ValueError("Could not find a 'CUSTOM CIRCUITS' row in FIELD PG.NO")


def get_next_circuit_after_custom_circuits(xlsx_path: str) -> str:
    """Reads the starting sheet number of whatever circuit follows CUSTOM CIRCUITS in FIELD PG.NO."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(row=r, column=2).value or "").strip().upper()
        if label == "CUSTOM CIRCUITS":
            value = ws.cell(row=r + 1, column=3).value
            return str(value).strip() if value is not None else ""
    return ""
