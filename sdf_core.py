"""
MSDAC SDF Circuit Generator - core assembly logic.

Derives entirely from the TRACK sheet (no dedicated SDF Excel sheet) -
CONFIRMED: "no need any Excel input, it will be based on Track input".

Structure (CONFIRMED so far, further refinement expected):
  - SDF_CIRCUIT.dxf is the base/first-track unit: FUZEBLOCK, FUZEENDBLOCK,
    first track's Back_Contact pair (TS(M)/TS(R)) + Front_Contact (TSPR),
    plus the shared JR (10-60 SEC timer relay) and SDF Front_Contact/coil.
  - SDF_CIRCUIT_CONTACTS.dxf is the repeatable per-track contact unit
    (Back_Contact TS(M)/TS(R) + Front_Contact TSPR only) for tracks 2+.
  - S_NAME on each track's contacts = that TRACK NAME (CONFIRMED,
    matching the same '*' substitution convention as Signal circuits).
  - Stacking (CONFIRMED): pick=(208.381, 217.809) [= SDF_CIRCUIT.dxf's own
    2nd Back_Contact/TS(R) native position], place point for track i
    (0-indexed, i=0 is the FIRST additional track after the base) =
    (208.3812, 217.8089 - i*32.1206).
  - CONFIRMED: only same-DIRECTION tracks (consecutive) stack onto one
    sheet; a direction change starts a new sheet.

NOTE: exact handling of JR/SDF coil when there are MULTIPLE tracks per
sheet is not yet fully confirmed - current implementation shares ONE
JR+SDF coil pair per sheet (from the base SDF_CIRCUIT.dxf unit), applied
once regardless of how many tracks stack onto that sheet. Flagged for
review/refinement per explicit note from the user ("let me change that later").
"""

import ezdxf
from ezdxf.addons import importer
import os

from signal_core import (
    TEMPLATES_DIR,
    resolve_template_path,
    _safe_str,
    insert_border_title as _insert_border_title_generic,
)

import re

CONTACT_PICK_POINT = (259.8977, 232.5538)
Y_SPACING = 32.5049


def _parse_contact_groups(value: str):
    """'A5,A6/A7,A8' -> [['A5','A6'], ['A7','A8']]"""
    groups = []
    for chunk in str(value).split("/"):
        chunk = chunk.strip()
        if not chunk:
            continue
        values = [v.strip() for v in chunk.split(",")]
        groups.append(values)
    return groups


def _afc_for_subvalue(subvalue: str):
    """'A5' -> (letter='A', number='5')."""
    m = re.match(r"^([A-Za-z]+)(\d+)$", subvalue)
    if not m:
        raise ValueError(f"Could not parse contact sub-value {subvalue!r} (expected e.g. 'A5')")
    return m.group(1), m.group(2)


def get_sdf_contact_groups(xlsx_path: str):
    """
    Reads FIELD PG.NO!L4 (for TS(M)/TS(R), CYCLED per track - same group
    used for BOTH contacts within a track) and M4 (for TSPR - uses the
    group's 1st value; the SDF relay's own shared Front_Contact uses the
    group's 2nd, "free" value).
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    l4 = ws["L4"].value
    m4 = ws["M4"].value
    if not l4 or not m4:
        raise ValueError("FIELD PG.NO!L4 (TS(M)/TS(R)) and M4 (TSPR) must both be filled in")
    return _parse_contact_groups(l4), _parse_contact_groups(m4)


def read_track_rows(xlsx_path: str):
    """Reads the TRACK sheet: list of dicts {TRACK NAME, DIRECTION, TYPICAL}."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["TRACK"]
    rows = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        direction = ws.cell(row=r, column=2).value
        typical = ws.cell(row=r, column=3).value
        if name is None or str(name).strip() == "":
            continue
        rows.append({
            "TRACK NAME": str(name).strip(),
            "DIRECTION": str(direction).strip() if direction else "",
            "TYPICAL": str(typical).strip() if typical else "",
        })
    return rows


def _group_by_direction(rows):
    """Groups CONSECUTIVE rows sharing the same DIRECTION into separate sheet-groups."""
    groups = []
    current = []
    current_dir = None
    for row in rows:
        if current and row["DIRECTION"] != current_dir:
            groups.append(current)
            current = []
        current.append(row)
        current_dir = row["DIRECTION"]
    if current:
        groups.append(current)
    return groups


def _import_whole_file_no_rename(doc, source_filename: str, pick_point, place_point):
    """
    Like _import_whole_file, but avoids block-name-renaming when the same
    block definition gets re-imported multiple times (e.g. tracks 2-6 all
    reusing SDF_CIRCUIT_CONTACTS.dxf). Imports each INSERT's block
    DEFINITION only once per document, then creates a fresh blockref
    preserving the original name at the translated position. Raw geometry
    (the connecting LWPOLYLINEs) is copied and translated fresh each call,
    since it has no renaming risk.
    """
    source_doc = ezdxf.readfile(resolve_template_path(source_filename))
    source_msp = source_doc.modelspace()
    target_msp = doc.modelspace()

    dx = place_point[0] - pick_point[0]
    dy = place_point[1] - pick_point[1]

    new_entities = []
    for e in source_msp:
        if e.dxftype() == "INSERT":
            block_name = e.dxf.name
            if block_name not in doc.blocks:
                imp = importer.Importer(source_doc, doc)
                imp.import_block(block_name)
                imp.finalize()
            new_x = e.dxf.insert.x + dx
            new_y = e.dxf.insert.y + dy
            ins = target_msp.add_blockref(block_name, (new_x, new_y))
            ins.dxf.xscale = e.dxf.xscale
            ins.dxf.yscale = e.dxf.yscale
            ins.dxf.zscale = e.dxf.zscale
            ins.dxf.rotation = e.dxf.rotation
            for att in e.attribs:
                dxfattribs = {
                    "height": att.dxf.height,
                    "style": att.dxf.style,
                    "rotation": att.dxf.rotation,
                    "flags": att.dxf.flags,
                }
                att_insert = (att.dxf.insert.x + dx, att.dxf.insert.y + dy, att.dxf.insert.z)
                ins.add_attrib(att.dxf.tag, att.dxf.text, att_insert, dxfattribs)
            new_entities.append(ins)
        else:
            existing_ids = set(id(x) for x in target_msp)
            imp = importer.Importer(source_doc, doc)
            imp.import_entity(e, target_msp)
            imp.finalize()
            copied = [x for x in target_msp if id(x) not in existing_ids]
            for c in copied:
                try:
                    c.translate(dx, dy, 0)
                except AttributeError:
                    pass
                new_entities.append(c)

    return new_entities


def _import_whole_file(doc, source_filename: str, pick_point=None, place_point=None):
    """Copy the ENTIRE source file's modelspace content into doc, optionally translated."""
    source_doc = ezdxf.readfile(resolve_template_path(source_filename))
    target_msp = doc.modelspace()
    existing_ids = set(id(e) for e in target_msp)

    imp = importer.Importer(source_doc, doc)
    imp.import_modelspace()
    imp.finalize()

    new_entities = [e for e in target_msp if id(e) not in existing_ids]
    if pick_point is not None and place_point is not None:
        dx = place_point[0] - pick_point[0]
        dy = place_point[1] - pick_point[1]
        for e in new_entities:
            try:
                e.translate(dx, dy, 0)
            except AttributeError:
                pass
    return new_entities


def _read_sdf_bit_names(xlsx_path: str) -> dict:
    """
    CONFIRMED: FIELD PG.NO!F2/F3 ("SDF BIT" column) gives the exact
    relay to use for the SDF_CONTACT/SDF_RELAY block per direction (e.g.
    F2="LOC DNSDF" for DN, F3="UP SDF" for UP) - NOT constant, since a
    station name prefix may or may not be present. FIELD PG.NO!E2/E3
    ("SDF TIMER" column) gives the same for the TIMER_RELAY block (e.g.
    "DN JR"/"UP JR"). Both get parsed via the same whitelist-based
    split_bit_name() used everywhere else in the project.
    Returns {"DN": {"sdf": (s_name, r_name), "timer": (s_name, r_name)},
             "UP": {...}}.
    """
    import openpyxl
    from relay_rack_core import split_bit_name, read_name2_whitelist

    result = {"DN": {"sdf": ("", ""), "timer": ("", "")}, "UP": {"sdf": ("", ""), "timer": ("", "")}}
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb["FIELD PG.NO"]
        whitelist = read_name2_whitelist(xlsx_path)
        dn_timer = ws.cell(row=2, column=5).value  # E2
        dn_sdf = ws.cell(row=2, column=6).value    # F2
        up_timer = ws.cell(row=3, column=5).value  # E3
        up_sdf = ws.cell(row=3, column=6).value    # F3

        if dn_timer:
            result["DN"]["timer"] = split_bit_name(str(dn_timer).strip(), whitelist)
        if dn_sdf:
            result["DN"]["sdf"] = split_bit_name(str(dn_sdf).strip(), whitelist)
        if up_timer:
            result["UP"]["timer"] = split_bit_name(str(up_timer).strip(), whitelist)
        if up_sdf:
            result["UP"]["sdf"] = split_bit_name(str(up_sdf).strip(), whitelist)
    except Exception:
        pass
    return result


def generate_one_sdf_sheet(
    track_group: list, sht_number: str, cont_number: str, border_template_path: str, l4_groups: list, m4_groups: list,
    registry=None, sdf_bit_names: dict = None,
):
    """
    Builds ONE SDF sheet covering a group of same-direction tracks.
    track_group: list of track row dicts (all same DIRECTION).
    l4_groups: parsed FIELD PG.NO!L4 groups, e.g. [['A5','A6'], ['A7','A8']] -
        CONFIRMED: cycles per track (same group used for BOTH TS(M) and TS(R)
        within one track).
    m4_groups: parsed FIELD PG.NO!M4 groups, e.g. [['D1','D2']] - the group's
        1st value goes to TSPR, the 2nd ("free") value goes to the shared
        SDF relay's own Front_Contact.
    Returns (output_filename, ezdxf.Drawing).
    """
    if not track_group:
        raise ValueError("track_group is empty")

    def _apply_back_contact(e, group, relay_name=None):
        letter1, num1 = _afc_for_subvalue(group[0])
        letter2, num2 = _afc_for_subvalue(group[1])
        for att in e.attribs:
            if att.dxf.tag == "A":
                att.dxf.text = num1
            elif att.dxf.tag == "B":
                att.dxf.text = num2
            elif att.dxf.tag == "C":
                att.dxf.text = letter1
        if registry is not None and relay_name:
            registry.register_pair(relay_name, letter1, num1, num2, "SDF", sht_number)

    def _apply_front_contact(e, subvalue, r_name_prefix_direction=None, relay_name=None, sdf_relay_suffix=""):
        letter, num = _afc_for_subvalue(subvalue)
        sdf_s_name, sdf_r_name = (sdf_bit_names or {}).get(r_name_prefix_direction, {}).get("sdf", ("", ""))
        for att in e.attribs:
            if att.dxf.tag == "F":
                att.dxf.text = num
            elif att.dxf.tag == "A":
                att.dxf.text = str(int(num) + 1)
            elif att.dxf.tag == "C":
                att.dxf.text = letter
            elif att.dxf.tag == "R_NAME" and r_name_prefix_direction is not None:
                # CONFIRMED: prefer the exact relay from FIELD PG.NO's
                # SDF BIT column (F2/F3) when given - it's not always
                # just "{direction} SDF", since a station prefix may or
                # may not apply. Falls back to the old direction-prefix
                # convention if FIELD PG.NO doesn't have a value. Either
                # way, an overflow relay suffix (e.g. "1" for tracks 9+)
                # gets appended, matching the "#SDF1" naming Krish
                # confirmed - only applies to the AUTO-generated SDF
                # bit contacts, never to TS(M)/TS(R)/TSPR.
                att.dxf.text = (sdf_r_name if sdf_r_name else f"{_safe_str(r_name_prefix_direction)} SDF") + sdf_relay_suffix
            elif att.dxf.tag == "S_NAME" and r_name_prefix_direction is not None and sdf_s_name:
                att.dxf.text = sdf_s_name
        if registry is not None and relay_name:
            registry.register_pair(relay_name, letter, num, int(num) + 1, "SDF", sht_number)

    # Base doc: SDF_CIRCUIT.dxf, unchanged position - provides the fuse,
    # first track's own contacts, and the shared JR/SDF coil.
    # CONFIRMED: when a direction has only ONE track (no additional
    # tracks to stack), use the corrected SDF_CIRCUIT_SINGLE_TSPR.dxf
    # template instead of the normal base one - Krish's own file, kept
    # as a SEPARATE template rather than replacing SDF_CIRCUIT.dxf
    # entirely, since multi-track directions should still use the
    # original base template as before.
    base_template_name = "SDF_CIRCUIT_SINGLE_TSPR.dxf" if len(track_group) == 1 else "SDF_CIRCUIT.dxf"
    doc = ezdxf.readfile(resolve_template_path(base_template_name))
    msp = doc.modelspace()

    first_track = track_group[0]
    first_l4_group = l4_groups[0 % len(l4_groups)]
    tspr_value = m4_groups[0][0]  # M4 provides TSPR's Excel-given value

    # CONFIRMED: DNSDF is a SEPARATE relay with NO Excel input at all -
    # its own contact is auto-generated independently (not derived from
    # TSPR), using its OWN independent A/B/C/D contact space (a different
    # physical relay from whatever L4/M4 reference). CONFIRMED (per the
    # 8F-8B configuration table): only F=1 or F=3 are valid Front_Contact
    # positions per letter (F=5/F=7 are reserved for Back_Contact pairs) -
    # giving 8 valid unique pairs total: A1,A3,B1,B3,C1,C3,D1,D3.
    SDF_VALID_PAIRS = [f"{letter}{num}" for letter in "ABCD" for num in (1, 3)]

    def _sdf_auto_value(track_index):
        # CONFIRMED: SDF's own "# SDF" repeater relay only has capacity
        # for 8 tracks worth of bit positions - once a direction has
        # MORE than 8 tracks, tracks 9+ overflow onto a SECOND repeater
        # relay ("# SDF1"), matching the same LCPR-style overflow
        # pattern (0 -> base relay, 1 -> "1" suffix, 2 -> "2" suffix...).
        relay_number = track_index // len(SDF_VALID_PAIRS)
        relay_suffix = "" if relay_number == 0 else str(relay_number)
        value = SDF_VALID_PAIRS[track_index % len(SDF_VALID_PAIRS)]
        return value, relay_suffix

    # Direction (#) replacement - CONFIRMED: can appear ANYWHERE (any
    # attribute, any entity) in either SDF_CIRCUIT.dxf or
    # SDF_CIRCUIT_CONTACTS.dxf, not just one specific attribute.
    def _replace_direction_hash(entities, direction):
        for e in entities:
            if e.dxftype() == "INSERT":
                for att in e.attribs:
                    if att.dxf.text and "#" in att.dxf.text:
                        att.dxf.text = att.dxf.text.replace("#", _safe_str(direction))

    direction = first_track.get("DIRECTION", "")
    _replace_direction_hash(list(msp), direction)

    # CONFIRMED: TIMER_RELAY's S_NAME comes from FIELD PG.NO's SDF TIMER
    # column (E2/E3), same convention as the SDF block's own S_NAME -
    # its R_NAME already gets the direction substituted via the generic
    # "#"-replacement above.
    timer_s_name, _timer_r_name = (sdf_bit_names or {}).get(direction, {}).get("timer", ("", ""))
    if timer_s_name:
        for e in msp:
            if e.dxftype() == "INSERT" and e.dxf.name == "TIMER_RELAY":
                for att in e.attribs:
                    if att.dxf.tag == "S_NAME":
                        att.dxf.text = timer_s_name

    # CONFIRMED: SDF_RELAY (the coil) was never touched at all before -
    # its S_NAME stayed at the template's raw placeholder text forever,
    # which never matches Relay Rack, making the coil always show as
    # "never used". Its S_NAME needs the SAME FIELD PG.NO SDF BIT value
    # as SDF_CONTACT (they're the same relay's contact vs coil) - its
    # R_NAME already gets the direction substituted via the generic
    # "#"-replacement above.
    sdf_s_name_for_coil, _sdf_r_name = (sdf_bit_names or {}).get(direction, {}).get("sdf", ("", ""))
    for e in msp:
        if e.dxftype() == "INSERT" and e.dxf.name == "SDF_RELAY":
            for att in e.attribs:
                if att.dxf.tag == "S_NAME":
                    att.dxf.text = sdf_s_name_for_coil
            # CONFIRMED: NOT registered with ContactRegistry - this
            # coil is read from a single fixed Excel cell per direction
            # (FIELD PG.NO F2/F3), so it's intentionally the SAME relay
            # across every SDF sheet of that direction. Registering it
            # produced a false "repetition" error on any station with
            # more than one SDF sheet per direction - reverted.

    for e in msp:
        if e.dxftype() == "INSERT" and e.dxf.name in ("Back_Contact", "Front_Contact", "SDF_CONTACT"):
            r_name = next((a.dxf.text for a in e.attribs if a.dxf.tag == "R_NAME"), "")
            if r_name in ("TS(M)", "TS(R)"):
                for att in e.attribs:
                    if att.dxf.tag == "S_NAME":
                        att.dxf.text = _safe_str(first_track["TRACK NAME"])
                _apply_back_contact(e, first_l4_group, relay_name=first_track['TRACK NAME'] if r_name == "TS(M)" else None)
            elif r_name == "TSPR":
                for att in e.attribs:
                    if att.dxf.tag == "S_NAME":
                        att.dxf.text = _safe_str(first_track["TRACK NAME"])
                _apply_front_contact(e, tspr_value, relay_name=first_track['TRACK NAME'])
            elif "SDF" in r_name:  # matches both plain "SDF" and "#SDF"
                # CONFIRMED: S_NAME left untouched (template default) - DNSDF
                # is one shared relay, not per-track like TS(M)/TS(R)/TSPR.
                sdf_value, sdf_relay_suffix = _sdf_auto_value(0)
                _apply_front_contact(e, sdf_value, r_name_prefix_direction=direction, sdf_relay_suffix=sdf_relay_suffix)

    # Additional tracks (2nd onward): stack SDF_CIRCUIT_CONTACTS.dxf below.
    # CONFIRMED: every gap (including 1st-to-2nd) is exactly Y_SPACING,
    # for tracks 2-4 (column 1, which only has room for 4 tracks total
    # including track 1). Track 5 onward overflows into a SECOND column:
    #   - Track 5 (first of column 2) needs TWO pieces placed together:
    #     SDF_CONT.dxf (place point (64.8089, 139.5203), CONFIRMED) and
    #     SDF_CONT_UP_1ST.dxf (pick (139.239, 213.247), place
    #     (66.2320, 166.6123), CONFIRMED).
    #   - Track 6 onward (column 2) goes back to SDF_CIRCUIT_CONTACTS.dxf,
    #     starting at (70.2629, 139.5203) for track 6, CONFIRMED, then
    #     continuing to stack downward by Y_SPACING same as column 1.
    #   - Whichever track is the LAST one overall (regardless of which
    #     column it falls in) uses the "_LAST" variant of whatever file
    #     it would have otherwise used - CONFIRMED, per Krish's own
    #     table: SDF_CIRCUIT_CONTACTS.dxf -> SDF_CIRCUIT_CONTACTS_LAST.dxf,
    #     SDF_CONT.dxf -> SDF_CIRCUIT_CONTACTS_LAST.dxf (track 5's first
    #     piece specifically, if track 5 is last), and
    #     SDF_CONT_UP_1ST.dxf -> SDF_CONT_UP_LAST.dxf.
    NATIVE_INHERENT_OFFSET = 32.1206
    # CONFIRMED: static, explicitly given place positions for tracks
    # 2-7 (the gaps between them aren't perfectly uniform, so these are
    # exact values rather than a formula). Track 8+ is calculated,
    # continuing the gap established between track 6 and track 7
    # (35.4206), since that's the last confirmed pair before the
    # explicit list ends.
    STATIC_COLUMN1_X = 259.8977
    STATIC_COLUMN2_X = 123.2021
    TRACK_STATIC_PLACE = {
        2: (STATIC_COLUMN1_X, 232.5538),
        3: (STATIC_COLUMN1_X, 197.1332),
        4: (STATIC_COLUMN1_X, 161.7126),
        6: (STATIC_COLUMN2_X, 183.7572),
        7: (STATIC_COLUMN2_X, 148.3366),
    }
    TRACK_6_TO_7_GAP = 183.7572 - 148.3366  # = 35.4206, confirmed
    SDF_CONT_PICK = (259.8977, 197.4175)
    SDF_CONT_PLACE = (259.8977, 126.0077)  # confirmed
    SDF_CONT_UP_1ST_PICK = (139.239, 213.247)
    SDF_CONT_UP_1ST_PLACE = (66.2320, 166.6123)
    COLUMN2_CONTACT_PICK_POINT = (259.898, 232.554)

    # CONFIRMED: SDF_CONT.dxf is a SEPARATE, one-time structural piece -
    # NOT part of any specific track's own content - inserted exactly
    # once whenever the direction has MORE than 4 tracks (i.e. whenever
    # column 2 is needed at all), alongside whatever track 5's own
    # SDF_CONT_UP content is.
    if len(track_group) > 4:
        pasted_cont = _import_whole_file_no_rename(doc, "SDF_CONT.dxf", SDF_CONT_PICK, SDF_CONT_PLACE)
        _replace_direction_hash(pasted_cont, direction)

    for i, track in enumerate(track_group[1:], start=1):
        track_index_1based = i + 1  # track 1 is track_group[0], so this is 2, 3, 4, 5...
        is_last_track = track_index_1based == len(track_group)
        track_l4_group = l4_groups[i % len(l4_groups)]  # CONFIRMED: cycles per track

        def _apply_track_contacts(pasted, track, is_last):
            _replace_direction_hash(pasted, track.get("DIRECTION", ""))
            for e in pasted:
                if e.dxftype() == "INSERT" and e.dxf.name.startswith(("Back_Contact", "Front_Contact", "SDF_CONTACT")):
                    r_name = next((a.dxf.text for a in e.attribs if a.dxf.tag == "R_NAME"), "")
                    if r_name in ("TS(M)", "TS(R)"):
                        for att in e.attribs:
                            if att.dxf.tag == "S_NAME":
                                att.dxf.text = _safe_str(track["TRACK NAME"])
                        _apply_back_contact(e, track_l4_group, relay_name=track['TRACK NAME'] if r_name == "TS(M)" else None)
                    elif r_name == "TSPR":
                        for att in e.attribs:
                            if att.dxf.tag == "S_NAME":
                                att.dxf.text = _safe_str(track["TRACK NAME"])
                        _apply_front_contact(e, tspr_value, relay_name=track['TRACK NAME'])
                    elif "SDF" in r_name:
                        sdf_value, sdf_relay_suffix = _sdf_auto_value(i)
                        _apply_front_contact(e, sdf_value, r_name_prefix_direction=track.get("DIRECTION", ""), sdf_relay_suffix=sdf_relay_suffix)

        if track_index_1based <= 4:
            # Column 1, tracks 2-4: static confirmed positions.
            place_point = TRACK_STATIC_PLACE[track_index_1based]
            filename = "SDF_CIRCUIT_CONTACTS_LAST.dxf" if is_last_track else "SDF_CIRCUIT_CONTACTS.dxf"
            pasted = _import_whole_file_no_rename(doc, filename, CONTACT_PICK_POINT, place_point)
            _apply_track_contacts(pasted, track, is_last_track)

        elif track_index_1based == 5:
            # Column 2, first track: CONFIRMED - only ONE piece placed,
            # not two. SDF_CONT.dxf is NOT used at track 5 at all -
            # only SDF_CONT_UP_1ST.dxf (or SDF_CONT_UP_LAST.dxf if this
            # is the last track), placed at SDF_CONT_UP_1ST_PLACE (its
            # own original, already-correct position - NOT the old
            # SDF_CONT.dxf position).
            up_1st_filename = "SDF_CONT_UP_LAST.dxf" if is_last_track else "SDF_CONT_UP_1ST.dxf"
            pasted_up = _import_whole_file_no_rename(doc, up_1st_filename, SDF_CONT_UP_1ST_PICK, SDF_CONT_UP_1ST_PLACE)
            _apply_track_contacts(pasted_up, track, is_last_track)

        elif track_index_1based in (6, 7):
            # Column 2, tracks 6-7: static confirmed positions.
            place_point = TRACK_STATIC_PLACE[track_index_1based]
            filename = "SDF_CIRCUIT_CONTACTS_LAST.dxf" if is_last_track else "SDF_CIRCUIT_CONTACTS.dxf"
            pasted = _import_whole_file_no_rename(doc, filename, COLUMN2_CONTACT_PICK_POINT, place_point)
            _apply_track_contacts(pasted, track, is_last_track)

        else:
            # Column 2, track 8 onward: CALCULATED, continuing the
            # gap established between track 6 and track 7 (35.4206).
            tracks_past_7 = track_index_1based - 7
            place_y = TRACK_STATIC_PLACE[7][1] - tracks_past_7 * TRACK_6_TO_7_GAP
            filename = "SDF_CIRCUIT_CONTACTS_LAST.dxf" if is_last_track else "SDF_CIRCUIT_CONTACTS.dxf"
            pasted = _import_whole_file_no_rename(doc, filename, COLUMN2_CONTACT_PICK_POINT, (STATIC_COLUMN2_X, place_y))
            _apply_track_contacts(pasted, track, is_last_track)

    title_text = f"SDF CIRCUIT - {direction}"
    _insert_border_title_generic(
        doc=doc,
        border_template_path=border_template_path,
        template_name="SDF",
        sig_name=None,
        ahead_sig=None,
        loc=None,
        sht_number=sht_number,
        cont_number=cont_number,
        title_text_override=title_text,
    )

    out_name = f"SDF_{first_track['TRACK NAME']}_SHT{sht_number}.dxf"
    return out_name, doc


def get_sdf_start_sheet_number(xlsx_path: str) -> int:
    """Reads SDF CIRCUITS's starting sheet number from FIELD PG.NO."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(row=r, column=2).value or "").strip().upper()
        if label == "SDF CIRCUITS":
            value = ws.cell(row=r, column=3).value
            if value is None or str(value).strip() == "":
                raise ValueError("FIELD PG.NO has no Sheet Number for the SDF CIRCUITS row")
            return int(str(value).strip())
    raise ValueError("Could not find a 'SDF CIRCUITS' row in FIELD PG.NO")


def get_next_circuit_after_sdf(xlsx_path: str) -> str:
    """Reads the starting sheet number of whatever circuit follows SDF CIRCUITS in FIELD PG.NO."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(row=r, column=2).value or "").strip().upper()
        if label == "SDF CIRCUITS":
            value = ws.cell(row=r + 1, column=3).value
            return str(value).strip() if value is not None else ""
    return ""


def generate_all_sdf_sheets(xlsx_path: str, start_sheet_number: int, border_template_path: str, registry=None):
    """
    Generates SDF sheets: one per group of consecutive same-direction tracks.
    Returns (list of (output_filename, ezdxf.Drawing, sht_number_str), next_sheet_number)
    """
    rows = read_track_rows(xlsx_path)
    if not rows:
        raise ValueError("No rows found in TRACK sheet")

    groups = _group_by_direction(rows)
    l4_groups, m4_groups = get_sdf_contact_groups(xlsx_path)
    sdf_bit_names = _read_sdf_bit_names(xlsx_path)

    results = []
    sheet_num = start_sheet_number
    for group in groups:
        sht = f"{sheet_num:03d}"
        cont = f"{sheet_num + 1:03d}"
        try:
            out_name, doc = generate_one_sdf_sheet(
                group, sht, cont, border_template_path, l4_groups, m4_groups,
                registry=registry, sdf_bit_names=sdf_bit_names,
            )
        except Exception as e:
            raise ValueError(f"Track group starting {group[0]['TRACK NAME']!r}: {e}") from e
        results.append((out_name, doc, sht))
        sheet_num += 1

    return results, sheet_num
