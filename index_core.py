"""
INDEX sheet generation - a table of contents listing every generated
sheet's S.NO / DESCRIPTION / SHEET NUMBER.

CONFIRMED: uses two real template files (not hand-drawn primitives):
  - INDEX_HEADER.dxf: one header row (S.NO / DESCRIPTION / SHEET NUMBER
    captions + its own grid lines), positioned at the LEFT column's
    coordinates in the source file.
  - INDEX_TITLE.dxf: one DATA row (grid lines + MTEXT placeholders
    "{1}" for S.NO, "{TITLE}" for description, "{FILE NAME}" for the
    sheet's final filename), also at LEFT column coordinates. Its own
    top edge coincides with the header's bottom edge, and each
    subsequent row's top coincides with the previous row's bottom -
    stacking copies translated down by ROW_SPACING recreates the full
    table grid with no gaps or overlaps.

CONFIRMED layout (measured from Krish's reference files):
  - Row spacing: 5.9867 units.
  - A FULL page holds 33 rows in a LEFT column + 33 in a RIGHT column
    (66 rows/sheet) - S.NO runs 1-33 left, 34-66 right, continuing
    across sheets (not reset per sheet or per column).
  - If a page's remaining content fits in ONE column only (<=33 rows -
    e.g. the last page), that column is CENTERED on the sheet instead
    of using the LEFT position (confirmed).
  - TITLE follows "INDEX - N" (N = page number), same pattern as
    Contact Analysis's own incrementing title.
  - SHEET NUMBER shows the final generated filename (already computed
    via _build_output_filename elsewhere), without its .dxf extension.
"""

import ezdxf
from ezdxf.addons import importer
import re

from signal_core import insert_border_title, resolve_template_path

TITLE_TEXT_PREFIX = "INDEX - "

ROW_SPACING = 5.9867
ROWS_PER_COLUMN = 33

# CONFIRMED: X-offset to shift a LEFT-column copy into the RIGHT column
# or the CENTERED single-column position, measured directly from the
# reference files' own grid line positions.
RIGHT_X_OFFSET = 193.99   # RIGHT_GRID_X[0] (216.79) - LEFT_GRID_X[0] (22.8)
CENTER_X_OFFSET = 98.02   # CENTER_GRID_X[0] (120.82) - LEFT_GRID_X[0] (22.8)


def _next_plain_sheet_number(sht: str) -> str:
    """Increments a purely-numeric sheet number, preserving its zero-padding
    (e.g. '062' -> '063', '099' -> '100')."""
    width = len(sht)
    return str(int(sht) + 1).zfill(width)


def _estimate_page_count(total_rows: int) -> int:
    """Lightweight version of generate_all_index_sheets' own pagination
    loop - just counts how many pages `total_rows` would need, without
    building any DXF content. Used to resolve the self-listing
    chicken-and-egg problem below."""
    if total_rows == 0:
        return 0
    pages = 0
    idx = 0
    while idx < total_rows:
        remaining = total_rows - idx
        take = remaining if remaining <= ROWS_PER_COLUMN else min(remaining, ROWS_PER_COLUMN * 2)
        idx += take
        pages += 1
    return pages


def add_index_self_listing(ordered_sheets: list, index_start_sht: str, filename_pattern: str) -> list:
    """
    CONFIRMED: INDEX lists ITSELF too (e.g. "INDEX - 1"/"INDEX - 2" rows
    inside the table, matching Krish's own reference file). Since how
    many INDEX pages are needed depends on the total row count, which
    includes INDEX's own rows, this resolves it with a small fixed-point
    iteration: estimate page count, add that many self-listing rows,
    re-estimate, repeat until the count stops changing (converges in 1-2
    iterations in practice).

    ordered_sheets: list of (description, sht_number, final_filename)
        for every OTHER sheet (NOT including INDEX itself), already
        sorted by sheet number.
    Returns: ordered_sheets with INDEX's own self-listing rows merged in
    (NOT yet re-sorted - caller should sort the result by sheet number
    using _sheet_number_sort_key before passing to
    build_index_entries_with_spare_gaps).
    """
    other_count = len(ordered_sheets)
    page_count = 0
    while True:
        new_page_count = _estimate_page_count(other_count + page_count)
        if new_page_count == page_count:
            break
        page_count = new_page_count

    if page_count == 0:
        return list(ordered_sheets)

    self_rows = []
    sht = index_start_sht
    for i in range(page_count):
        filename = _build_output_filename_range(sht, filename_pattern)
        self_rows.append((f"{TITLE_TEXT_PREFIX}{i + 1}", sht, filename))
        sht = _increment_alpha_suffix(sht)

    return list(ordered_sheets) + self_rows


def build_index_entries_with_spare_gaps(ordered_sheets: list, filename_pattern: str) -> list:
    """
    CONFIRMED: if there's a GAP between two consecutively-generated
    sheets' numbers (e.g. sheet 66 is followed by sheet 70, meaning
    67/68/69 were reserved but never actually generated), the INDEX
    collapses the whole gap into ONE row: DESCRIPTION="SPARE",
    SHEET NUMBER=the filename pattern applied to "067-069" (a range,
    not a single number) - matching Krish's confirmed example and the
    "NNN-NNN" range pattern already seen in his own reference file.

    ordered_sheets: list of (description, sht_number, final_filename_no_ext)
        tuples for every sheet ACTUALLY generated, in generation order.
        final_filename_no_ext is the real filename already built via
        _build_output_filename() elsewhere in the app.
    filename_pattern: the SAME pattern used everywhere else (passed to
        _build_output_filename), reused here (via the same '[]'
        convention) so a SPARE range's filename looks consistent with
        every other row.

    Returns: list of (description, filename) tuples ready for
    generate_all_index_sheets(), with SPARE gap rows inserted as needed.
    Only handles gaps between purely-numeric sheet numbers (e.g. "062"
    -> "066") - alphanumeric sheet number sequences (e.g. "00A"->"00C")
    are not gap-checked, since those sections don't use the "SPARE
    SHEETS" reservation convention.
    """
    entries = []
    for i, (description, sht_number, final_filename) in enumerate(ordered_sheets):
        if i > 0:
            prev_sht = ordered_sheets[i - 1][1]
            if prev_sht.isdigit() and sht_number.isdigit():
                expected_next = _next_plain_sheet_number(prev_sht)
                if sht_number != expected_next and int(sht_number) > int(expected_next):
                    gap_start = expected_next
                    gap_end = str(int(sht_number) - 1).zfill(len(sht_number))
                    range_sht = f"{gap_start}-{gap_end}"
                    spare_filename = _build_output_filename_range(range_sht, filename_pattern)
                    entries.append(("SPARE", spare_filename))
        entries.append((description, final_filename))
    return entries


def _build_output_filename_range(range_sht: str, pattern: str) -> str:
    pattern = (pattern or "").strip()
    if not pattern:
        return range_sht
    if "[]" in pattern:
        return pattern.replace("[]", range_sht)
    return f"{pattern}{range_sht}"


def _copy_translated(source_doc, target_doc, dx, dy):
    """Copies every modelspace entity from source_doc into target_doc's
    modelspace, translated by (dx, dy). Returns the list of new entities
    (so callers can further edit e.g. MTEXT placeholder substitution)."""
    target_msp = target_doc.modelspace()
    existing_ids = set(id(e) for e in target_msp)
    imp = importer.Importer(source_doc, target_doc)
    imp.import_modelspace()
    imp.finalize()
    new_entities = [e for e in target_msp if id(e) not in existing_ids]
    for e in new_entities:
        e.translate(dx, dy, 0)
    return new_entities


def _substitute_row_placeholders(entities, sno, description, sheet_number):
    for e in entities:
        if e.dxftype() == "MTEXT":
            new_text = (
                e.text.replace("{1}", str(sno))
                .replace("{TITLE}", description)
                .replace("{FILE NAME}", sheet_number)
            )
            e.text = new_text
        elif e.dxftype() == "TEXT":
            new_text = (
                e.dxf.text.replace("{1}", str(sno))
                .replace("{TITLE}", description)
                .replace("{FILE NAME}", sheet_number)
            )
            e.dxf.text = new_text


def get_index_start_sheet_number(xlsx_path: str) -> str:
    """Reads INDEX's starting sheet number from FIELD PG.NO (e.g. '00A') -
    a string, not an int, since INDEX sheet numbers are alphanumeric."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(row=r, column=2).value or "").strip().upper()
        if label == "INDEX":
            value = ws.cell(row=r, column=3).value
            if value is None or str(value).strip() == "":
                raise ValueError("FIELD PG.NO has no Sheet Number for the INDEX row")
            return str(value).strip()
    raise ValueError("Could not find an 'INDEX' row in FIELD PG.NO")


def _sheet_number_sort_key(sht: str):
    """
    CONFIRMED FIX: sheet numbers like '00F' (Relay Rack) need to sort
    BETWEEN '000' and '001', not after every purely-numeric sheet number.
    Splits into (leading digits as int, trailing letters) - e.g. '00F' ->
    (0, 'F'), '001' -> (1, ''), '000' -> (0, ''). Since '' < 'F'
    alphabetically, '000' < '00A' < ... < '00F' < '001' < '002' sorts
    correctly as one continuous, correctly-ordered sequence.
    """
    m = re.match(r"^(\d+)([A-Za-z]*)$", sht)
    if not m:
        return (float("inf"), sht)
    return (int(m.group(1)), m.group(2))


def generate_one_index_sheet(rows: list, sht_number: str, cont_number: str,
                              border_template_path: str, page_number: int):
    """
    rows: list of (sno, description, sheet_number) tuples, ALREADY split
    for this one page (up to 66 - 33 left + 33 right, or fewer for a
    centered partial page).
    Returns (out_name, ezdxf.Drawing).
    """
    header_doc = ezdxf.readfile(resolve_template_path("INDEX_HEADER.dxf"))
    row_doc = ezdxf.readfile(resolve_template_path("INDEX_TITLE.dxf"))

    doc = ezdxf.new()

    if len(rows) <= ROWS_PER_COLUMN:
        # CONFIRMED: a page whose content fits in ONE column is
        # CENTERED on the sheet, not anchored at the left position.
        x_offset = CENTER_X_OFFSET
        columns = [rows]
    else:
        x_offset = 0  # LEFT column uses the template's own native position
        columns = [rows[:ROWS_PER_COLUMN], rows[ROWS_PER_COLUMN:]]

    for col_index, col_rows in enumerate(columns):
        # Second column (when there are two) is the RIGHT column, offset
        # from the LEFT position - the CENTERED single-column case never
        # reaches here with 2 columns.
        dx = x_offset if len(columns) == 1 else (0 if col_index == 0 else RIGHT_X_OFFSET)
        _copy_translated(header_doc, doc, dx, 0)
        for i, (sno, description, sheet_number) in enumerate(col_rows):
            row_entities = _copy_translated(row_doc, doc, dx, -i * ROW_SPACING)
            _substitute_row_placeholders(row_entities, sno, description, sheet_number)

    insert_border_title(
        doc=doc,
        border_template_path=border_template_path,
        template_name="INDEX",
        sig_name="", ahead_sig="", loc="",
        sht_number=str(sht_number),
        cont_number=str(cont_number),
        title_text_override=f"{TITLE_TEXT_PREFIX}{page_number}",
    )

    out_name = f"INDEX_SHT{sht_number}.dxf"
    return out_name, doc


def _increment_alpha_suffix(sht: str) -> str:
    """'00A' -> '00B' -> ... -> '00Z' -> '00AA' (prefix stays, trailing
    letters increment like a base-26 counter)."""
    m = 0
    while m < len(sht) and sht[len(sht) - 1 - m].isalpha():
        m += 1
    if m == 0:
        return str(int(sht) + 1)
    prefix, suffix = sht[:-m], sht[-m:]
    chars = list(suffix)
    i = len(chars) - 1
    while i >= 0:
        if chars[i] != "Z":
            chars[i] = chr(ord(chars[i]) + 1)
            break
        chars[i] = "A"
        i -= 1
    else:
        chars.insert(0, "A")
    return prefix + "".join(chars)


def generate_all_index_sheets(entries: list, start_sheet_number: str, border_template_path: str):
    """
    entries: list of (description, final_filename_no_ext) tuples, in the
    order they should be numbered.
    start_sheet_number: e.g. "00A" (from FIELD PG.NO's INDEX row) - a
    string, not an int, since INDEX sheet numbers are alphanumeric.
    Returns (list of (out_name, doc, sht_number_str), next_sheet_number).
    """
    numbered_rows = [(i + 1, desc, fname) for i, (desc, fname) in enumerate(entries)]

    results = []
    sht = start_sheet_number
    page_number = 1
    idx = 0
    while idx < len(numbered_rows):
        remaining = len(numbered_rows) - idx
        take = remaining if remaining <= ROWS_PER_COLUMN else min(remaining, ROWS_PER_COLUMN * 2)
        page_rows = numbered_rows[idx: idx + take]
        idx += take

        next_sht = _increment_alpha_suffix(sht)
        out_name, doc = generate_one_index_sheet(page_rows, sht, next_sht, border_template_path, page_number)
        results.append((out_name, doc, sht))
        sht = next_sht
        page_number += 1

    return results, sht
