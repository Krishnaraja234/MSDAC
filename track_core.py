"""
MSDAC Track Circuit Generator - core substitution logic.

Mirrors signal_core.py's architecture (TYPICAL-column-driven sheet
selection, same token substitution rules, sheet numbering sourced from
FIELD PG.NO), with one Track-specific twist: up to TWO track rows share
a single physical sheet, placed side-by-side at a fixed X offset.

TRACK sheet columns: TRACK NAME, DIRECTION, TYPICAL (this column was
named REQUIREMENT before - both names are supported so this keeps
working whichever header the workbook currently uses).

CONFIRMED rules:
  - TYPICAL value picks a set of template files, exactly like Signal's
    TYPICAL_SHEET_SETS. Currently only "TSPR" exists, mapping to
    [TSPR.dxf, TSPR1.dxf] - two independent sheets per pair of tracks.
  - Up to 2 TRACK rows are paired onto one sheet, side-by-side. If an odd
    row is left over, it occupies slot 1 alone (slot 2 stays empty).
  - The second track's circuit is shifted by X_OFFSET (measured from the
    reference file: ~175.1634 units), Y unchanged.
  - '*' token (S_NAME attributes, TITLE text) -> TRACK NAME, per slot.
  - '#' inside FUZEBLOCK/FUZEENDBLOCK VOLT -> DIRECTION, per slot.
  - TITLE combines both track names in one sheet: "TSPR CIRCUITS - <name1>
    & <name2>" (or just "<name1>" if a lone track occupies the last slot).
  - Sheet numbering starts from FIELD PG.NO's TRACK row, and the very
    last Track sheet's CONT points to whatever circuit type follows
    TRACK in FIELD PG.NO (e.g. "AR & AZR CIRCUITS").

ASSUMPTIONS still to be confirmed once real (non-placeholder) TSPR1.dxf
content is available - see NOTE near TRACK_TYPICAL_SHEET_SETS below.
"""

import ezdxf
from ezdxf.addons import importer
import os

from signal_core import (
    TEMPLATES_DIR,
    resolve_template_path,
    TOKEN_STAR,
    _safe_str,
    insert_border_title as _insert_border_title_generic,
)

# Measured from track_reference.dxf: X distance between slot 1 and slot 2
# (e.g. Relay_Coil_ACI TS(A) at x=145.46 for track 1 vs x=320.62 for
# track 2 -> offset = 175.1634). Y is unchanged between slots.
X_OFFSET = 175.1634301509164

# TYPICAL value -> ordered list of template sheets. Same pattern as
# Signal's TYPICAL_SHEET_SETS.
# NOTE: TSPR.dxf and TSPR1.dxf are currently IDENTICAL placeholder files
# (both show R_NAME='TSPR', not 'TSPR1' for the second one) - this is
# expected to be corrected on your end. The code below works with
# whatever these two files actually contain; once they're finalized,
# re-run a test generation to confirm titles/positioning still look right.
TRACK_TYPICAL_SHEET_SETS = {
    "TSPR": ["TSPR.dxf"],
    "TSPR1": ["TSPR1.dxf"],
}

# TITLE text per track typical template (same idea as Signal's TITLE_TEXT_MAP).
TRACK_TITLE_TEXT_MAP = {
    "TSPR.dxf": "TSPR CIRCUITS - {names}",
    "TSPR1.dxf": "TSPR CIRCUITS - {names}",
}


def resolve_track_sheet_set(typical: str) -> list[str]:
    """Return the ordered list of template filenames for a TRACK TYPICAL value."""
    typical = (typical or "").strip()
    sheet_set = TRACK_TYPICAL_SHEET_SETS.get(typical)
    if sheet_set is None:
        raise ValueError(f"Unknown TRACK TYPICAL value: {typical!r}")
    return list(sheet_set)


def _get_track_row_typical(row: dict) -> str:
    """TYPICAL column, falling back to the old REQUIREMENT header name."""
    return row.get("TYPICAL") or row.get("REQUREMENT") or row.get("REQUIREMENT") or ""


def _substitute_track_template(template_path: str, track_name, direction: str, hut_name_input=None):
    """
    Load a track typical template and substitute placeholders for one
    track instance. Returns an ezdxf Drawing object.
    """
    doc = ezdxf.readfile(template_path)
    msp = doc.modelspace()

    def _apply_hut_name(text):
        if text and "HUT NAME" in text:
            return text.replace("HUT NAME", _safe_str(hut_name_input))
        return text

    to_delete = []
    for e in msp:
        if e.dxftype() == "TEXT":
            if e.dxf.text and TOKEN_STAR in e.dxf.text:
                e.dxf.text = e.dxf.text.replace(TOKEN_STAR, _safe_str(track_name))
            e.dxf.text = _apply_hut_name(e.dxf.text)
            continue
        if e.dxftype() == "MTEXT":
            if e.text and TOKEN_STAR in e.text:
                e.text = e.text.replace(TOKEN_STAR, _safe_str(track_name))
            e.text = _apply_hut_name(e.text)
            continue

        if e.dxftype() != "INSERT":
            continue

        if e.dxf.name == "TITLE":
            # Old baked-in title block - border template supplies TITLE now.
            to_delete.append(e)
            continue

        for att in e.attribs:
            tag = att.dxf.tag
            val = att.dxf.text
            if tag == "S_NAME" and val == TOKEN_STAR:
                att.dxf.text = _safe_str(track_name)
            elif tag == "VOLT" and "#" in (val or ""):
                att.dxf.text = val.replace("#", _safe_str(direction))
            elif val and "HUT NAME" in val:
                att.dxf.text = _apply_hut_name(val)

    for e in to_delete:
        msp.delete_entity(e)

    return doc


def _merge_second_track_into(base_doc, second_doc, x_offset: float):
    """Copy every entity from second_doc's modelspace into base_doc's
    modelspace, shifted by x_offset (Y unchanged)."""
    imp = importer.Importer(second_doc, base_doc)
    base_msp = base_doc.modelspace()
    existing_ids = set(id(e) for e in base_msp)

    imp.import_modelspace()
    imp.finalize()

    new_entities = [e for e in base_msp if id(e) not in existing_ids]
    for e in new_entities:
        try:
            e.translate(x_offset, 0, 0)
        except AttributeError:
            pass  # entity type without translate() - leave as-is


def generate_track_sheet_pair(
    template_name: str,
    track_row_a: dict,
    track_row_b: dict | None,
    sht_number: str,
    cont_number: str,
    border_template_path: str,
    hut_name_input=None,
):
    """
    Build ONE sheet from a single track typical template file, containing
    track_row_a in slot 1 and (if given) track_row_b in slot 2 at X_OFFSET.
    Returns (output_filename, ezdxf.Drawing).
    """
    template_path = resolve_template_path(template_name)

    name_a = track_row_a.get("TRACK NAME")
    direction_a = track_row_a.get("DIRECTION") or ""
    doc = _substitute_track_template(template_path, name_a, direction_a, hut_name_input)

    names_for_title = [str(name_a)]

    if track_row_b is not None:
        name_b = track_row_b.get("TRACK NAME")
        direction_b = track_row_b.get("DIRECTION") or ""
        doc_b = _substitute_track_template(template_path, name_b, direction_b, hut_name_input)
        _merge_second_track_into(doc, doc_b, X_OFFSET)
        names_for_title.append(str(name_b))

    title_template = TRACK_TITLE_TEXT_MAP.get(template_name, "TSPR CIRCUITS - {names}")
    title_text = title_template.format(names=" & ".join(names_for_title))

    _insert_border_title_generic(
        doc=doc,
        border_template_path=border_template_path,
        template_name=template_name,
        sig_name=None,
        ahead_sig=None,
        loc=None,
        sht_number=sht_number,
        cont_number=cont_number,
        title_text_override=title_text,
    )

    out_name = f"{'_'.join(names_for_title)}_{template_name.replace('.dxf', '')}_SHT{sht_number}.dxf"
    return out_name, doc


def generate_all_track_sheets(rows: list, start_sheet_number: int, border_template_path: str, hut_name_input=None):
    """
    rows: list of dicts with TRACK NAME, DIRECTION, TYPICAL (or REQUREMENT)
    start_sheet_number: first sheet number to assign
    border_template_path: path to the uploaded border/title-block template

    Returns: (list of (output_filename, ezdxf.Drawing, sht_number_str), next_sheet_number)
    """
    # Pair up rows 2-at-a-time, in order. (Rows with different TYPICAL
    # values are still paired sequentially - if that's wrong for your
    # real data, flag it and pairing can be restricted to same-TYPICAL rows only.)
    pairs = []
    i = 0
    while i < len(rows):
        a = rows[i]
        b = rows[i + 1] if i + 1 < len(rows) else None
        pairs.append((a, b))
        i += 2

    results = []
    sheet_num = start_sheet_number
    for row_a, row_b in pairs:
        typical = _get_track_row_typical(row_a)
        sheet_files = resolve_track_sheet_set(typical)

        for template_name in sheet_files:
            sht = f"{sheet_num:03d}"
            cont = f"{sheet_num + 1:03d}"
            out_name, doc = generate_track_sheet_pair(
                template_name, row_a, row_b, sht, cont, border_template_path, hut_name_input
            )
            results.append((out_name, doc, sht))
            sheet_num += 1

    return results, sheet_num


def get_track_start_sheet_number(xlsx_path: str) -> int:
    """Read TRACK's starting sheet number from FIELD PG.NO (row where
    Circuits == 'TRACK')."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(row=r, column=2).value or "").strip().upper() == "TRACK":
            value = ws.cell(row=r, column=3).value
            if value is None or str(value).strip() == "":
                raise ValueError("FIELD PG.NO has no Sheet Number for the TRACK row")
            return int(str(value).strip())
    raise ValueError("Could not find a 'TRACK' row in FIELD PG.NO")


def get_next_circuit_after_track(xlsx_path: str) -> str:
    """Read the starting sheet number of whatever circuit type follows
    TRACK in FIELD PG.NO (for the last Track sheet's CONT value)."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(row=r, column=2).value or "").strip().upper() == "TRACK":
            value = ws.cell(row=r + 1, column=3).value
            if value is None or str(value).strip() == "":
                raise ValueError("FIELD PG.NO has no Sheet Number for the row after TRACK")
            v = str(value).strip()
            return v.zfill(3) if v.isdigit() else v
    raise ValueError("Could not find a 'TRACK' row in FIELD PG.NO")
