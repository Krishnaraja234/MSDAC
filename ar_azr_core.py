"""
MSDAC AR & AZR Circuit Generator - core assembly logic.

Unlike Signal/Track (fixed templates with token substitution), AR & AZR
circuits are ASSEMBLED PROGRAMMATICALLY from four separate building
blocks, because the number of "provings" varies per row:

  - POSITIVE_FUSE.dxf  (FUZEBLOCK)    - circuit start
  - CONTACTS.dxf       (Front_Contact x2, a "pair") - one per proving
  - RELAY_COIL.dxf     (Relay_Coil)   - the AR/AZR relay itself
  - NEGATIVE_FUSE.dxf  (FUZEENDBLOCK) - circuit end

Layout rule (CONFIRMED): the Relay_Coil and NEGATIVE_FUSE always shift
right by (N-1) * CONTACT_PAIR_WIDTH, where N = number of provings, so
extra provings never overlap the coil/end-fuse. CONTACT_PAIR_WIDTH is
measured directly from the Front_Contact block's own bounding box.

AR & AZR sheet columns: "AR/AZR" (the row's own name, e.g. "SKL 2"),
DIRECTION, TYPE (AR or AZR - becomes the coil's R_NAME), PROVINGS.

PROVINGS format (CONFIRMED): comma-separated entries, each
"<name> <relay_type>", e.g. "2 TSPR, 45 TSPR, 21 AR1". Each entry
becomes ONE contact pair: S_NAME=<name>, R_NAME=<relay_type>.

A/F/C values on each contact (CONFIRMED): per FIELD PG.NO!H3, formatted
as groups separated by "/", each group containing exactly 2 comma-
separated values - one per contact in that proving's pair. E.g.
"D1,D2/D3,D4/D5,D6" gives proving 1 -> (D1,D2), proving 2 -> (D3,D4),
proving 3 -> (D5,D6).
  - C is always the fixed letter prefix of the value (e.g. "D")
  - F = the numeric part of the value
  - A = F + 1

NOTE: the exact relay contact-position convention (e.g. 8F-8B series
numbering) is still being finalized as part of RELAY RACK generation -
per instruction, this module proceeds now with the example logic above
and can be refined once that's confirmed, without needing to be rebuilt
from scratch.
"""

import ezdxf
from ezdxf import bbox as ezdxf_bbox
from ezdxf.addons import importer
import os
import re

from signal_core import (
    TEMPLATES_DIR,
    resolve_template_path,
    _safe_str,
    insert_border_title as _insert_border_title_generic,
)

# Measured directly from CONTACTS.dxf's Front_Contact block bounding box.
def _measure_contact_pair_width():
    doc = ezdxf.readfile(resolve_template_path("CONTACTS.dxf"))
    blk = doc.blocks["Front_Contact"]
    box = ezdxf_bbox.extents(blk)
    return box.extmax.x - box.extmin.x


CONTACT_PAIR_WIDTH = _measure_contact_pair_width()

TITLE_TEXT = "AR & AZR CIRCUITS - {name}"


def read_ar_azr_position_config(xlsx_path: str) -> dict:
    """
    Reads every position value from the 'AR_AZR_POSITIONS' sheet, so
    positions can be adjusted directly in Excel without touching code.
    Every element's spacing/position is independently controllable -
    nothing is shared or derived between elements.

    Expected sheet layout (header row + one row per named value):
        Parameter                  | X        | Y
        CONTACT_SPACING            | 23.1426  |
        FUSE_TO_PROVING_SPACING    | 23.1426  |
        CONTACT_PICK_POINT         | 115.6222 | 143.4991
        LAST_PROVING_PLACE_POINT   | 300.9720 | 255.0941
        COIL_PICK_POINT            | 164.0562 | 143.4991
        COIL_PLACE_POINT           | 339.6605 | 255.0941
        NEGFUSE_PICK_POINT         | 172.0300 | 143.4991
        NEGFUSE_PLACE_POINT        | 320.3911 | 255.0940
        POSFUSE_PICK_POINT         | 78.5799  | 143.4991
        AZR_SHIFT                  | 0.4143   | -50.2322

    Returns a dict: {name: (x, y)} for point values, {name: x} for
    single-value rows (like CONTACT_SPACING, which has no Y).
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "AR_AZR_POSITIONS" not in wb.sheetnames:
        raise ValueError(
            "Workbook has no 'AR_AZR_POSITIONS' sheet. Add one with columns "
            "Parameter/X/Y to control AR & AZR element positions - see "
            "ar_azr_core.py's read_ar_azr_position_config() docstring for the layout."
        )

    ws = wb["AR_AZR_POSITIONS"]
    config = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or str(row[0]).strip() == "":
            continue
        name = str(row[0]).strip()
        x_val = row[1]
        y_val = row[2] if len(row) > 2 else None
        if x_val is None:
            raise ValueError(f"AR_AZR_POSITIONS row {name!r} has no X value")
        if y_val is None or str(y_val).strip() == "":
            config[name] = float(x_val)
        else:
            config[name] = (float(x_val), float(y_val))

    required = [
        "CONTACT_SPACING", "FUSE_TO_PROVING_SPACING", "CONTACT_PICK_POINT", "LAST_PROVING_PLACE_POINT",
        "COIL_PICK_POINT", "COIL_PLACE_POINT", "NEGFUSE_PICK_POINT",
        "NEGFUSE_PLACE_POINT", "POSFUSE_PICK_POINT", "AZR_SHIFT",
    ]
    missing = [name for name in required if name not in config]
    if missing:
        raise ValueError(f"AR_AZR_POSITIONS sheet is missing required rows: {missing}")

    return config


def _parse_provings(provings_str: str, name2_whitelist: list = None):
    """
    '2 TSPR, 45 TSPR, 21 AR1' -> [('2','TSPR'), ('45','TSPR'), ('21','AR1')]
    CONFIRMED: tries the known NAME2 whitelist first (handles multi-word
    NAME2 values like 'DN SDF' correctly, which a plain last-space split
    can't). Falls back to the original "everything before the last
    whitespace-separated token is the name" behavior (supports
    multi-word NAME1 like 'SKL 8 TSPR' -> ('SKL 8','TSPR')) if nothing
    in the whitelist matches - e.g. genuinely unrecognized/new relay
    types not yet in RELAY_RACK_POSITIONS!E2.
    """
    if not provings_str or not str(provings_str).strip():
        return []
    entries = []
    for chunk in str(provings_str).split(","):
        chunk = chunk.strip()
        if not chunk:
            continue

        name = None
        relay_type = None
        if name2_whitelist:
            chunk_upper = chunk.upper()
            for candidate in name2_whitelist:
                cand_upper = candidate.strip().upper()
                if not cand_upper:
                    continue
                if chunk_upper == cand_upper:
                    name, relay_type = "", chunk
                    break
                suffix = " " + cand_upper
                if chunk_upper.endswith(suffix):
                    name = chunk[: -len(suffix)].strip()
                    relay_type = chunk[-len(candidate):].strip()
                    break

        if name is None:
            parts = chunk.rsplit(" ", 1)
            if len(parts) != 2:
                raise ValueError(
                    f"Could not parse PROVINGS entry {chunk!r} - expected '<name> <relay_type>'"
                )
            name, relay_type = parts[0].strip(), parts[1].strip()

        entries.append((name, relay_type))
    return entries


def _parse_contact_groups(h3_value: str):
    """
    CONFIRMED convention: groups are separated by '/', and each group's
    two comma-separated values are that proving's two contact sub-values.
    'D1,D2/D3,D4/D5,D6' -> [['D1','D2'], ['D3','D4'], ['D5','D6']]
    Each group must have exactly 2 values (one per contact in the pair).
    """
    if not h3_value or not str(h3_value).strip():
        raise ValueError("FIELD PG.NO!H3 is empty - needed for AR & AZR contact numbering")
    groups = []
    for chunk in str(h3_value).split("/"):
        chunk = chunk.strip()
        if not chunk:
            continue
        values = [v.strip() for v in chunk.split(",")]
        if len(values) != 2:
            raise ValueError(
                f"FIELD PG.NO!H3 group {chunk!r} has {len(values)} value(s), expected exactly 2 "
                f"(e.g. 'D1,D2'). Full H3 value: {h3_value!r}"
            )
        groups.append(values)
    return groups


def _afc_for_subvalue(subvalue: str):
    """'D1' -> (C='D', F='1', A='2'). A is always F+1."""
    m = re.match(r"^([A-Za-z]+)(\d+)$", subvalue)
    if not m:
        raise ValueError(f"Could not parse contact sub-value {subvalue!r} (expected e.g. 'D1')")
    letter, num = m.group(1), int(m.group(2))
    return letter, str(num), str(num + 1)


def _import_whole_file_no_rename(target_doc, source_filename: str, pick_point, place_point):
    """
    Like _import_whole_file, but avoids the block-name-renaming that
    happens when the same block definition gets re-imported multiple
    times (e.g. 'Front_Contact' -> 'Front_Contact0', 'Front_Contact1'...).

    For each INSERT in the source file: imports its block DEFINITION only
    ONCE per target document (checked via target_doc.blocks - safe, no
    aliasing risk), then creates a fresh blockref (not a re-import) at the
    correct translated position - preserving the exact original block
    name on every instance.

    For raw geometry (LINE, LWPOLYLINE, etc. not inside a block): copies
    and translates it fresh each call, since raw geometry has no renaming
    issue at all.

    Returns the list of newly-added entities in target_doc.
    """
    source_doc = ezdxf.readfile(resolve_template_path(source_filename))
    source_msp = source_doc.modelspace()
    target_msp = target_doc.modelspace()

    dx = place_point[0] - pick_point[0]
    dy = place_point[1] - pick_point[1]

    new_entities = []

    for e in source_msp:
        if e.dxftype() == "INSERT":
            block_name = e.dxf.name
            if block_name not in target_doc.blocks:
                imp = importer.Importer(source_doc, target_doc)
                imp.import_block(block_name)
                imp.finalize()

            new_x = e.dxf.insert.x + dx
            new_y = e.dxf.insert.y + dy
            ins = target_msp.add_blockref(block_name, (new_x, new_y))
            ins.dxf.xscale = e.dxf.xscale
            ins.dxf.yscale = e.dxf.yscale
            ins.dxf.zscale = e.dxf.zscale
            ins.dxf.rotation = e.dxf.rotation
            # CONFIRMED: use add_attrib per source attribute (not
            # add_auto_attribs) - some INSERT instances carry attributes
            # that aren't part of the block's own ATTDEF template (e.g.
            # Relay_Coil's S_NAME), and add_auto_attribs silently drops
            # anything not matching the template.
            for att in e.attribs:
                dxfattribs = {
                    "height": att.dxf.height,
                    "style": att.dxf.style,
                    "rotation": att.dxf.rotation,
                    "flags": att.dxf.flags,
                }
                att_insert = (att.dxf.insert.x + dx, att.dxf.insert.y + dy, att.dxf.insert.z)
                ins.add_attrib(att.dxf.tag, att.dxf.text, att_insert, dxfattribs)
            new_entities.append(ins)
        else:
            # Raw geometry (LINE, LWPOLYLINE, etc.) - copy fresh each time, no renaming risk.
            existing_ids = set(id(x) for x in target_msp)
            imp = importer.Importer(source_doc, target_doc)
            imp.import_entity(e, target_msp)
            imp.finalize()
            copied = [x for x in target_msp if id(x) not in existing_ids]
            for c in copied:
                # CONFIRMED FIX: a silent `except AttributeError: pass`
                # here meant that if .translate() ever failed for any
                # reason, the entity was left stranded at its RAW,
                # untranslated source coordinates with no error, no
                # warning, nothing - producing exactly this kind of
                # "unwanted line in the wrong place" bug with no trace
                # of why. Explicit manual fallback for LINE specifically
                # (the most common raw-geometry type in these
                # templates) instead of relying solely on .translate().
                if c.dxftype() == "LINE":
                    c.dxf.start = (c.dxf.start.x + dx, c.dxf.start.y + dy, c.dxf.start.z)
                    c.dxf.end = (c.dxf.end.x + dx, c.dxf.end.y + dy, c.dxf.end.z)
                else:
                    c.translate(dx, dy, 0)
                new_entities.append(c)

    return new_entities


def _import_whole_file(target_doc, source_filename: str, pick_point, place_point):
    """
    Copy the ENTIRE content of a source template file's modelspace (every
    entity - INSERTs, LINEs, LWPOLYLINEs, everything) into target_doc's
    modelspace as ONE atomic paste operation, then translate ALL of it
    together so pick_point lands exactly on place_point.

    This replaces looking up individual named entities one at a time,
    which risked subtle mismatches between separate import calls (e.g.
    a proving's contacts and its connecting line being imported/moved in
    separate steps). Copying and moving everything together guarantees
    each proving's full geometry stays correctly grouped.

    Returns the list of newly-added entities (so the caller can find and
    edit specific attributes on them, e.g. the INSERT entities' ATTRIBs).
    """
    source_doc = ezdxf.readfile(resolve_template_path(source_filename))
    target_msp = target_doc.modelspace()
    existing_ids = set(id(e) for e in target_msp)

    imp = importer.Importer(source_doc, target_doc)
    imp.import_modelspace()
    imp.finalize()

    new_entities = [e for e in target_msp if id(e) not in existing_ids]

    dx = place_point[0] - pick_point[0]
    dy = place_point[1] - pick_point[1]
    for e in new_entities:
        try:
            e.translate(dx, dy, 0)
        except AttributeError:
            pass

    return new_entities


def _inserts_named(entities, block_name_prefix):
    """Filter a list of entities to just the INSERTs whose block name
    starts with block_name_prefix (handles ezdxf's auto-renaming of
    duplicate block names, e.g. 'Front_Contact' -> 'Front_Contact0')."""
    return [e for e in entities if e.dxftype() == "INSERT" and e.dxf.name.startswith(block_name_prefix)]


def _move_to(insert_entity, target_point):
    """Translate insert_entity so its own insertion point lands exactly
    on target_point (x, y)."""
    dx = target_point[0] - insert_entity.dxf.insert.x
    dy = target_point[1] - insert_entity.dxf.insert.y
    insert_entity.translate(dx, dy, 0)


def _draw_ar_azr_circuit(doc, row: dict, groups: list, positions: dict, offset=(0.0, 0.0), registry=None, sht_number=None, name2_whitelist=None):
    """
    Draws one AR/AZR row's full circuit (provings, both fuses, coil) into
    an existing doc, with every position shifted by `offset`. Used both
    for a standalone sheet (offset=(0,0)) and for a second row combined
    onto the same sheet as a paired row (offset=positions['AZR_SHIFT']).

    positions: dict from read_ar_azr_position_config() - all coordinates
    come from here rather than hardcoded constants, so they can be
    adjusted directly in Excel (AR_AZR_POSITIONS sheet) without touching code.
    """
    own_name = row.get("AR/AZR")
    direction = row.get("DIRECTION") or ""
    relay_type = row.get("TYPE") or ""
    provings = _parse_provings(row.get("PROVINGS"), name2_whitelist)

    msp = doc.modelspace()
    ox, oy = offset

    contact_spacing = positions["CONTACT_SPACING"]
    fuse_to_proving_spacing = positions["FUSE_TO_PROVING_SPACING"]
    contact_pick_point = positions["CONTACT_PICK_POINT"]
    last_proving_place_point = positions["LAST_PROVING_PLACE_POINT"]
    coil_pick_point = positions["COIL_PICK_POINT"]
    coil_place_point = positions["COIL_PLACE_POINT"]
    negfuse_pick_point = positions["NEGFUSE_PICK_POINT"]
    negfuse_place_point = positions["NEGFUSE_PLACE_POINT"]
    posfuse_pick_point = positions["POSFUSE_PICK_POINT"]

    def _apply_proving(pasted_entities, proving_index, name, r_type):
        # CONFIRMED: groups[0] is the "top group", groups[1] is the "bottom
        # group" - every proving (regardless of index) uses the FIRST value
        # from each: top always groups[0][0] (e.g. D1), bottom always
        # groups[1][0] (e.g. D3). Fixed for every proving (temporary
        # simplification - proper distinct-contact restriction planned later).
        if len(groups) < 2:
            raise ValueError(
                f"FIELD PG.NO!H3 needs at least 2 groups (top group and bottom group), "
                f"got {len(groups)}: {groups!r}"
            )
        top_value = groups[0][0]
        bottom_value = groups[1][0]
        contacts = _inserts_named(pasted_entities, "Front_Contact")
        pair_sorted = sorted(contacts, key=lambda e: -e.dxf.insert.y)
        subvalues = [top_value, bottom_value]

        for e, subval in zip(pair_sorted, subvalues):
            letter, f_val, a_val = _afc_for_subvalue(subval)
            for att in e.attribs:
                if att.dxf.tag == "S_NAME":
                    att.dxf.text = _safe_str(name)
                elif att.dxf.tag == "R_NAME":
                    att.dxf.text = _safe_str(r_type)
                elif att.dxf.tag == "C":
                    att.dxf.text = letter
                elif att.dxf.tag == "F":
                    att.dxf.text = f_val
                elif att.dxf.tag == "A":
                    att.dxf.text = a_val
            if registry is not None:
                registry.register_pair(name, letter, f_val, a_val, "AR & AZR", sht_number)

    num_provings = max(len(provings), 1)

    last_place_x, place_y = last_proving_place_point
    last_place_x += ox
    place_y += oy

    for i in range(len(provings)):
        steps_from_last = (len(provings) - 1) - i
        x = last_place_x - steps_from_last * contact_spacing
        pasted = _import_whole_file_no_rename(doc, "CONTACTS.dxf", contact_pick_point, (x, place_y))
        _apply_proving(pasted, i, provings[i][0], provings[i][1])

    base_x = last_place_x - (num_provings - 1) * contact_spacing

    fuse_target_x = base_x - fuse_to_proving_spacing
    pf_pasted = _import_whole_file_no_rename(doc, "POSITIVE_FUSE.dxf", posfuse_pick_point, (fuse_target_x, place_y))
    for e in _inserts_named(pf_pasted, "FUZEBLOCK"):
        for att in e.attribs:
            if att.dxf.tag == "VOLT":
                att.dxf.text = att.dxf.text.replace("#", _safe_str(direction))

    coil_place = (coil_place_point[0] + ox, coil_place_point[1] + oy)
    coil_pasted = _import_whole_file_no_rename(doc, "RELAY_COIL.dxf", coil_pick_point, coil_place)
    for e in _inserts_named(coil_pasted, "Relay_Coil"):
        for att in e.attribs:
            if att.dxf.tag == "S_NAME":
                att.dxf.text = _safe_str(own_name)
            elif att.dxf.tag == "R_NAME":
                att.dxf.text = _safe_str(relay_type)
        # CONFIRMED FIX: the coil terminal itself (R1/R2) was never
        # registered with the ContactRegistry before - only the front
        # contact provings were, meaning the same relay's coil being
        # wired twice would pass through undetected. Register both
        # coil terminals, same critical-error/halt behavior as contacts.
        # CONFIRMED FIX: the coil's real identity is (own_name,
        # relay_type) together, not own_name alone - an AR relay and a
        # separately-typed AZR relay can legitimately share the same
        # bare number (e.g. both named "32"), since AR/AZR convention
        # often numbers pairs by shared track/signal number. Registering
        # by own_name alone falsely flagged these as the same relay.
        if registry is not None:
            registry.register_pair(f"{own_name} {relay_type}", "R", 1, 2, "AR & AZR", sht_number)

    negfuse_place = (negfuse_place_point[0] + ox, negfuse_place_point[1] + oy)
    nf_pasted = _import_whole_file_no_rename(doc, "NEGATIVE_FUSE.dxf", negfuse_pick_point, negfuse_place)
    for e in _inserts_named(nf_pasted, "FUZEENDBLOCK"):
        for att in e.attribs:
            if att.dxf.tag == "VOLT":
                att.dxf.text = att.dxf.text.replace("#", _safe_str(direction))


def generate_ar_azr_sheet(
    row: dict,
    contact_groups_h3: str,
    contact_groups_g3: str,
    sht_number: str,
    cont_number: str,
    border_template_path: str,
    positions: dict,
    second_row: dict = None,
    registry=None,
    name2_whitelist: list = None,
):
    """
    row: dict with keys 'AR/AZR' (own name), DIRECTION, TYPE (AR/AZR), PROVINGS
    contact_groups_h3: raw FIELD PG.NO!H3 value, e.g. 'D1,D3' - used for AR-type provings.
    contact_groups_g3: raw FIELD PG.NO!G3 value - used for AZR-type provings
        (CONFIRMED: separate from H3 so AZR's contacts don't repeat AR's).
    positions: dict from read_ar_azr_position_config(xlsx_path)
    second_row: optional second row (matching DIRECTION) drawn on the SAME
        sheet, shifted by positions['AZR_SHIFT'].

    Returns (output_filename, ezdxf.Drawing).
    """
    groups_h3 = _parse_contact_groups(contact_groups_h3)
    if not groups_h3:
        raise ValueError(f"FIELD PG.NO!H3 defines no usable contact groups ({contact_groups_h3!r})")
    groups_g3 = _parse_contact_groups(contact_groups_g3) if contact_groups_g3 else []

    # Base doc starts from POSITIVE_FUSE.dxf - just a convenient starting
    # container; its own original (unplaced) FUZEBLOCK gets removed once
    # the real, correctly-positioned one is pasted in.
    doc = ezdxf.readfile(resolve_template_path("POSITIVE_FUSE.dxf"))
    msp = doc.modelspace()
    pf_ins = next(e for e in msp if e.dxftype() == "INSERT" and e.dxf.name == "FUZEBLOCK")
    # CONFIRMED FIX: POSITIVE_FUSE.dxf also has a LINE baked in (the
    # fuse's own connecting lead wire) alongside its FUZEBLOCK - the
    # cleanup below already correctly removed the placeholder FUZEBLOCK,
    # but never removed this LINE, leaving it stranded at its raw,
    # untranslated source coordinates in every generated sheet (an
    # "unwanted line" with no connection to anything, since the real,
    # correctly-positioned fuse+lead gets pasted in fresh via
    # _import_whole_file_no_rename elsewhere - this base-document copy
    # was only ever meant to provide a valid starting container with
    # the right layers/styles, not to keep its own drawing content).
    pf_line = next((e for e in msp if e.dxftype() == "LINE"), None)

    def _groups_for_row(r):
        r_type = (r.get("TYPE") or "").strip().upper()
        if r_type == "AZR":
            if not groups_g3:
                raise ValueError("FIELD PG.NO!G3 defines no usable contact groups - needed for AZR-type provings")
            return groups_g3
        return groups_h3

    _draw_ar_azr_circuit(
        doc, row, _groups_for_row(row), positions, offset=(0.0, 0.0), registry=registry,
        sht_number=sht_number, name2_whitelist=name2_whitelist,
    )
    if second_row is not None:
        _draw_ar_azr_circuit(
            doc, second_row, _groups_for_row(second_row), positions,
            offset=positions["AZR_SHIFT"], registry=registry, sht_number=sht_number,
            name2_whitelist=name2_whitelist,
        )

    if pf_ins.is_alive:
        msp.delete_entity(pf_ins)
    if pf_line is not None and pf_line.is_alive:
        msp.delete_entity(pf_line)

    own_name = row.get("AR/AZR")
    relay_type = row.get("TYPE") or ""
    title_names = [_safe_str(own_name)]
    if second_row is not None:
        title_names.append(_safe_str(second_row.get("AR/AZR")))
    title_text = TITLE_TEXT.format(name=" & ".join(title_names))

    _insert_border_title_generic(
        doc=doc,
        border_template_path=border_template_path,
        template_name="AR_AZR",
        sig_name=None,
        ahead_sig=None,
        loc=None,
        sht_number=sht_number,
        cont_number=cont_number,
        title_text_override=title_text,
    )

    out_name = f"{own_name}_{relay_type}_SHT{sht_number}.dxf"
    return out_name, doc


def generate_all_ar_azr_sheets(
    rows: list, contact_groups_h3: str, contact_groups_g3: str, start_sheet_number: int,
    border_template_path: str, positions: dict, registry=None, name2_whitelist: list = None,
):
    """
    rows: list of dicts with 'AR/AZR', DIRECTION, TYPE, PROVINGS
    positions: dict from read_ar_azr_position_config(xlsx_path)

    CONFIRMED pairing rule: consecutive rows with the SAME DIRECTION value
    combine onto ONE sheet (AZR positioned per positions['AZR_SHIFT']). A
    row whose DIRECTION doesn't match the next one gets its own standalone sheet.

    Returns (list of (output_filename, ezdxf.Drawing, sht_number_str), next_sheet_number)
    """
    # Build the pairing plan first: list of (row, second_row_or_None, first_row_index)
    plan = []
    i = 0
    while i < len(rows):
        row = rows[i]
        if i + 1 < len(rows) and (row.get("DIRECTION") or "") == (rows[i + 1].get("DIRECTION") or ""):
            plan.append((row, rows[i + 1], i))
            i += 2
        else:
            plan.append((row, None, i))
            i += 1

    results = []
    sheet_num = start_sheet_number
    for row, second_row, row_index in plan:
        sht = f"{sheet_num:03d}"
        cont = f"{sheet_num + 1:03d}"
        try:
            out_name, doc = generate_ar_azr_sheet(
                row, contact_groups_h3, contact_groups_g3, sht, cont, border_template_path, positions,
                second_row=second_row, registry=registry, name2_whitelist=name2_whitelist,
            )
        except Exception as e:
            raise ValueError(f"Row {row_index + 2} (AR/AZR={row.get('AR/AZR')}): {e}") from e
        results.append((out_name, doc, sht))
        sheet_num += 1
    return results, sheet_num


def get_ar_azr_start_sheet_number(xlsx_path: str) -> int:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(row=r, column=2).value or "").strip().upper()
        if "AR" in label and "AZR" in label:
            value = ws.cell(row=r, column=3).value
            if value is None or str(value).strip() == "":
                raise ValueError("FIELD PG.NO has no Sheet Number for the AR & AZR row")
            return int(str(value).strip())
    raise ValueError("Could not find an 'AR & AZR' row in FIELD PG.NO")


def get_next_circuit_after_ar_azr(xlsx_path: str) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(row=r, column=2).value or "").strip().upper()
        if "AR" in label and "AZR" in label:
            value = ws.cell(row=r + 1, column=3).value
            if value is None or str(value).strip() == "":
                raise ValueError("FIELD PG.NO has no Sheet Number for the row after AR & AZR")
            v = str(value).strip()
            return v.zfill(3) if v.isdigit() else v
    raise ValueError("Could not find an 'AR & AZR' row in FIELD PG.NO")


def get_ar_azr_contact_groups_h3(xlsx_path: str) -> str:
    """Read FIELD PG.NO!H3 - the contact-numbering groups string, e.g. 'D1/D2,D3,D4'."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    value = ws["H3"].value
    if value is None or str(value).strip() == "":
        raise ValueError("FIELD PG.NO!H3 is empty - needed for AR & AZR contact numbering")
    return str(value).strip()


def get_ar_azr_contact_groups_g3(xlsx_path: str) -> str:
    """Read FIELD PG.NO!G3 - the contact-numbering groups string for AZR-type provings specifically
    (CONFIRMED: separate from H3, so AZR's contacts don't repeat AR's)."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    value = ws["G3"].value
    return str(value).strip() if value is not None and str(value).strip() != "" else ""
