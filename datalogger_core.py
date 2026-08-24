"""
MSDAC Data Logger Circuit Generator - core assembly logic.

Derives from the RELAY RACK sheet's relay listing (reuses
relay_rack_core.read_all_relay_racks) - one contact per relay entry
(skipping SPARE cells), 3 columns x 12 rows per sheet (36 relays/sheet).

CONFIRMED layout:
  - DL_BACK_CONTACTS.dxf / DL_FRONT_CONTACTS.dxf: single-block templates,
    whole-file-copy + translate per relay (matching the established pattern).
  - Column start X positions (all at Y=269.46): col1=64.2, col2=174.2744,
    col3=297.5974.
  - Row spacing within a column: 17.6737 (downward).
  - 12 rows per column, 3 columns per sheet = 36 relays/sheet.
  - DL_FRONT_CONTACTS.dxf's own pick point is (0,0) (its native position).

CONFIRMED contact selection: reads FIELD PG.NO!I3 as an override (e.g.
'D7,D8') applied to every relay. When I3 is genuinely BLANK, each relay
gets its own free-contact search instead of a fixed default: BACK bank
tried first (highest pair first, e.g. 8F-8B tries C7,C8 before C5,C6),
then FRONT bank, where "free" means not already used by any other
already-generated circuit (Signal, Track, AR & AZR, Communication, SDF)
for that SAME relay - see build_used_contacts_registry(). Only
meaningful within the Full IFC job; a standalone Data Logger run has no
other circuits to check against, so it always gets the top-priority
back pair. Falls back to the old fixed default (8F-8B/8T-8B -> C7,C8;
4F-4B/4T-4B -> B3,B4) if nothing is free.
"""

import ezdxf
from ezdxf.addons import importer
import os

from signal_core import (
    TEMPLATES_DIR,
    resolve_template_path,
    _safe_str,
    insert_border_title as _insert_border_title_generic,
)
from relay_rack_core import read_all_relay_racks

COLUMN_X_POSITIONS = [64.2, 174.2744, 297.5974]
ROW_Y_START = 269.4600
ROW_SPACING = 17.6737
ROWS_PER_COLUMN = 12
RELAYS_PER_SHEET = len(COLUMN_X_POSITIONS) * ROWS_PER_COLUMN  # 36

BACK_CONTACT_PICK_POINT = (64.2, 269.46)
FRONT_CONTACT_PICK_POINT = (0.0, 0.0)


def _max_position_for_cntcnf(cntcnf: str) -> int:
    """8F-8B/8T-8B relays have positions 1-8; 4F-4B/4T-4B relays only have 1-4."""
    cntcnf = (cntcnf or "").upper()
    if "8F" in cntcnf or "8T" in cntcnf:
        return 8
    elif "4F" in cntcnf or "4T" in cntcnf:
        return 4
    raise ValueError(f"Unrecognized relay configuration {cntcnf!r} - expected 8F-8B/8T-8B or 4F-4B/4T-4B style")


def _parse_i3_override(i3_value: str):
    """'D7,D8' -> (letter='D', num1=7, num2=8), or None if blank."""
    if not i3_value or not str(i3_value).strip():
        return None
    parts = str(i3_value).strip().split(",")
    if len(parts) != 2:
        raise ValueError(f"FIELD PG.NO!I3 must be 'LETTER#,LETTER#' (e.g. 'D7,D8'), got {i3_value!r}")
    import re
    m1 = re.match(r"^([A-Za-z])(\d+)$", parts[0].strip())
    m2 = re.match(r"^([A-Za-z])(\d+)$", parts[1].strip())
    if not m1 or not m2:
        raise ValueError(f"FIELD PG.NO!I3 must be 'LETTER#,LETTER#' (e.g. 'D7,D8'), got {i3_value!r}")
    return m1.group(1).upper(), int(m1.group(2)), int(m2.group(2))


_FIXED_WIRING_DOC_CACHE = {}


def _read_fixed_wiring_doc(source_file: str, block_name: str):
    """
    CONFIRMED: DL_FRONT_CONTACTS.dxf/DL_BACK_CONTACTS.dxf's OWN reference
    instance already carries the correct fixed WIRING/DOC values (e.g.
    Front_Contact: WIRING='21'/DOC='RF'; Back_Contact: WIRING='21'/
    DOC='LB') - the same value applies to every placement regardless of
    position, since Data Logger's contacts are all the same physical
    connection type. The gap was that _import_block_definition_once()
    only imports the block DEFINITION (blank ATTDEF defaults) and a
    fresh add_blockref() doesn't inherit the template file's own
    instance values - so they have to be read and applied explicitly.
    Cached per (file, block) since this doesn't change across a run.
    """
    cache_key = (source_file, block_name)
    if cache_key in _FIXED_WIRING_DOC_CACHE:
        return _FIXED_WIRING_DOC_CACHE[cache_key]
    wiring, doc_value = "", ""
    try:
        source_doc = ezdxf.readfile(resolve_template_path(source_file))
        for e in source_doc.modelspace():
            if e.dxftype() == "INSERT" and e.dxf.name == block_name:
                atts = {a.dxf.tag: a.dxf.text for a in e.attribs}
                wiring = atts.get("WIRING", "")
                doc_value = atts.get("DOC", "")
                break
    except Exception:
        pass
    _FIXED_WIRING_DOC_CACHE[cache_key] = (wiring, doc_value)
    return wiring, doc_value


def _contact_for_relay(cntcnf: str, i3_override, name1: str = None, name2: str = None, used_contacts: dict = None, log=None):
    """
    CONFIRMED: if I3 has a value AND it's a valid position for this relay's
    own configuration (position number within that config's valid range),
    use it - correctly determining Front vs Back based on which starting
    position it matches (per CONTACTS_CONFIGURATION.xlsx):
      8F-8B: FRONT starts at 1 or 3 (pairs 1-2, 3-4); BACK starts at 5 or 7 (pairs 5-6, 7-8).
      4F-4B: FRONT starts at 1 (pair 1-2); BACK starts at 3 (pair 3-4).
    Otherwise falls back to the config-based default.

    CONFIRMED (new): when I3 is genuinely BLANK (not just an invalid
    value), don't use the old fixed default - instead search for a
    contact pair that's actually FREE (not already used by any other
    already-generated circuit for this SAME relay), trying the BACK
    bank first (highest pair first), then the FRONT bank (highest pair
    first). Falls back to the old fixed default only if nothing is free.
    Returns (is_front: bool, val1, val2, letter).
    """
    if i3_override is not None:
        letter, num1, num2 = i3_override
        cntcnf_upper = (cntcnf or "").upper()
        if "8F" in cntcnf_upper or "8T" in cntcnf_upper:
            front_starts, back_starts = (1, 3), (5, 7)
        elif "4F" in cntcnf_upper or "4T" in cntcnf_upper:
            front_starts, back_starts = (1,), (3,)
        else:
            raise ValueError(f"Unrecognized relay configuration {cntcnf!r}")

        if num1 in front_starts:
            return True, str(num1), str(num2), letter
        elif num1 in back_starts:
            return False, str(num1), str(num2), letter
        # num1 doesn't match a valid starting position for this config - fall through to default.
        return _default_contact_for_cntcnf(cntcnf)

    if used_contacts is not None and name1 and name2:
        return _free_priority_contact_for_cntcnf(cntcnf, name1, name2, used_contacts, log)
    return _default_contact_for_cntcnf(cntcnf)


def build_used_contacts_registry(docs: list) -> dict:
    """
    docs: ezdxf.Drawing objects already generated earlier in THIS SAME
    run (Signal, Track, AR & AZR, Communication, SDF - Relay Rack has no
    contact blocks to find, Data Logger/Custom Circuits haven't run yet).
    Returns {(s_name_upper, r_name_upper): set of "LETTERNUM" strings}
    e.g. {"C7", "C8"} - CONFIRMED this is what "free" is checked against:
    not already used by any OTHER already-generated circuit for the SAME
    relay.
    """
    used = {}
    for doc in docs:
        for e in doc.modelspace():
            if e.dxftype() != "INSERT":
                continue
            if not (e.dxf.name.startswith("Front_Contact") or e.dxf.name.startswith("Back_Contact")
                    or e.dxf.name.startswith("SDF_CONTACT") or e.dxf.name.startswith("LCPR_FRONT")
                    or e.dxf.name.startswith("LCPR_BACK CONTACT")):
                continue
            att = {a.dxf.tag: (a.dxf.text or "").strip() for a in e.attribs}
            r_name = att.get("R_NAME", "")
            s_name = att.get("S_NAME", "")
            letter = att.get("C", "")
            f_val = att.get("F", "")
            a_val = att.get("A", "")
            if not (s_name and r_name and letter):
                continue
            key = (s_name.upper(), r_name.upper())
            used.setdefault(key, set())
            if f_val:
                used[key].add(f"{letter}{f_val}")
            if a_val:
                used[key].add(f"{letter}{a_val}")
    return used


def _free_priority_contact_for_cntcnf(cntcnf: str, name1: str, name2: str, used_contacts: dict, log=None):
    """
    CONFIRMED priority order: BACK bank first (highest pair first, e.g.
    8F-8B tries 7-8 before 5-6), then FRONT bank (ASSUMED same
    highest-pair-first convention, not yet explicitly confirmed). Letter
    stays fixed per config (same as the old default: "C" for 8F-8B/8T-8B,
    "B" for 4F-4B/4T-4B) - only which NUMBER pair is picked varies.
    Falls back to the old fixed default (with a warning) if nothing is
    free at all.
    """
    cntcnf_upper = (cntcnf or "").upper()
    if "8F" in cntcnf_upper or "8T" in cntcnf_upper:
        back_pairs = [(7, 8), (5, 6)]
        front_pairs = [(3, 4), (1, 2)]
        letter = "C"
    elif "4F" in cntcnf_upper or "4T" in cntcnf_upper:
        back_pairs = [(3, 4)]
        front_pairs = [(1, 2)]
        letter = "B"
    else:
        raise ValueError(f"Unrecognized relay configuration {cntcnf!r} - expected 8F-8B/8T-8B or 4F-4B/4T-4B style")

    key = ((name1 or "").upper(), (name2 or "").upper())
    used_for_relay = used_contacts.get(key, set())

    for num1, num2 in back_pairs:
        if f"{letter}{num1}" not in used_for_relay and f"{letter}{num2}" not in used_for_relay:
            return False, str(num1), str(num2), letter

    for num1, num2 in front_pairs:
        if f"{letter}{num1}" not in used_for_relay and f"{letter}{num2}" not in used_for_relay:
            return True, str(num1), str(num2), letter

    if log is not None:
        log.warning(f"[DATA LOGGER] No free contact: {name1} {name2} - using default.")
    return _default_contact_for_cntcnf(cntcnf)


def _default_contact_for_cntcnf(cntcnf: str):
    """
    CONFIRMED fixed defaults per relay configuration:
      8F-8B (or 8T-8B) -> Back_Contact, A=7,B=8,C=C
      4F-4B (or 4T-4B) -> Back_Contact, A=3,B=4,C=B
    Returns (is_front: bool, A, B_or_F, C).
    """
    cntcnf = (cntcnf or "").upper()
    if "8F" in cntcnf or "8T" in cntcnf:
        return False, "7", "8", "C"
    elif "4F" in cntcnf or "4T" in cntcnf:
        return False, "3", "4", "B"
    else:
        raise ValueError(f"Unrecognized relay configuration {cntcnf!r} - expected 8F-8B/8T-8B or 4F-4B/4T-4B style")


def _flatten_relay_entries(racks):
    """
    Flattens all racks' non-SPARE cells into one sequential list of
    dicts: {S_NAME, R_NAME, CNTCNF}.
    """
    entries = []
    for rack_num, grid_rows in racks:
        for row in grid_rows:
            for cell in row:
                if cell["NAME3"] == "SPARE":
                    continue
                entries.append({
                    "S_NAME": cell["NAME1"],
                    "R_NAME": cell["NAME2"],
                    "CNTCNF": cell["CNTCNF"],
                    "RELTYP": cell.get("RELTYP", ""),
                })
    return entries


def _import_whole_file(doc, source_filename: str, pick_point, place_point):
    """Copy the ENTIRE source file's modelspace content into doc, translated."""
    source_doc = ezdxf.readfile(resolve_template_path(source_filename))
    target_msp = doc.modelspace()
    existing_ids = set(id(e) for e in target_msp)

    imp = importer.Importer(source_doc, doc)
    imp.import_modelspace()
    imp.finalize()

    new_entities = [e for e in target_msp if id(e) not in existing_ids]
    dx = place_point[0] - pick_point[0]
    dy = place_point[1] - pick_point[1]
    for e in new_entities:
        try:
            e.translate(dx, dy, 0)
        except AttributeError:
            pass
    return new_entities


_transform_cache = {}


def _get_original_transform(source_filename: str, block_name: str):
    """Reads the original xscale/yscale/zscale/rotation for a block's INSERT in its source file (cached)."""
    cache_key = (source_filename, block_name)
    if cache_key not in _transform_cache:
        source_doc = ezdxf.readfile(resolve_template_path(source_filename))
        for e in source_doc.modelspace():
            if e.dxftype() == "INSERT" and e.dxf.name == block_name:
                _transform_cache[cache_key] = (e.dxf.xscale, e.dxf.yscale, e.dxf.zscale, e.dxf.rotation)
                break
        else:
            _transform_cache[cache_key] = (1.0, 1.0, 1.0, 0.0)
    return _transform_cache[cache_key]


def _apply_transform(ins, source_filename: str, block_name: str):
    """Applies the original scale/rotation to a newly-created blockref, so it renders identically to the source."""
    xscale, yscale, zscale, rotation = _get_original_transform(source_filename, block_name)
    ins.dxf.xscale = xscale
    ins.dxf.yscale = yscale
    ins.dxf.zscale = zscale
    ins.dxf.rotation = rotation


def _import_block_definition_once(doc, source_filename: str, block_name: str):
    """Import just ONE named block's DEFINITION (not the whole modelspace) - avoids duplicating unrelated content."""
    if block_name in doc.blocks:
        return
    source_doc = ezdxf.readfile(resolve_template_path(source_filename))
    imp = importer.Importer(source_doc, doc)
    imp.import_block(block_name)
    imp.finalize()


def _sorted_inc_instances(msp):
    """
    CONFIRMED: the base datalogger.dxf template already has 36 'INC'
    block instances pre-placed, one per relay slot, in the SAME 3-column
    x 12-row grid as the relay entries themselves (COLUMN_X_POSITIONS/
    ROW_Y_START/ROW_SPACING) - just at different absolute coordinates
    since they're a separate fixed grid baked into the template. Sorts
    them into the SAME col-major, top-to-bottom order used for relay
    entries (col = i // ROWS_PER_COLUMN, row = i % ROWS_PER_COLUMN), so
    the i-th relay entry can be matched to the i-th INC instance
    positionally, without needing their absolute coordinates to match.
    """
    instances = [e for e in msp if e.dxftype() == "INSERT" and e.dxf.name == "INC"]
    # Group by column (rounded X, clustered) then sort each column top-to-bottom (descending Y).
    instances.sort(key=lambda e: (round(e.dxf.insert.x / 50) * 50, -e.dxf.insert.y))
    return instances


def _letter_for(n):
    """0->'A', 1->'B', ..., 25->'Z', 26->'A1', 27->'B1', ..."""
    cycle, remainder = divmod(n, 26)
    letter = chr(ord("A") + remainder)
    return f"{letter}{cycle}" if cycle > 0 else letter


def generate_one_datalogger_sheet(
    entries: list, sht_number: str, cont_number: str, border_template_path: str,
    i3_override=None, is_last_sheet=True, letter_start=0, next_sht=None, is_first_sheet=True,
    previous_sht=None, sheet_index_number=1, boundary_letter_in=None,
    used_contacts=None, log=None, start_inc_number=1,
):
    """
    Builds ONE Data Logger sheet covering up to 36 relay entries.
    Returns (output_filename, ezdxf.Drawing).
    """
    if not entries:
        raise ValueError("entries is empty")

    # Base doc: datalogger.dxf provides the FUZEBLOCK/RRT grid framework,
    # unchanged position (pick=place=(0,0)).
    doc = ezdxf.readfile(resolve_template_path("datalogger.dxf"))
    msp = doc.modelspace()

    # Remove the sample Back_Contact/Front_Contact instances that came
    # with the reference file - they'll be replaced with our own relay data.
    # Also remove the baked-in FUZEBLOCK - CONFIRMED: it's only re-inserted
    # (via DL_1ST_SHEETS_FUSE.dxf) on the FIRST sheet, not every sheet.
    for e in list(msp):
        if e.dxftype() == "INSERT" and e.dxf.name in ("Back_Contact", "Front_Contact", "FUZEBLOCK"):
            msp.delete_entity(e)

    if is_first_sheet:
        _import_block_definition_once(doc, "DL_1ST_SHEETS_FUSE.dxf", "FUZEBLOCK")
        ins = msp.add_blockref("FUZEBLOCK", (28.19, 269.46))
        _apply_transform(ins, "DL_1ST_SHEETS_FUSE.dxf", "FUZEBLOCK")
        ins.add_auto_attribs({"VOLT": "N24", "FTOT": "XXX", "FUSEVALUE": ""})

    _import_block_definition_once(doc, "DL_FRONT_CONTACTS.dxf", "Front_Contact")
    _import_block_definition_once(doc, "DL_BACK_CONTACTS.dxf", "Back_Contact")

    # CONFIRMED: INC block's own "INC" attribute (tag == block name) gets
    # a simple incrementing count (1, 2, 3...) that continues across
    # every Data Logger sheet in the run, never resetting - matched to
    # each relay entry by position (i-th entry -> i-th pre-placed INC
    # instance), not computed/placed fresh.
    inc_instances = _sorted_inc_instances(msp)
    inc_counter = start_inc_number

    for i, entry in enumerate(entries):
        col = i // ROWS_PER_COLUMN
        row = i % ROWS_PER_COLUMN
        x = COLUMN_X_POSITIONS[col]
        y = ROW_Y_START - row * ROW_SPACING

        if i < len(inc_instances):
            inc_instances[i].add_auto_attribs({"INC": str(inc_counter)})
        inc_counter += 1

        is_spare_entry = (entry.get("R_NAME") or "").strip().upper() == "SPARE"
        if is_spare_entry:
            # CONFIRMED FIX: SPARE entries (Data Logger's own visual
            # padding for the last sheet's unused slots) previously
            # still fell through to _contact_for_relay's fixed-default
            # fallback (since S_NAME is blank, the free-search branch
            # never ran) - drawing a real contact position on a
            # placeholder that has no relay behind it. A spare doesn't
            # need any contact allocated at all, so skip the call and
            # leave the position fields blank. CONFIRMED (from Krish):
            # place a Back_Contact block for spares, not Front_Contact.
            is_front, val1, val2, letter = False, "", "", ""
        else:
            is_front, val1, val2, letter = _contact_for_relay(
                entry["CNTCNF"], i3_override, entry["S_NAME"], entry["R_NAME"], used_contacts, log,
            )
        block_name = "Front_Contact" if is_front else "Back_Contact"

        ins = msp.add_blockref(block_name, (x, y))
        source_file = "DL_FRONT_CONTACTS.dxf" if is_front else "DL_BACK_CONTACTS.dxf"
        _apply_transform(ins, source_file, block_name)
        wiring_val, doc_val = _read_fixed_wiring_doc(source_file, block_name)
        attribs = {
            "S_NAME": _safe_str(entry["S_NAME"]),
            "R_NAME": _safe_str(entry["R_NAME"]),
            "C": letter,
            "F" if is_front else "A": val1,
            "A" if is_front else "B": val2,
            "WIRING": wiring_val,
            "DOC": doc_val,
        }
        ins.add_auto_attribs(attribs)

    # RRT markers: matched by POSITION (not tag - it's genuinely blank in
    # this template). One RRT per relay slot; its FIRST attribute (default
    # 'X') gets replaced with an incrementing number (1-36 per sheet).
    rrt_instances = [e for e in msp if e.dxftype() == "INSERT" and e.dxf.name == "RRT"]
    for i, entry in enumerate(entries):
        col = i // ROWS_PER_COLUMN
        row = i % ROWS_PER_COLUMN
        expected_y = ROW_Y_START - row * ROW_SPACING
        # Match the RRT closest to this relay's row/column (within tolerance).
        candidates = [
            e for e in rrt_instances
            if abs(e.dxf.insert.y - expected_y) < 1.0
            and abs(e.dxf.insert.x - (COLUMN_X_POSITIONS[col] + 58)) < 40
        ]
        if candidates:
            rrt = candidates[0]
            if rrt.attribs:
                rrt.attribs[0].dxf.text = str(i + 1)

    # Continuation blocks - CONFIRMED: global letter counter starting at
    # 'A', incrementing per entity added, shared across ALL these block types.
    letter_counter = letter_start

    # DL_INSHEET_CONT/DL_INSHEET_CONT1 are now BAKED IN to the base
    # template (no longer dynamically inserted) - find and update the
    # EXISTING instances, sorted by X so "1st"/"2nd" is unambiguous.
    # CONFIRMED: 1st entities of both blocks share one letter; 2nd
    # entities of both blocks share the next letter.
    inshcont = sorted(
        (e for e in msp if e.dxftype() == "INSERT" and e.dxf.name == "DL_INSHEET_CONT"),
        key=lambda e: e.dxf.insert.x,
    )
    inshcont1 = sorted(
        (e for e in msp if e.dxftype() == "INSERT" and e.dxf.name == "DL_INSHEET_CONT1"),
        key=lambda e: e.dxf.insert.x,
    )

    letter_pos1 = _letter_for(letter_counter)
    letter_counter += 1
    letter_pos2 = _letter_for(letter_counter)
    letter_counter += 1

    for e, letter in zip(inshcont, (letter_pos1, letter_pos2)):
        for att in e.attribs:
            if att.dxf.tag == "CONT":
                att.dxf.text = letter
    for e, letter in zip(inshcont1, (letter_pos1, letter_pos2)):
        for att in e.attribs:
            if att.dxf.tag == "CONT":
                att.dxf.text = letter

    # PREVIOUS/NEXT-sheet continuation markers - CONFIRMED: a sheet's
    # "next" marker and the FOLLOWING sheet's "previous" marker share the
    # SAME letter (one letter per sheet-boundary, not one per sheet).
    if not is_first_sheet:
        _import_block_definition_once(doc, "DL_PREVIOUS_SHEETS_CONT.dxf", "CONT' PREVIOS SHEET")
        ins = msp.add_blockref("CONT' PREVIOS SHEET", (0, 0))
        _apply_transform(ins, "DL_PREVIOUS_SHEETS_CONT.dxf", "CONT' PREVIOS SHEET")
        ins.add_auto_attribs({"CONT": boundary_letter_in, "CONDSHEET": _safe_str(previous_sht) if previous_sht else ""})

    next_boundary_letter = None
    if not is_last_sheet:
        _import_block_definition_once(doc, "DL_NEXT_SHEETS_CONT.dxf", "DL_NXTSHEET_CONT")
        next_boundary_letter = _letter_for(letter_counter)
        letter_counter += 1
        ins = msp.add_blockref("DL_NXTSHEET_CONT", (0, 0))
        _apply_transform(ins, "DL_NEXT_SHEETS_CONT.dxf", "DL_NXTSHEET_CONT")
        ins.add_auto_attribs({"CONT": next_boundary_letter, "CONDSHEET": _safe_str(next_sht) if next_sht else ""})

    title_text = f"DATALOGGER INPUT CIRCUITS - {sheet_index_number}"
    _insert_border_title_generic(
        doc=doc,
        border_template_path=border_template_path,
        template_name="DATALOGGER",
        sig_name=None,
        ahead_sig=None,
        loc=None,
        sht_number=sht_number,
        cont_number=cont_number,
        title_text_override=title_text,
    )

    out_name = f"DATALOGGER_SHT{sht_number}.dxf"
    return out_name, doc, letter_counter, next_boundary_letter, inc_counter


def get_datalogger_start_sheet_number(xlsx_path: str) -> int:
    """Reads DATA LOGGER's starting sheet number from FIELD PG.NO."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(row=r, column=2).value or "").strip().upper()
        if label == "DATA LOGGER":
            value = ws.cell(row=r, column=3).value
            if value is None or str(value).strip() == "":
                raise ValueError("FIELD PG.NO has no Sheet Number for the DATA LOGGER row")
            return int(str(value).strip())
    raise ValueError("Could not find a 'DATA LOGGER' row in FIELD PG.NO")


def get_next_circuit_after_datalogger(xlsx_path: str) -> str:
    """Reads the starting sheet number of whatever circuit follows DATA LOGGER in FIELD PG.NO."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(row=r, column=2).value or "").strip().upper()
        if label == "DATA LOGGER":
            value = ws.cell(row=r + 1, column=3).value
            return str(value).strip() if value is not None else ""
    return ""


def get_datalogger_i3_override(xlsx_path: str):
    """Reads FIELD PG.NO!I3 - the contact override preference (e.g. 'D7,D8'), or None if blank."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["FIELD PG.NO"]
    return _parse_i3_override(ws["I3"].value)


def generate_all_datalogger_sheets(xlsx_path: str, start_sheet_number: int, border_template_path: str, used_contacts: dict = None, log=None):
    """
    Generates all Data Logger sheets, 36 relay entries per sheet.
    used_contacts: optional {(s_name_upper, r_name_upper): set of "LETTERNUM"}
        from build_used_contacts_registry() - only meaningful when called
        from the Full IFC job (individual Data Logger runs have nothing
        else generated yet to check against, so this is empty/None there
        and every relay just gets the top-priority back pair).
    Returns (list of (output_filename, ezdxf.Drawing, sht_number_str), next_sheet_number)
    """
    racks = read_all_relay_racks(xlsx_path)
    entries = _flatten_relay_entries(racks)

    # CONFIRMED: TIMER relays (RELTYP="TIMER", CNTCNF holding a duration
    # label like "10-60 SEC" instead of a normal 8F-8B/4F-4B contact
    # configuration) don't have a contact bank to select from - excluded
    # here rather than crashing when _contact_for_relay hits an
    # unrecognized CNTCNF.
    timer_entries = [e for e in entries if (e.get("RELTYP") or "").strip().upper() == "TIMER"]
    entries = [e for e in entries if e not in timer_entries]
    if timer_entries and log is not None:
        for e in timer_entries:
            log.warning(f"[DATA LOGGER] Timer relay excluded (no contact bank): {e['S_NAME']} {e['R_NAME']}")

    if not entries:
        raise ValueError("No relay entries found in RELAY RACK sheet")

    i3_override = get_datalogger_i3_override(xlsx_path)

    pages = [entries[i:i + RELAYS_PER_SHEET] for i in range(0, len(entries), RELAYS_PER_SHEET)]

    # CONFIRMED: pad the last sheet's remaining empty slots with SPARE entries.
    if pages:
        last_page = pages[-1]
        while len(last_page) < RELAYS_PER_SHEET:
            last_page.append({"S_NAME": "", "R_NAME": "SPARE", "CNTCNF": "8F-8B"})

    results = []
    sheet_num = start_sheet_number
    letter_counter = 0
    boundary_letter = None
    inc_counter = 1
    for page_index, page in enumerate(pages):
        sht = f"{sheet_num:03d}"
        cont = f"{sheet_num + 1:03d}"
        is_last_sheet = page_index == len(pages) - 1
        is_first_sheet = page_index == 0
        next_sht = None if is_last_sheet else f"{sheet_num + 1:03d}"
        previous_sht = None if is_first_sheet else f"{sheet_num - 1:03d}"
        try:
            out_name, doc, letter_counter, boundary_letter, inc_counter = generate_one_datalogger_sheet(
                page, sht, cont, border_template_path, i3_override, is_last_sheet,
                letter_counter, next_sht, is_first_sheet, previous_sht, page_index + 1, boundary_letter,
                used_contacts, log, inc_counter,
            )
        except Exception as e:
            raise ValueError(f"Sheet starting at relay {page[0]['S_NAME']!r}: {e}") from e
        results.append((out_name, doc, sht))
        sheet_num += 1

    return results, sheet_num
